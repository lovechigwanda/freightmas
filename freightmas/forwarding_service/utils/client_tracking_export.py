# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd and contributors
# For license information, please see license.txt

"""Client-facing tracking export helpers (portal Excel workbook)."""

import io

from frappe.utils import getdate
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from freightmas.forwarding_service.utils.client_tracking_view import (
	JOURNEY_SECTION_ORDER,
	build_pdf_job_context,
)

XL_PURPLE = "2C2560"
XL_PURPLE_LIGHT = "4B3F8F"
XL_TEAL = "178A8A"
XL_OVERDUE_FILL = "FCE9E8"
XL_OVERDUE_FONT = "A3241F"

STATUS_FILLS = {
	"In Progress": ("FDEEE0", "9A4A10"),
	"Completed": ("E5F3EA", "186429"),
	"Delayed": ("FCE9E8", "A3241F"),
	"Delivered": ("E5F3EA", "186429"),
}


def sort_job_contexts(contexts):
	status_priority = {"red": 0, "orange": 1, "gray": 1, "green": 2}
	return sorted(
		contexts,
		key=lambda ctx: (
			status_priority.get(ctx["status_key"], 1),
			ctx.get("sort_date") or getdate("2099-12-31"),
		),
	)


def build_job_contexts_from_docs(docs):
	return sort_job_contexts([build_pdf_job_context(doc) for doc in docs])


def _style_header_row(ws, row, ncols):
	for col in range(1, ncols + 1):
		cell = ws.cell(row=row, column=col)
		cell.font = Font(bold=True, color="FFFFFF", size=10.5)
		cell.fill = PatternFill("solid", fgColor=XL_PURPLE_LIGHT)
		cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_title_row(ws, ncols, title):
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
	title_cell = ws.cell(row=1, column=1, value=title)
	title_cell.font = Font(bold=True, size=13, color="FFFFFF")
	title_cell.fill = PatternFill("solid", fgColor=XL_PURPLE)
	title_cell.alignment = Alignment(vertical="center", indent=1)
	ws.row_dimensions[1].height = 26


def _format_date_cell(cell, value):
	if value:
		cell.value = value
		cell.number_format = "DD-MMM-YY"


def _journey_lookup(journey_rows):
	return {row["title"]: row for row in journey_rows}


def _phase_detail_columns():
	columns = []
	for title in JOURNEY_SECTION_ORDER:
		short = title.replace(" / ", "/")
		columns.extend([
			(f"{short} Summary", f"{title}:summary"),
			(f"{short} %", f"{title}:percent"),
			(f"{short} State", f"{title}:state"),
		])
	return columns


def build_client_tracking_workbook(customer_label, job_contexts):
	wb = Workbook()

	# ---- Sheet 1: Shipments ----
	ws = wb.active
	ws.title = "Shipments"
	ws.sheet_properties.tabColor = XL_PURPLE

	shipment_headers = [
		"Reference",
		"Job Ref",
		"BL Number",
		"Route",
		"Direction",
		"Mode",
		"Phase",
		"ETA / ETD",
		"Overdue",
		"Latest Update",
		"Progress %",
		"Status",
		"Cargo",
	]
	ncols = len(shipment_headers)
	_write_title_row(ws, ncols, f"SHIPMENT TRACKING — {customer_label.upper()}")
	for col, label in enumerate(shipment_headers, start=1):
		ws.cell(row=2, column=col, value=label)
	_style_header_row(ws, 2, ncols)

	for row_idx, ctx in enumerate(job_contexts, start=3):
		glance = ctx["glance"]
		values = [
			glance["primary_label"],
			ctx["ref"],
			ctx["bl_number"],
			glance["route"],
			ctx["direction"],
			ctx["mode"],
			glance["phase_label"],
			glance["eta_display"],
			"Yes" if ctx["is_overdue"] else "No",
			glance["headline"],
			ctx["progress_percent"],
			ctx["status_label"],
			ctx["cargo_display"],
		]
		for col, val in enumerate(values, start=1):
			cell = ws.cell(row=row_idx, column=col, value=val)

		phase_cell = ws.cell(row=row_idx, column=7)
		phase_color = str(glance.get("phase_color", "64748b")).lstrip("#")
		phase_cell.font = Font(color=phase_color, bold=True)

		if ctx["is_overdue"]:
			for col in (8, 9):
				cell = ws.cell(row=row_idx, column=col)
				cell.fill = PatternFill("solid", fgColor=XL_OVERDUE_FILL)
				cell.font = Font(color=XL_OVERDUE_FONT, bold=True)

		progress_cell = ws.cell(row=row_idx, column=11)
		progress_cell.alignment = Alignment(horizontal="center")

		status_cell = ws.cell(row=row_idx, column=12)
		fill, text = STATUS_FILLS.get(ctx["status_label"], ("F3F4F6", "5F6368"))
		status_cell.fill = PatternFill("solid", fgColor=fill)
		status_cell.font = Font(bold=True, color=text)
		status_cell.alignment = Alignment(horizontal="center")

	last_shipment_row = max(2 + len(job_contexts), 2)
	if job_contexts:
		ws.conditional_formatting.add(
			f"K3:K{last_shipment_row}",
			DataBarRule(
				start_type="num",
				start_value=0,
				end_type="num",
				end_value=100,
				color="F6C199",
				showValue=True,
			),
		)

	shipment_widths = [16, 16, 15, 22, 10, 14, 18, 16, 9, 34, 11, 13, 22]
	for col, width in enumerate(shipment_widths, start=1):
		ws.column_dimensions[get_column_letter(col)].width = width
	ws.freeze_panes = "C3"
	ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{last_shipment_row}"
	ws.sheet_view.showGridLines = False

	# ---- Sheet 2: Detail ----
	ws_detail = wb.create_sheet("Detail")
	ws_detail.sheet_properties.tabColor = XL_PURPLE_LIGHT

	detail_base_headers = [
		"Job Ref",
		"Reference",
		"BL Number",
		"ETD",
		"ATD",
		"ETA",
		"ATA",
		"Discharge",
		"Completed On",
	]
	phase_columns = _phase_detail_columns()
	detail_headers = detail_base_headers + [label for label, _key in phase_columns]
	detail_ncols = len(detail_headers)
	_write_title_row(ws_detail, detail_ncols, "SHIPMENT DETAIL")
	for col, label in enumerate(detail_headers, start=1):
		ws_detail.cell(row=2, column=col, value=label)
	_style_header_row(ws_detail, 2, detail_ncols)

	for row_idx, ctx in enumerate(job_contexts, start=3):
		glance = ctx["glance"]
		dates = ctx["dates"]
		journey = _journey_lookup(ctx["journey"])
		values = [
			ctx["ref"],
			glance["primary_label"],
			ctx["bl_number"],
			dates.get("etd"),
			dates.get("atd"),
			dates.get("eta"),
			dates.get("ata"),
			dates.get("discharge_date"),
			dates.get("completed_on"),
		]
		for _label, key in phase_columns:
			title, field = key.split(":", 1)
			phase = journey.get(title)
			if not phase:
				values.append(None)
				continue
			if field == "summary":
				values.append(phase["summary"])
			elif field == "percent":
				values.append(phase["progress_percent"])
			else:
				values.append(phase["state"].title() if phase.get("state") else None)

		for col, val in enumerate(values, start=1):
			cell = ws_detail.cell(row=row_idx, column=col, value=val)
			if col in (4, 5, 6, 7, 8, 9) and val:
				cell.number_format = "DD-MMM-YY"

	last_detail_row = max(2 + len(job_contexts), 2)
	detail_widths = [16, 16, 15, 11, 11, 11, 11, 11, 13] + [22, 8, 10] * len(JOURNEY_SECTION_ORDER)
	for col, width in enumerate(detail_widths[:detail_ncols], start=1):
		ws_detail.column_dimensions[get_column_letter(col)].width = width
	ws_detail.freeze_panes = "D3"
	ws_detail.auto_filter.ref = f"A2:{get_column_letter(detail_ncols)}{last_detail_row}"
	ws_detail.sheet_view.showGridLines = False

	# ---- Sheet 3: Containers ----
	ws_containers = wb.create_sheet("Containers")
	ws_containers.sheet_properties.tabColor = XL_TEAL

	container_headers = [
		"Job Ref",
		"Reference",
		"Container",
		"Type",
		"Status",
		"Last Event",
		"Date",
	]
	container_ncols = len(container_headers)
	_write_title_row(ws_containers, container_ncols, "CONTAINERS")
	for col, label in enumerate(container_headers, start=1):
		ws_containers.cell(row=2, column=col, value=label)
	_style_header_row(ws_containers, 2, container_ncols)

	container_row_idx = 3
	for ctx in job_contexts:
		glance = ctx["glance"]
		for container in ctx["containers"]:
			values = [
				ctx["ref"],
				glance["primary_label"],
				container["container_number"],
				container["container_type"],
				container["status"],
				container["last_event"],
				container["last_event_date"],
			]
			for col, val in enumerate(values, start=1):
				ws_containers.cell(row=container_row_idx, column=col, value=val)
			container_row_idx += 1

	last_container_row = max(container_row_idx - 1, 2)
	container_widths = [16, 16, 18, 10, 16, 22, 14]
	for col, width in enumerate(container_widths, start=1):
		ws_containers.column_dimensions[get_column_letter(col)].width = width
	ws_containers.freeze_panes = "C3"
	if container_row_idx > 3:
		ws_containers.auto_filter.ref = f"A2:{get_column_letter(container_ncols)}{last_container_row}"
	ws_containers.sheet_view.showGridLines = False

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)
	return output.getvalue()
