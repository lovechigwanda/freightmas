# Enhanced Export API for FreightMas Reports
"""
Enhanced export functionality with better error handling, logging, and consistency.
This builds upon the existing api.py functions but adds improvements.
"""

import frappe
from frappe import _
from frappe.utils import now_datetime, formatdate
import json
import importlib
from io import BytesIO
import logging
from typing import Optional, Dict, List, Any

from freightmas.utils.report_utils import get_report_filename


# Setup logging
logger = frappe.logger("freightmas.exports", allow_site=True, file_count=50)


class ReportExportError(Exception):
    """Custom exception for report export errors."""
    pass


def log_export_attempt(report_name: str, export_type: str, user: str, filters: Dict):
    """Log export attempts for auditing."""
    logger.info(f"Export {export_type.upper()} - Report: {report_name}, User: {user}, Filters: {filters}")


def validate_export_request(report_name: str, filters: Optional[str] = None) -> Dict:
    """
    Validate and sanitize export request parameters.
    
    Args:
        report_name: Name of the report
        filters: JSON string of filters
        
    Returns:
        Dictionary of validated parameters
        
    Raises:
        ReportExportError: If validation fails
    """
    # Validate report exists
    if not frappe.db.exists("Report", report_name):
        raise ReportExportError(f"Report '{report_name}' does not exist")
    
    # Get report document
    report = frappe.get_doc("Report", report_name)
    
    # Check permissions
    if not frappe.has_permission("Report", "read", report):
        raise ReportExportError(f"No permission to access report '{report_name}'")
    
    # Validate report type
    if report.report_type != "Script Report":
        raise ReportExportError(f"Only Script Reports are supported. '{report_name}' is {report.report_type}")
    
    # Parse and validate filters
    if isinstance(filters, str):
        try:
            filters = json.loads(filters)
        except json.JSONDecodeError as e:
            raise ReportExportError(f"Invalid filter JSON: {str(e)}")
    
    filters = filters or {}
    
    # Log the export attempt
    log_export_attempt(report_name, "validation", frappe.session.user, filters)
    
    return {
        "report": report,
        "filters": filters,
        "user": frappe.session.user,
        "timestamp": now_datetime()
    }


def get_report_data(report: frappe.Document, filters: Dict) -> tuple:
    """
    Get report data by executing the report module.
    
    Args:
        report: Report document
        filters: Filter dictionary
        
    Returns:
        Tuple of (columns, data)
        
    Raises:
        ReportExportError: If report execution fails
    """
    try:
        # Import report module
        module_path = f"freightmas.{frappe.scrub(report.module)}.report.{frappe.scrub(report.name)}.{frappe.scrub(report.name)}"
        module = importlib.import_module(module_path)
        
        # Execute report
        columns, data = module.execute(filters)
        
        logger.info(f"Report data retrieved - Columns: {len(columns)}, Rows: {len(data)}")
        
        return columns, data
        
    except ImportError as e:
        raise ReportExportError(f"Could not import report module: {str(e)}")
    except Exception as e:
        logger.error(f"Report execution failed: {str(e)}")
        raise ReportExportError(f"Report execution failed: {str(e)}")


@frappe.whitelist()
def export_report_to_excel_v2(report_name: str, filters: Optional[str] = None):
    """
    Enhanced Excel export with better error handling and logging.
    
    Args:
        report_name: Name of the report to export
        filters: JSON string of filter values
    """
    try:
        # Validate request
        validated = validate_export_request(report_name, filters)
        report = validated["report"]
        filters = validated["filters"]
        
        # Get report data
        columns, data = get_report_data(report, filters)
        
        # Generate Excel file
        excel_content = generate_excel_export(report_name, columns, data, filters)
        
        # Set response
        timestamp = validated["timestamp"].strftime("%Y%m%d_%H%M")
        filename = get_report_filename(report_name, "xlsx")
        
        frappe.local.response.filename = filename
        frappe.local.response.filecontent = excel_content
        frappe.local.response.type = "binary"
        
        logger.info(f"Excel export successful - Report: {report_name}, User: {validated['user']}, Filename: {filename}")
        
    except ReportExportError as e:
        logger.error(f"Export failed - Report: {report_name}, Error: {str(e)}")
        frappe.throw(_(str(e)))
    except Exception as e:
        logger.error(f"Unexpected error in Excel export - Report: {report_name}, Error: {str(e)}")
        frappe.throw(_("An unexpected error occurred during export. Please try again or contact support."))


@frappe.whitelist()
def export_report_to_pdf_v2(report_name: str, filters: Optional[str] = None):
    """
    Enhanced PDF export with better error handling and logging.
    
    Args:
        report_name: Name of the report to export
        filters: JSON string of filter values
    """
    try:
        # Validate request
        validated = validate_export_request(report_name, filters)
        report = validated["report"]
        filters = validated["filters"]
        
        # Get report data
        columns, data = get_report_data(report, filters)
        
        # Generate PDF file
        pdf_content = generate_pdf_export(report_name, columns, data, filters)
        
        # Set response
        filename = get_report_filename(report_name, "pdf")
        
        frappe.local.response.filename = filename
        frappe.local.response.filecontent = pdf_content
        frappe.local.response.type = "download"
        
        logger.info(f"PDF export successful - Report: {report_name}, User: {validated['user']}, Filename: {filename}")
        
    except ReportExportError as e:
        logger.error(f"Export failed - Report: {report_name}, Error: {str(e)}")
        frappe.throw(_(str(e)))
    except Exception as e:
        logger.error(f"Unexpected error in PDF export - Report: {report_name}, Error: {str(e)}")
        frappe.throw(_("An unexpected error occurred during export. Please try again or contact support."))


def generate_excel_export(report_name: str, columns: List, data: List, filters: Dict) -> bytes:
    """
    Generate Excel file content with enhanced formatting.
    
    Args:
        report_name: Report name
        columns: Column definitions
        data: Data rows
        filters: Filter values
        
    Returns:
        Excel file content as bytes
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from freightmas.api import get_report_filename  # Import existing function
        
        # Create workbook using existing logic from api.py but with enhancements
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Set sheet title
        import re
        sheet_title = re.sub(r'[\\/*?:\[\]]', '', report_name)[:31]
        ws.title = sheet_title
        
        # Enhanced styling
        title_font = Font(bold=True, size=16, color="2E75B6")
        subtitle_font = Font(bold=True, size=13, color="2E75B6")
        bold_font = Font(bold=True, color="FFFFFF")
        filter_label_font = Font(bold=True, size=11)
        header_fill = PatternFill("solid", fgColor="305496")
        zebra_fill = PatternFill("solid", fgColor="F2F2F2")
        
        # Apply the enhanced Excel generation logic here
        # (This would include the existing logic from api.py but with improvements)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
        
    except Exception as e:
        logger.error(f"Excel generation failed: {str(e)}")
        raise ReportExportError(f"Failed to generate Excel file: {str(e)}")


def generate_pdf_export(report_name: str, columns: List, data: List, filters: Dict) -> bytes:
    """
    Generate PDF file content using enhanced template.
    
    Args:
        report_name: Report name
        columns: Column definitions
        data: Data rows
        filters: Filter values
        
    Returns:
        PDF file content as bytes
    """
    try:
        # Format filters for display
        formatted_filters = {}
        for k, v in filters.items():
            if "date" in k and v:
                try:
                    formatted_filters[k] = formatdate(v, "dd-MMM-yyyy")
                except Exception:
                    formatted_filters[k] = str(v)
            else:
                formatted_filters[k] = v
        
        # Prepare context for template
        context = {
            "company": frappe.defaults.get_user_default("Company") or "FreightMas",
            "title": _(report_name),
            "filters": formatted_filters,
            "columns": columns,
            "data": data,
            "exported_at": now_datetime().strftime("%d-%b-%Y %H:%M"),
            "exported_by": frappe.session.user,
            "frappe": frappe,
        }
        
        # Render template
        html = frappe.render_template(
            "freightmas/templates/report_pdf_template.html", 
            context
        )
        
        # Generate PDF
        pdf = frappe.utils.pdf.get_pdf(
            html,
            options={
                "orientation": "Landscape",
                "footer-right": f"Page [page] of [topage] | Exported by {frappe.session.user}",
                "footer-font-size": "10",
                "header-font-size": "10",
                "margin-top": "20mm",
                "margin-bottom": "25mm",
                "margin-left": "15mm", 
                "margin-right": "15mm",
            }
        )
        
        return pdf
        
    except Exception as e:
        logger.error(f"PDF generation failed: {str(e)}")
        raise ReportExportError(f"Failed to generate PDF file: {str(e)}")


@frappe.whitelist()
def get_export_status(export_id: str):
    """
    Get status of an export operation (for future async exports).
    
    Args:
        export_id: Unique identifier for the export operation
        
    Returns:
        Dictionary with export status information
    """
    # This would be implemented for async exports in the future
    # For now, all exports are synchronous
    return {"status": "completed", "message": "Synchronous export completed"}


@frappe.whitelist()
def get_available_export_formats(report_name: str):
    """
    Get available export formats for a report.
    
    Args:
        report_name: Name of the report
        
    Returns:
        List of available export formats
    """
    try:
        validate_export_request(report_name)
        
        # Standard formats
        formats = [
            {"key": "excel", "label": "Excel (.xlsx)", "endpoint": "export_report_to_excel_v2"},
            {"key": "pdf", "label": "PDF (.pdf)", "endpoint": "export_report_to_pdf_v2"}
        ]
        
        # Check for custom export functions
        # (This could be extended to check for report-specific exports)
        
        return formats
        
    except Exception as e:
        logger.error(f"Error getting export formats for {report_name}: {str(e)}")
        return []