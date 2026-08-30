# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd and contributors
# For license information, please see license.txt

"""Shared Statement of Account (Job Linked) export helpers for Desk and Portal."""

from io import BytesIO

import frappe
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.properties import PageSetupProperties

from freightmas.api import get_report_filename
from freightmas.utils.company_branding import company_logo_data_uri, read_company_logo_bytes

XL_PURPLE = "2C2560"
XL_ACCENT = "3E3B92"
XL_GREY = "F2F2F7"
XL_MUTED = "64748B"
XL_OPENING = "8B7FC7"
XL_SUBTITLE = "C7C2E6"
XL_WHITE = "FFFFFF"
XL_BODY = "2C2560"

LEDGER_HEADERS = ["Date", "Type", "Voucher No", "Job ID", "Debit", "Credit", "Balance", "Remarks"]
COL_WIDTHS = [9, 14, 21, 14, 11, 11, 12, 48]
NUM_COLS = len(LEDGER_HEADERS)
LAST_COL = get_column_letter(NUM_COLS)

_FILL_PURPLE = PatternFill("solid", fgColor=XL_PURPLE)
_FILL_GREY = PatternFill("solid", fgColor=XL_GREY)
_FILL_WHITE = PatternFill("solid", fgColor=XL_WHITE)

_FONT_TITLE = Font(bold=True, size=13, color=XL_WHITE)
_FONT_SUBTITLE = Font(bold=False, size=10, color=XL_SUBTITLE)
_FONT_STAT_LABEL = Font(bold=True, size=8, color=XL_SUBTITLE)
_FONT_STAT_VALUE = Font(bold=True, size=11, color=XL_WHITE)
_FONT_META_LABEL = Font(bold=True, size=8, color=XL_MUTED)
_FONT_META_PARTY = Font(bold=True, size=11, color=XL_ACCENT)
_FONT_META_VALUE = Font(bold=True, size=10, color=XL_BODY)
_FONT_HEADER = Font(bold=True, size=9, color=XL_WHITE)
_FONT_BODY = Font(size=9.5, color=XL_BODY)
_FONT_BODY_BOLD = Font(bold=True, size=9.5, color=XL_BODY)
_FONT_OPENING = Font(size=9.5, color=XL_OPENING)
_FONT_CLOSING_BALANCE = Font(bold=True, size=9.5, color=XL_ACCENT)
_FONT_SUMMARY_LABEL = Font(bold=True, size=8, color=XL_MUTED)
_FONT_SUMMARY_VALUE = Font(bold=True, size=13, color=XL_BODY)
_FONT_SUMMARY_ACCENT = Font(bold=True, size=13, color=XL_ACCENT)
_FONT_FOOTER = Font(italic=True, size=8, color=XL_MUTED)
_FONT_LOGO_FALLBACK = Font(bold=True, size=16, color=XL_PURPLE)

_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_ALIGN_FOOTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

_CLOSING_BORDER = Border(bottom=Side(style="medium", color=XL_PURPLE))
_CURRENCY_FORMAT = "#,##0.00"


def _report_module():
	return frappe.get_module(
		"freightmas.freightmas.report.statement_of_account_job_linked.statement_of_account_job_linked"
	)


def _format_display_date(date_str):
	if date_str:
		return frappe.utils.formatdate(date_str, "dd-MMM-yy")
	return ""


def _party_display_name(party_type, party):
	if not party:
		return None
	field = "customer_name" if party_type == "Customer" else "supplier_name"
	return frappe.db.get_value(party_type, party, field) or party


def _report_context(filters, data):
	party_type = filters.get("party_type")
	party = filters.get("party")
	total_debit = sum(row.get("debit", 0) for row in data if row.get("debit"))
	total_credit = sum(row.get("credit", 0) for row in data if row.get("credit"))
	closing_balance = data[-1].get("balance", 0) if data else 0
	company = filters.get("company")
	company_currency = frappe.db.get_value("Company", company, "default_currency") or "USD"
	company_name = frappe.db.get_value("Company", company, "company_name") or company
	statement_date = _format_display_date(frappe.utils.today())

	return {
		"company": company,
		"company_name": company_name,
		"logo": company_logo_data_uri(company),
		"party_type": party_type,
		"party_name": _party_display_name(party_type, party) or party,
		"report_name": "Statement of Account Job Linked",
		"from_date": _format_display_date(filters.get("from_date")),
		"to_date": _format_display_date(filters.get("to_date")),
		"statement_date": statement_date,
		"currency": company_currency,
		"include_draft": filters.get("include_draft_invoices", False),
		"data": data,
		"totals": {
			"debit": total_debit,
			"credit": total_credit,
			"balance": closing_balance,
		},
		"frappe": frappe,
		"today": frappe.utils.today(),
	}


def _fill_range(ws, row_start, row_end, col_start, col_end, fill):
	for row in range(row_start, row_end + 1):
		for col in range(col_start, col_end + 1):
			ws.cell(row=row, column=col).fill = fill


def _style_cell(cell, value=None, font=None, fill=None, alignment=None, border=None, number_format=None):
	if value is not None:
		cell.value = value
	if font:
		cell.font = font
	if fill:
		cell.fill = fill
	if alignment:
		cell.alignment = alignment
	if border:
		cell.border = border
	if number_format:
		cell.number_format = number_format


def _normalize_remarks(text):
	if not text:
		return ""
	return " ".join(str(text).replace("\r", " ").replace("\n", " ").split())


def _disclaimer_text(party_type):
	if party_type == "Customer":
		return (
			"Statement may include proforma invoices — amounts subject to change "
			"until final invoices are issued"
		)
	return (
		"Statement may include draft purchase invoices — amounts subject to change "
		"until final invoices are issued"
	)


def _write_masthead(ws, context):
	_fill_range(ws, 1, 2, 1, NUM_COLS, _FILL_PURPLE)
	ws.merge_cells("A1:A2")
	ws.merge_cells("B1:E1")
	ws.merge_cells("B2:E2")
	ws.merge_cells("F1:G1")
	ws.merge_cells("F2:G2")

	logo_result = read_company_logo_bytes(context["company"])
	logo_cell = ws["A1"]
	logo_cell.fill = _FILL_WHITE
	logo_cell.alignment = _ALIGN_CENTER
	if logo_result:
		content, mime, _filename = logo_result
		if mime in ("image/png", "image/jpeg", "image/gif"):
			img = XLImage(BytesIO(content))
			img.width = 38
			img.height = 38
			ws.add_image(img, "A1")
		else:
			_style_cell(
				logo_cell,
				(context["company_name"] or context["company"])[:1].upper(),
				font=_FONT_LOGO_FALLBACK,
				fill=_FILL_WHITE,
				alignment=_ALIGN_CENTER,
			)
	else:
		_style_cell(
			logo_cell,
			(context["company_name"] or context["company"])[:1].upper(),
			font=_FONT_LOGO_FALLBACK,
			fill=_FILL_WHITE,
			alignment=_ALIGN_CENTER,
		)

	_style_cell(ws["B1"], "STATEMENT OF ACCOUNT", font=_FONT_TITLE, fill=_FILL_PURPLE, alignment=_ALIGN_LEFT)
	_style_cell(
		ws["B2"],
		f"{context['company_name']} · Job Linked",
		font=_FONT_SUBTITLE,
		fill=_FILL_PURPLE,
		alignment=_ALIGN_LEFT,
	)
	_style_cell(ws["F1"], "STATEMENT DATE", font=_FONT_STAT_LABEL, fill=_FILL_PURPLE, alignment=_ALIGN_RIGHT)
	_style_cell(ws["F2"], context["statement_date"], font=_FONT_STAT_VALUE, fill=_FILL_PURPLE, alignment=_ALIGN_RIGHT)
	_style_cell(ws["H1"], "CURRENCY", font=_FONT_STAT_LABEL, fill=_FILL_PURPLE, alignment=_ALIGN_RIGHT)
	_style_cell(ws["H2"], context["currency"], font=_FONT_STAT_VALUE, fill=_FILL_PURPLE, alignment=_ALIGN_RIGHT)

	ws.row_dimensions[1].height = 22
	ws.row_dimensions[2].height = 20


def _write_meta_bar(ws, context):
	_fill_range(ws, 3, 4, 1, NUM_COLS, _FILL_GREY)
	ws.merge_cells("A3:C3")
	ws.merge_cells("A4:C4")
	ws.merge_cells("D3:F3")
	ws.merge_cells("D4:F4")
	ws.merge_cells("G3:H3")
	ws.merge_cells("G4:H4")

	_style_cell(ws["A3"], context["party_type"].upper(), font=_FONT_META_LABEL, fill=_FILL_GREY, alignment=_ALIGN_LEFT)
	_style_cell(ws["A4"], context["party_name"], font=_FONT_META_PARTY, fill=_FILL_GREY, alignment=_ALIGN_LEFT)
	_style_cell(ws["D3"], "STATEMENT PERIOD", font=_FONT_META_LABEL, fill=_FILL_GREY, alignment=_ALIGN_CENTER)
	_style_cell(
		ws["D4"],
		f"{context['from_date']} to {context['to_date']}",
		font=_FONT_META_VALUE,
		fill=_FILL_GREY,
		alignment=_ALIGN_CENTER,
	)
	_style_cell(
		ws["G3"],
		"INCLUDE DRAFT INVOICES",
		font=_FONT_META_LABEL,
		fill=_FILL_GREY,
		alignment=_ALIGN_RIGHT,
	)
	_style_cell(
		ws["G4"],
		"Yes" if context["include_draft"] else "No",
		font=_FONT_META_VALUE,
		fill=_FILL_GREY,
		alignment=_ALIGN_RIGHT,
	)
	ws.row_dimensions[3].height = 14
	ws.row_dimensions[4].height = 18


def _write_ledger_header(ws, row):
	_fill_range(ws, row, row, 1, NUM_COLS, _FILL_PURPLE)
	for col, label in enumerate(LEDGER_HEADERS, start=1):
		alignment = _ALIGN_RIGHT if col >= 5 else _ALIGN_LEFT
		_style_cell(
			ws.cell(row=row, column=col),
			label.upper(),
			font=_FONT_HEADER,
			fill=_FILL_PURPLE,
			alignment=alignment,
		)
	ws.row_dimensions[row].height = 18


def _write_amount_cell(cell, value, font, fill, bold_balance=False):
	if value is not None and value != "":
		cell.value = value
		cell.number_format = _CURRENCY_FORMAT
	_style_cell(
		cell,
		font=font or (_FONT_BODY_BOLD if bold_balance else _FONT_BODY),
		fill=fill,
		alignment=_ALIGN_RIGHT,
	)


def _write_ledger_row(ws, row_num, row_data, body_index, fill=None):
	voucher_type = row_data.get("voucher_type")
	is_opening = voucher_type == "Opening Balance"
	is_closing = voucher_type == "Closing Balance"
	is_data = not is_opening and not is_closing
	zebra = body_index % 2 == 0
	row_fill = _FILL_GREY if is_data and zebra else _FILL_WHITE

	if is_opening:
		values = ["Opening", "", "", "", None, None, row_data.get("balance"), "Opening Balance"]
	elif is_closing:
		values = [
			row_data.get("posting_date") or "",
			"",
			"",
			"",
			None,
			None,
			row_data.get("balance"),
			"Closing Balance",
		]
	else:
		values = [
			row_data.get("posting_date") or "",
			row_data.get("voucher_type") or "",
			row_data.get("voucher_no") or "",
			row_data.get("job_id") or "",
			row_data.get("debit") or None,
			row_data.get("credit") or None,
			row_data.get("balance"),
			_normalize_remarks(row_data.get("remarks")),
		]

	for col, value in enumerate(values, start=1):
		cell = ws.cell(row=row_num, column=col)
		use_fill = fill or row_fill

		if col in (5, 6, 7):
			if is_opening:
				font = _FONT_BODY_BOLD if col == 7 else _FONT_OPENING
			elif is_closing:
				font = _FONT_CLOSING_BALANCE if col == 7 else _FONT_BODY
			else:
				font = _FONT_BODY_BOLD if col == 7 else _FONT_BODY
			_write_amount_cell(cell, value, font, use_fill, bold_balance=(col == 7))
			continue

		if is_opening:
			font = _FONT_OPENING
		elif is_closing:
			font = _FONT_BODY
		else:
			font = _FONT_BODY

		alignment = _ALIGN_WRAP if col == 8 else _ALIGN_LEFT
		_style_cell(cell, value, font=font, fill=use_fill, alignment=alignment)

		if is_closing:
			cell.border = _CLOSING_BORDER

	ws.row_dimensions[row_num].height = 16 if is_data else 18


def _write_summary(ws, row, totals):
	label_row = row
	value_row = row + 1
	for col, label in ((5, "TOTAL DEBITS"), (6, "TOTAL CREDITS"), (7, "CLOSING BALANCE")):
		_style_cell(
			ws.cell(row=label_row, column=col),
			label,
			font=_FONT_SUMMARY_LABEL,
			fill=_FILL_WHITE,
			alignment=_ALIGN_RIGHT,
		)
		value_cell = ws.cell(row=value_row, column=col)
		value = totals.get("debit") if col == 5 else totals.get("credit") if col == 6 else totals.get("balance")
		value_cell.value = value
		value_cell.number_format = _CURRENCY_FORMAT
		value_font = _FONT_SUMMARY_ACCENT if col == 7 else _FONT_SUMMARY_VALUE
		_style_cell(value_cell, font=value_font, fill=_FILL_WHITE, alignment=_ALIGN_RIGHT)
	ws.row_dimensions[label_row].height = 14
	ws.row_dimensions[value_row].height = 20


def _write_footer(ws, row, context):
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
	ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
	_style_cell(
		ws.cell(row=row, column=1),
		_disclaimer_text(context["party_type"]),
		font=_FONT_FOOTER,
		fill=_FILL_WHITE,
		alignment=_ALIGN_FOOTER,
	)
	_style_cell(
		ws.cell(row=row, column=7),
		context["statement_date"],
		font=_FONT_FOOTER,
		fill=_FILL_WHITE,
		alignment=_ALIGN_RIGHT,
	)
	ws.row_dimensions[row].height = 24


def _configure_sheet(ws):
	for idx, width in enumerate(COL_WIDTHS, start=1):
		ws.column_dimensions[get_column_letter(idx)].width = width
	ws.sheet_properties.tabColor = XL_PURPLE
	ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
	ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
	ws.page_setup.fitToWidth = 1
	ws.page_setup.fitToHeight = 0
	ws.print_title_rows = "6:6"


def build_statement_job_linked_pdf(filters):
	_columns, data = _report_module().execute(filters)
	context = _report_context(filters, data)
	html = frappe.render_template("freightmas/templates/statement_of_account_job_linked.html", context)
	pdf = frappe.utils.pdf.get_pdf(
		html,
		{
			"orientation": "Landscape",
			"page-size": "A4",
			"margin-top": "10mm",
			"margin-right": "10mm",
			"margin-bottom": "14mm",
			"margin-left": "10mm",
			"footer-right": "Page [page] of [topage]",
			"footer-font-size": "8",
			"print-media-type": True,
		},
	)
	filename = get_report_filename("Statement_of_Account_Job_Linked", "pdf", filters.get("party"))
	return filename, pdf


def build_statement_job_linked_excel(filters):
	_columns, data = _report_module().execute(filters)
	context = _report_context(filters, data)

	wb = Workbook()
	ws = wb.active
	ws.title = "Statement"

	_write_masthead(ws, context)
	_write_meta_bar(ws, context)

	header_row = 6
	_write_ledger_header(ws, header_row)

	current_row = header_row + 1
	for body_index, row_data in enumerate(data, start=1):
		_write_ledger_row(ws, current_row, row_data, body_index)
		current_row += 1

	current_row += 1
	_write_summary(ws, current_row, context["totals"])
	current_row += 3
	_write_footer(ws, current_row, context)

	_configure_sheet(ws)

	output = BytesIO()
	wb.save(output)
	output.seek(0)
	filename = get_report_filename("Statement_of_Account_Job_Linked", "xlsx", filters.get("party"))
	return filename, output.read()
