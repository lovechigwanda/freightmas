# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd
# Management Accounts Multi-Sheet Excel Export
"""
Generates a comprehensive Management Accounts Excel workbook with multiple
worksheets — combining ERPNext standard financial reports and Freightmas
custom reports into a single downloadable .xlsx file.
"""

import importlib
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt, cstr, now_datetime, getdate

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styles (shared across all sheets)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="305496")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
COVER_FILL = PatternFill("solid", fgColor="1F4E79")
BOLD_WHITE = Font(bold=True, color="FFFFFF")
BOLD_BLACK = Font(bold=True)
TITLE_FONT = Font(bold=True, size=16)
SUBTITLE_FONT = Font(bold=True, size=13)
COVER_TITLE = Font(bold=True, size=22, color="FFFFFF")
COVER_SUBTITLE = Font(bold=True, size=14, color="BDD7EE")
COVER_ITEM = Font(size=11)
FILTER_LABEL_FONT = Font(bold=True)
THIN_BORDER = Border(
	left=Side(style="thin", color="DDDDDD"),
	right=Side(style="thin", color="DDDDDD"),
	top=Side(style="thin", color="DDDDDD"),
	bottom=Side(style="thin", color="DDDDDD"),
)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# ---------------------------------------------------------------------------
# Report Registry
# ---------------------------------------------------------------------------
REPORT_CONFIGS = [
	{
		"key": "pl",
		"sheet_name": "P&L Statement",
		"module": "erpnext.accounts.report.profit_and_loss_statement.profit_and_loss_statement",
		"group": "financial_statement",
		"is_tree": True,
	},
	{
		"key": "pl_cc",
		"sheet_name": "P&L by Cost Center",
		"module": "freightmas.freightmas.report.profit_and_loss_by_cost_center_with_template.profit_and_loss_by_cost_center_with_template",
		"group": "pl_cost_center",
		"is_tree": True,
	},
	{
		"key": "bs",
		"sheet_name": "Balance Sheet",
		"module": "erpnext.accounts.report.balance_sheet.balance_sheet",
		"group": "financial_statement",
		"is_tree": True,
	},
	{
		"key": "cf",
		"sheet_name": "Cash Flow",
		"module": "erpnext.accounts.report.cash_flow.cash_flow",
		"group": "financial_statement",
		"is_tree": True,
	},
	{
		"key": "tb",
		"sheet_name": "Trial Balance",
		"module": "erpnext.accounts.report.trial_balance.trial_balance",
		"group": "trial_balance",
		"is_tree": True,
	},
	{
		"key": "budget",
		"sheet_name": "Budget Variance",
		"module": "erpnext.accounts.report.budget_variance_report.budget_variance_report",
		"group": "budget",
		"is_tree": False,
	},
	{
		"key": "revenue",
		"sheet_name": "Revenue Detail",
		"module": "freightmas.freightmas.report.revenue_detail_report.revenue_detail_report",
		"group": "gl_range",
		"is_tree": False,
	},
	{
		"key": "direct_exp",
		"sheet_name": "Direct Expenses",
		"module": "freightmas.freightmas.report.direct_expenses_detail_report.direct_expenses_detail_report",
		"group": "gl_range",
		"is_tree": False,
	},
	{
		"key": "indirect_exp",
		"sheet_name": "Indirect Expenses",
		"module": "freightmas.freightmas.report.indirect_expenses_detail_report.indirect_expenses_detail_report",
		"group": "gl_range",
		"is_tree": False,
	},
	{
		"key": "cash_bank",
		"sheet_name": "Cash & Bank Balance",
		"module": "freightmas.freightmas.report.cash_and_bank_balance_report.cash_and_bank_balance_report",
		"group": "point_in_time",
		"is_tree": False,
	},
	{
		"key": "debtors",
		"sheet_name": "Debtors Listing",
		"module": "freightmas.freightmas.report.debtors_listing.debtors_listing",
		"group": "point_in_time",
		"is_tree": False,
	},
	{
		"key": "creditors",
		"sheet_name": "Creditors Listing",
		"module": "freightmas.freightmas.report.creditors_listing.creditors_listing",
		"group": "point_in_time",
		"is_tree": False,
	},
	{
		"key": "assets",
		"sheet_name": "Asset Register",
		"module": "erpnext.assets.report.fixed_asset_register.fixed_asset_register",
		"group": "asset_register",
		"is_tree": False,
	},
	# --- Operational Reports ---
	# Add operational report configs here as needed. Each entry needs:
	#   key, sheet_name, module (full Python path to report's execute),
	#   group ("operational" for from_date/to_date, "operational_point_in_time" for as_of_date),
	#   is_tree (False for flat reports)
	{
		"key": "fuel_eff",
		"sheet_name": "Fuel Efficiency",
		"module": "freightmas.trucking_service.report.fuel_efficiency_report.fuel_efficiency_report",
		"group": "operational",
		"is_tree": False,
	},
	{
		"key": "driver_perf",
		"sheet_name": "Driver Performance",
		"module": "freightmas.trucking_service.report.driver_performance_report.driver_performance_report",
		"group": "operational",
		"is_tree": False,
	},
	{
		"key": "fwd_margin",
		"sheet_name": "Fwd Shipment Margin",
		"module": "freightmas.forwarding_service.report.forwarding_shipment_margin_analysis.forwarding_shipment_margin_analysis",
		"group": "operational",
		"is_tree": False,
	},
]
# ---------------------------------------------------------------------------
def _resolve_dates(master):
	"""Derive from_date / to_date from fiscal year selections when needed."""
	if master.get("filter_based_on") == "Fiscal Year":
		from_fy = frappe.get_cached_doc("Fiscal Year", master["from_fiscal_year"])
		to_fy = frappe.get_cached_doc("Fiscal Year", master["to_fiscal_year"])
		return str(from_fy.year_start_date), str(to_fy.year_end_date)
	return master.get("from_date"), master.get("to_date")


def _resolve_fiscal_year(master):
	"""Resolve a fiscal year name from from_date when in Date Range mode."""
	if master.get("filter_based_on") == "Fiscal Year":
		return master["from_fiscal_year"]
	from_date = master.get("from_date")
	if from_date:
		fy = frappe.db.get_value(
			"Fiscal Year",
			{"year_start_date": ["<=", from_date], "year_end_date": [">=", from_date]},
			"name",
		)
		return fy
	return None


def build_report_filters(config, master):
	"""
	Map harmonized master filters to report-specific filter dicts.

	Args:
		config: dict from REPORT_CONFIGS
		master: dict of harmonized master filter values

	Returns:
		frappe._dict of filters suitable for the target report's execute()
	"""
	group = config["group"]
	from_date, to_date = _resolve_dates(master)

	if group == "financial_statement":
		# P&L, Balance Sheet, Cash Flow
		filters = frappe._dict({
			"company": master["company"],
			"filter_based_on": master.get("filter_based_on", "Fiscal Year"),
			"periodicity": master.get("periodicity", "Yearly"),
			"accumulated_values": 0,
			"include_default_book_entries": 1,
		})
		if filters["filter_based_on"] == "Fiscal Year":
			filters["from_fiscal_year"] = master["from_fiscal_year"]
			filters["to_fiscal_year"] = master["to_fiscal_year"]
		else:
			filters["period_start_date"] = from_date
			filters["period_end_date"] = to_date
		if master.get("cost_center"):
			filters["cost_center"] = master["cost_center"]
		if master.get("finance_book"):
			filters["finance_book"] = master["finance_book"]
		if master.get("presentation_currency"):
			filters["presentation_currency"] = master["presentation_currency"]
		return filters

	if group == "pl_cost_center":
		# P&L by Cost Center With Template
		filters = frappe._dict({
			"company": master["company"],
			"filter_based_on": master.get("filter_based_on", "Fiscal Year"),
		})
		if filters["filter_based_on"] == "Fiscal Year":
			filters["from_fiscal_year"] = master["from_fiscal_year"]
			filters["to_fiscal_year"] = master["to_fiscal_year"]
		filters["from_date"] = from_date
		filters["to_date"] = to_date
		if master.get("cost_center"):
			filters["cost_center"] = master["cost_center"]
		if master.get("presentation_currency"):
			filters["presentation_currency"] = master["presentation_currency"]
		return filters

	if group == "trial_balance":
		fiscal_year = _resolve_fiscal_year(master)
		filters = frappe._dict({
			"company": master["company"],
			"fiscal_year": fiscal_year,
			"from_date": from_date,
			"to_date": to_date,
			"with_period_closing_entry_for_opening": 1,
			"with_period_closing_entry_for_current_period": 1,
			"show_net_values": 1,
			"show_group_accounts": 1,
			"include_default_book_entries": 1,
		})
		if master.get("cost_center"):
			filters["cost_center"] = master["cost_center"]
		if master.get("finance_book"):
			filters["finance_book"] = master["finance_book"]
		if master.get("presentation_currency"):
			filters["presentation_currency"] = master["presentation_currency"]
		return filters

	if group == "budget":
		fy_from = master.get("from_fiscal_year") or _resolve_fiscal_year(master)
		fy_to = master.get("to_fiscal_year") or fy_from
		return frappe._dict({
			"company": master["company"],
			"from_fiscal_year": fy_from,
			"to_fiscal_year": fy_to,
			"period": master.get("periodicity", "Yearly"),
			"budget_against": "Cost Center",
		})

	if group == "gl_range":
		# Revenue Detail, Direct Expenses, Indirect Expenses
		filters = frappe._dict({
			"company": master["company"],
			"from_date": from_date,
			"to_date": to_date,
		})
		if master.get("cost_center"):
			filters["cost_center"] = master["cost_center"]
		return filters

	if group == "point_in_time":
		# Cash & Bank Balance, Debtors Listing, Creditors Listing
		return frappe._dict({
			"company": master["company"],
			"as_of_date": to_date,
		})

	if group == "asset_register":
		filters = frappe._dict({
			"company": master["company"],
			"status": "In Location",
		})
		if master.get("cost_center"):
			filters["cost_center"] = master["cost_center"]
		if master.get("finance_book"):
			filters["finance_book"] = master["finance_book"]
		return filters

	if group == "operational":
		# Operational reports use from_date / to_date / company
		from_date, to_date = _resolve_dates(master)
		return frappe._dict({
			"company": master["company"],
			"from_date": from_date,
			"to_date": to_date,
		})

	if group == "operational_point_in_time":
		# Unbilled Revenue Aging uses as_of_date
		_, to_date = _resolve_dates(master)
		return frappe._dict({
			"company": master["company"],
			"as_of_date": to_date,
		})

	return frappe._dict({"company": master["company"]})


# ---------------------------------------------------------------------------
# Sheet Writer
# ---------------------------------------------------------------------------
def _safe_sheet_title(title):
	"""Sanitize sheet title for Excel (max 31 chars, no special chars)."""
	import re

	title = re.sub(r"[\\/*?:\[\]]", "", title)
	return title[:31]


def _write_sheet_header(ws, title, report_filters, ncols):
	"""Write company name, report title, filters, and export timestamp at the
	top of a worksheet. Returns the next available row index."""
	row_idx = 1

	# Company name
	ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max(ncols, 2))
	ws.cell(row=row_idx, column=1, value=frappe.defaults.get_user_default("Company")).font = TITLE_FONT
	row_idx += 1

	# Report title
	ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max(ncols, 2))
	ws.cell(row=row_idx, column=1, value=title).font = SUBTITLE_FONT
	row_idx += 1

	# Filter summary
	if report_filters:
		for label, val in report_filters.items():
			if val and label not in (
				"accumulated_values",
				"include_default_book_entries",
				"with_period_closing_entry_for_opening",
				"with_period_closing_entry_for_current_period",
				"show_net_values",
				"show_group_accounts",
				"show_zero_values",
				"include_closing_entries",
				"budget_against",
				"status",
				"_direct_accounts",
				"_excluded_accounts",
			):
				display_val = val
				if "date" in str(label).lower() and val:
					try:
						display_val = frappe.utils.formatdate(val, "dd-MMM-yy")
					except Exception:
						pass
				ws.cell(row=row_idx, column=1, value=f"{label.replace('_', ' ').title()}:").font = FILTER_LABEL_FONT
				ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=max(ncols, 2))
				ws.cell(row=row_idx, column=2, value=cstr(display_val))
				row_idx += 1

	# Timestamp
	ws.cell(row=row_idx, column=1, value="Exported:").font = FILTER_LABEL_FONT
	ws.cell(row=row_idx, column=2, value=now_datetime().strftime("%d-%b-%Y %H:%M"))
	row_idx += 1

	return row_idx


def write_report_to_sheet(ws, columns, data, title, report_filters, is_tree=False):
	"""
	Write a single report's columns + data to an openpyxl Worksheet.

	Handles:
	- Header/filter rows
	- Column headers with formatting
	- Numeric / date / currency formatting
	- Tree-structured data (indent levels, bold group rows)
	- Zebra striping
	- Auto-fit column widths
	"""
	if not columns:
		ws.cell(row=1, column=1, value=f"No data available for {title}")
		return

	ncols = len(columns)
	row_idx = _write_sheet_header(ws, title, report_filters, ncols)

	# Blank spacer row
	row_idx += 1

	# ---- Column headers ----
	header_row = row_idx
	for col_idx, col in enumerate(columns, start=1):
		label = col.get("label", col.get("fieldname", ""))
		cell = ws.cell(row=header_row, column=col_idx, value=label)
		cell.font = BOLD_WHITE
		cell.alignment = LEFT
		cell.fill = HEADER_FILL
		cell.border = THIN_BORDER
	row_idx += 1

	# Freeze panes below header
	ws.freeze_panes = ws[f"A{header_row + 1}"]

	# ---- Data rows ----
	for i, row in enumerate(data, start=1):
		# Determine if this is a group / parent row in tree reports
		indent = 0
		is_group = False
		if is_tree:
			indent = int(flt(row.get("indent", 0)))
			is_group = bool(row.get("is_group") or row.get("parent_account") == "" or indent == 0)

		fill = ZEBRA_FILL if i % 2 == 0 else None

		for col_idx, col in enumerate(columns, start=1):
			fieldname = col.get("fieldname", "")
			value = row.get(fieldname, "") if isinstance(row, dict) else ""
			cell = ws.cell(row=row_idx, column=col_idx)

			# ---- First column in tree: apply indent ----
			if is_tree and col_idx == 1 and isinstance(value, str):
				value = ("  " * indent) + value

			# ---- Format numbers / currency ----
			fieldtype = col.get("fieldtype", "")
			if fieldtype in ("Int", "Float", "Currency"):
				if isinstance(value, (int, float)):
					cell.value = value
					cell.number_format = "#,##0.00"
				else:
					try:
						cell.value = flt(value)
						cell.number_format = "#,##0.00"
					except Exception:
						cell.value = value
				cell.alignment = RIGHT
			elif "date" in fieldname and value:
				try:
					cell.value = frappe.utils.formatdate(value, "dd-MMM-yy")
				except Exception:
					cell.value = cstr(value)
				cell.alignment = LEFT
			else:
				cell.value = cstr(value) if value is not None else ""
				cell.alignment = LEFT

			cell.border = THIN_BORDER

			# Zebra fill
			if fill:
				cell.fill = fill

			# Bold group / total rows
			if is_group or row.get("is_total_row"):
				cell.font = BOLD_BLACK

		row_idx += 1

	# ---- Auto-fit column widths ----
	for col_idx in range(1, ncols + 1):
		max_length = 0
		for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
			for cell in row:
				try:
					cell_length = len(str(cell.value)) if cell.value else 0
					if cell_length > max_length:
						max_length = cell_length
				except Exception:
					pass
		ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_length + 2, 45))

	# Hide gridlines
	ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Cover Sheet
# ---------------------------------------------------------------------------
def write_cover_sheet(ws, master_filters, sheet_results):
	"""
	Generate a cover / table-of-contents sheet.

	Args:
		ws: openpyxl Worksheet (first sheet)
		master_filters: harmonized filter dict
		sheet_results: list of dicts with keys: sheet_name, success, error
	"""
	ws.sheet_view.showGridLines = False
	ws.column_dimensions["A"].width = 5
	ws.column_dimensions["B"].width = 45
	ws.column_dimensions["C"].width = 20

	row = 1

	# Title block
	ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=3)
	cell = ws.cell(row=row, column=1, value="Management Accounts")
	cell.font = COVER_TITLE
	cell.fill = COVER_FILL
	cell.alignment = Alignment(horizontal="center", vertical="center")
	# Fill merged area
	for r in range(row, row + 2):
		for c in range(1, 4):
			ws.cell(row=r, column=c).fill = COVER_FILL
	row += 3

	# Company
	company = master_filters.get("company", "")
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	cell = ws.cell(row=row, column=1, value=company)
	cell.font = SUBTITLE_FONT
	cell.alignment = CENTER
	row += 2

	# Period
	from_date, to_date = _resolve_dates(master_filters)
	period_text = ""
	if master_filters.get("filter_based_on") == "Fiscal Year":
		fy_from = master_filters.get("from_fiscal_year", "")
		fy_to = master_filters.get("to_fiscal_year", "")
		if fy_from == fy_to:
			period_text = f"Fiscal Year: {fy_from}"
		else:
			period_text = f"Fiscal Years: {fy_from} to {fy_to}"
	else:
		try:
			period_text = f"Period: {frappe.utils.formatdate(from_date, 'dd-MMM-yyyy')} to {frappe.utils.formatdate(to_date, 'dd-MMM-yyyy')}"
		except Exception:
			period_text = f"Period: {from_date} to {to_date}"

	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	cell = ws.cell(row=row, column=1, value=period_text)
	cell.font = Font(size=12, italic=True)
	cell.alignment = CENTER
	row += 2

	# Filter summary
	if master_filters.get("cost_center"):
		ws.cell(row=row, column=1, value="").font = FILTER_LABEL_FONT
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
		ws.cell(row=row, column=1, value="Cost Center:").font = FILTER_LABEL_FONT
		ws.cell(row=row, column=2, value=cstr(master_filters["cost_center"]))
		row += 1
	if master_filters.get("finance_book"):
		ws.cell(row=row, column=1, value="Finance Book:").font = FILTER_LABEL_FONT
		ws.cell(row=row, column=2, value=cstr(master_filters["finance_book"]))
		row += 1
	if master_filters.get("periodicity"):
		ws.cell(row=row, column=1, value="Periodicity:").font = FILTER_LABEL_FONT
		ws.cell(row=row, column=2, value=master_filters["periodicity"])
		row += 1

	row += 1

	# Table of contents
	ws.cell(row=row, column=1, value="").font = BOLD_BLACK
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	tc_cell = ws.cell(row=row, column=1, value="Contents")
	tc_cell.font = Font(bold=True, size=13)
	row += 1

	# Header row
	for c, label in enumerate(["#", "Report", "Status"], start=1):
		cell = ws.cell(row=row, column=c, value=label)
		cell.font = BOLD_WHITE
		cell.fill = HEADER_FILL
		cell.border = THIN_BORDER
	row += 1

	for idx, sr in enumerate(sheet_results, start=1):
		ws.cell(row=row, column=1, value=idx).border = THIN_BORDER
		ws.cell(row=row, column=2, value=sr["sheet_name"]).border = THIN_BORDER
		status_cell = ws.cell(
			row=row,
			column=3,
			value="OK" if sr["success"] else f"Error: {sr.get('error', 'Unknown')}",
		)
		status_cell.border = THIN_BORDER
		if not sr["success"]:
			status_cell.font = Font(color="FF0000")
		else:
			status_cell.font = Font(color="008000")
		row += 1

	row += 2

	# Export metadata
	ws.cell(row=row, column=1, value="Exported by:").font = FILTER_LABEL_FONT
	ws.cell(row=row, column=2, value=frappe.session.user)
	row += 1
	ws.cell(row=row, column=1, value="Export date:").font = FILTER_LABEL_FONT
	ws.cell(row=row, column=2, value=now_datetime().strftime("%d-%b-%Y %H:%M"))


# ---------------------------------------------------------------------------
# Execute a single report safely
# ---------------------------------------------------------------------------
def _execute_report(config, report_filters):
	"""
	Import and execute a report module. Returns (columns, data) or raises.
	Handles reports that return 2-tuple, 4-tuple, or 5-tuple.
	"""
	module = importlib.import_module(config["module"])
	result = module.execute(report_filters)

	# Normalise to (columns, data)
	if isinstance(result, (list, tuple)):
		columns = result[0] if len(result) > 0 else []
		data = result[1] if len(result) > 1 else []
	else:
		columns, data = [], []

	# Some reports return columns as list of dicts or list of strings
	# Normalise string columns to dicts
	normalised = []
	for col in (columns or []):
		if isinstance(col, str):
			parts = col.split(":")
			normalised.append({
				"fieldname": parts[0].lower().replace(" ", "_"),
				"label": parts[0],
				"fieldtype": parts[1] if len(parts) > 1 else "Data",
			})
		elif isinstance(col, dict):
			normalised.append(col)
	columns = normalised

	# Normalise data rows to list of dicts
	normalised_data = []
	for row in (data or []):
		if isinstance(row, dict):
			normalised_data.append(row)
		elif isinstance(row, frappe._dict):
			normalised_data.append(dict(row))
		elif isinstance(row, (list, tuple)):
			d = {}
			for idx, col in enumerate(columns):
				d[col.get("fieldname", f"col_{idx}")] = row[idx] if idx < len(row) else ""
			normalised_data.append(d)

	return columns, normalised_data


# ---------------------------------------------------------------------------
# Main Orchestrator
# ---------------------------------------------------------------------------
def generate_management_accounts(master_filters):
	"""
	Generate the multi-sheet Management Accounts workbook.

	Args:
		master_filters: dict with harmonized filter values

	Returns:
		BytesIO containing the .xlsx file
	"""
	wb = openpyxl.Workbook()

	# We'll write the cover sheet last (needs results from all reports)
	cover_ws = wb.active
	cover_ws.title = _safe_sheet_title("Cover")

	sheet_results = []

	for config in REPORT_CONFIGS:
		sheet_name = _safe_sheet_title(config["sheet_name"])

		try:
			report_filters = build_report_filters(config, master_filters)
			columns, data = _execute_report(config, report_filters)

			ws = wb.create_sheet(title=sheet_name)
			write_report_to_sheet(
				ws,
				columns,
				data,
				config["sheet_name"],
				report_filters,
				is_tree=config.get("is_tree", False),
			)
			sheet_results.append({"sheet_name": config["sheet_name"], "success": True})

		except Exception as e:
			frappe.log_error(
				f"Management Accounts - {config['sheet_name']}: {str(e)}",
				"Management Accounts Export Error",
			)
			# Create a sheet with the error message
			ws = wb.create_sheet(title=sheet_name)
			ws.cell(row=1, column=1, value=config["sheet_name"]).font = SUBTITLE_FONT
			ws.cell(row=2, column=1, value=f"Error generating this report: {str(e)}")
			ws.cell(row=2, column=1).font = Font(color="FF0000")
			sheet_results.append({
				"sheet_name": config["sheet_name"],
				"success": False,
				"error": str(e)[:80],
			})

	# Now write the cover sheet with results
	write_cover_sheet(cover_ws, master_filters, sheet_results)

	# Save
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	return output
