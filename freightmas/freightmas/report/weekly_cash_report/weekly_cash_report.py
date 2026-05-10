# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd and contributors
# For license information, please see license.txt

"""
Weekly Cash Report

Matrix report showing every Cash/Bank account as a column with:
  - Opening balance (Sunday morning position)
  - Receipts section: one row per party/label, amounts in each account
  - Payments section: one row per party/label, amounts in each account
  - Closing balance = Opening + Receipts - Payments
  - Balance as per statement (from Weekly Cash Statement Balance doctype)
  - Difference = Statement - Closing

All amounts are in company base currency (using GL debit/credit fields).
Week runs Sunday (week_ending_date - 6) to Saturday (week_ending_date).
"""

import re
from collections import defaultdict

import frappe
from frappe import _
from frappe.utils import flt, getdate, add_days, formatdate


def execute(filters=None):
	filters = frappe._dict(filters or {})
	_validate(filters)

	week_end = getdate(filters.week_ending_date)
	week_start = add_days(week_end, -6)

	accounts = _get_accounts(filters.company)
	if not accounts:
		frappe.throw(_("No active Cash or Bank accounts found for this company."))

	acc_fieldnames = _build_fieldname_map(accounts)
	columns = _build_columns(accounts, acc_fieldnames, filters)

	account_name_map = {acc.name: acc.account_name for acc in accounts}
	account_names = set(acc_fieldnames.keys())

	opening_balances = _get_opening_balances(filters.company, week_start, account_names)
	entries = _get_weekly_entries(filters.company, week_start, week_end, account_names)

	inter_account = _identify_inter_account(entries, account_names)

	je_vouchers = [
		e.voucher_no for e in entries
		if e.voucher_type == "Journal Entry"
		and not (e.party or "").strip()
		and e.voucher_no not in inter_account
	]
	je_contra = _get_je_contra_accounts(je_vouchers, account_names)

	receipts = defaultdict(lambda: defaultdict(float))
	payments = defaultdict(lambda: defaultdict(float))

	for entry in entries:
		label = _get_label(entry, inter_account, je_contra, account_name_map)
		if flt(entry.debit, 6) > 0:
			receipts[label][entry.account] += flt(entry.debit, 6)
		if flt(entry.credit, 6) > 0:
			payments[label][entry.account] += flt(entry.credit, 6)

	statement_bals = _get_statement_balances(filters.company, filters.week_ending_date, account_names)

	data, summary_data = _assemble_rows(
		accounts, acc_fieldnames, opening_balances, receipts, payments, statement_bals
	)

	report_summary = _get_report_summary(summary_data, filters)

	return columns, data, None, None, report_summary


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _validate(filters):
	if not filters.get("company"):
		frappe.throw(_("Company is required"))
	if not filters.get("week_ending_date"):
		frappe.throw(_("Week Ending Date is required"))
	d = getdate(filters.week_ending_date)
	if d.weekday() != 5:  # 5 = Saturday
		frappe.throw(_("Week Ending Date must be a Saturday"))


# ---------------------------------------------------------------------------
# Accounts
# ---------------------------------------------------------------------------

def _get_accounts(company):
	return frappe.db.sql("""
		SELECT name, account_name, account_type
		FROM `tabAccount`
		WHERE company = %(company)s
		  AND account_type IN ('Bank', 'Cash')
		  AND is_group = 0
		  AND disabled = 0
		ORDER BY account_name ASC
	""", {"company": company}, as_dict=True)


def _build_fieldname_map(accounts):
	"""Return {account.name: safe_fieldname} ensuring uniqueness."""
	seen = {}
	result = {}
	for acc in accounts:
		base = re.sub(r"[^a-z0-9]", "_", acc.account_name.lower()).strip("_") or "acct"
		fn = base
		if fn in seen:
			seen[fn] += 1
			fn = f"{base}_{seen[fn]}"
		else:
			seen[fn] = 0
		result[acc.name] = fn
	return result


def _build_columns(accounts, acc_fieldnames, filters):
	currency = _get_currency(filters)
	cols = [
		{
			"label": _("Description"),
			"fieldname": "description",
			"fieldtype": "Data",
			"width": 230,
		}
	]
	for acc in accounts:
		cols.append({
			"label": acc.account_name,
			"fieldname": acc_fieldnames[acc.name],
			"fieldtype": "Currency",
			"options": currency,
			"width": 145,
		})
	cols.append({
		"label": _("Total"),
		"fieldname": "total",
		"fieldtype": "Currency",
		"options": currency,
		"width": 155,
	})
	return cols


# ---------------------------------------------------------------------------
# GL Data Queries (all amounts in company base currency via debit/credit)
# ---------------------------------------------------------------------------

def _get_opening_balances(company, week_start, account_names):
	"""Balance per account up to (but not including) the week start."""
	rows = frappe.db.sql("""
		SELECT gle.account, SUM(gle.debit) - SUM(gle.credit) AS balance
		FROM `tabGL Entry` gle
		INNER JOIN `tabAccount` acc ON acc.name = gle.account
		WHERE gle.company = %(company)s
		  AND acc.account_type IN ('Bank', 'Cash')
		  AND acc.is_group = 0
		  AND acc.disabled = 0
		  AND gle.posting_date < %(week_start)s
		  AND gle.is_cancelled = 0
		GROUP BY gle.account
	""", {"company": company, "week_start": week_start}, as_dict=True)

	balances = {name: 0.0 for name in account_names}
	for row in rows:
		if row.account in balances:
			balances[row.account] = flt(row.balance, 2)
	return balances


def _get_weekly_entries(company, week_start, week_end, account_names):
	"""All GL entries for Cash/Bank accounts within the week."""
	rows = frappe.db.sql("""
		SELECT
			gle.posting_date,
			gle.account,
			gle.voucher_type,
			gle.voucher_no,
			gle.party_type,
			gle.party,
			gle.debit,
			gle.credit,
			gle.remarks
		FROM `tabGL Entry` gle
		INNER JOIN `tabAccount` acc ON acc.name = gle.account
		WHERE gle.company = %(company)s
		  AND acc.account_type IN ('Bank', 'Cash')
		  AND acc.is_group = 0
		  AND acc.disabled = 0
		  AND gle.posting_date BETWEEN %(week_start)s AND %(week_end)s
		  AND gle.is_cancelled = 0
		ORDER BY gle.posting_date, gle.voucher_no
	""", {"company": company, "week_start": week_start, "week_end": week_end}, as_dict=True)

	# Filter to tracked accounts (guards against any schema drift between queries)
	return [r for r in rows if r.account in account_names]


# ---------------------------------------------------------------------------
# Inter-account transfer detection
# ---------------------------------------------------------------------------

def _identify_inter_account(entries, account_names):
	"""
	A voucher is an inter-account transfer when the GL entries we collected for
	it show BOTH a debit side and a credit side, all within our Cash/Bank account
	set.  This catches Payment Entries of type 'Internal Transfer' and Journal
	Entries that move money between cash/bank accounts.
	"""
	by_voucher = defaultdict(list)
	for entry in entries:
		by_voucher[entry.voucher_no].append(entry)

	result = {}
	for voucher_no, ventries in by_voucher.items():
		debit_accts = {e.account for e in ventries if flt(e.debit, 6) > 0}
		credit_accts = {e.account for e in ventries if flt(e.credit, 6) > 0}

		if debit_accts and credit_accts and debit_accts != credit_accts:
			result[voucher_no] = {
				"from": next(iter(credit_accts)),
				"to": next(iter(debit_accts)),
			}
	return result


# ---------------------------------------------------------------------------
# Journal Entry contra-account lookup
# ---------------------------------------------------------------------------

def _get_je_contra_accounts(je_vouchers, account_names):
	"""
	For Journal Entries with no party and not classified as inter-account,
	find the first non-Cash/Bank account on the same JE to use as the label.
	"""
	if not je_vouchers:
		return {}

	vouchers_tuple = tuple(set(je_vouchers))

	rows = frappe.db.sql("""
		SELECT jea.parent AS voucher_no, acc.account_name
		FROM `tabJournal Entry Account` jea
		INNER JOIN `tabAccount` acc ON acc.name = jea.account
		WHERE jea.parent IN %(vouchers)s
		  AND acc.account_type NOT IN ('Bank', 'Cash')
		ORDER BY jea.parent, jea.idx
	""", {"vouchers": vouchers_tuple}, as_dict=True)

	result = {}
	for row in rows:
		if row.voucher_no not in result:
			result[row.voucher_no] = row.account_name
	return result


# ---------------------------------------------------------------------------
# Row label determination
# ---------------------------------------------------------------------------

def _get_label(entry, inter_account, je_contra, account_name_map):
	"""Determine the display label for a GL entry."""
	# 1. Party always takes precedence (customer, supplier, employee, etc.)
	if (entry.party or "").strip():
		return entry.party.strip()

	voucher_no = entry.voucher_no

	# 2. Inter-account transfer — label is the counterpart account name
	if voucher_no in inter_account:
		transfer = inter_account[voucher_no]
		if flt(entry.debit, 6) > 0:
			# Receiving side: label = source account (the one that was credited)
			return account_name_map.get(transfer["from"], transfer["from"])
		else:
			# Paying side: label = destination account (the one that was debited)
			return account_name_map.get(transfer["to"], transfer["to"])

	# 3. Journal Entry with a known contra account
	if voucher_no in je_contra:
		return je_contra[voucher_no]

	# 4. Fallback: remarks or voucher type
	remarks = (entry.remarks or "").strip()
	if remarks:
		return remarks[:80]
	return entry.voucher_type or _("Unknown")


# ---------------------------------------------------------------------------
# Statement balances
# ---------------------------------------------------------------------------

def _get_statement_balances(company, week_ending_date, account_names):
	"""Retrieve user-entered statement balances for the week from the balances table."""
	rows = frappe.db.sql("""
		SELECT item.account, item.statement_balance
		FROM `tabWeekly Cash Statement Balance Item` item
		INNER JOIN `tabWeekly Cash Statement Balance` parent 
			ON parent.name = item.parent
		WHERE parent.company = %(company)s
		  AND parent.week_ending_date = %(week_ending_date)s
	""", {"company": company, "week_ending_date": week_ending_date}, as_dict=True)

	result = {}
	for row in rows:
		if row.account in account_names:
			result[row.account] = flt(row.statement_balance, 2)
	return result


# ---------------------------------------------------------------------------
# Row assembly
# ---------------------------------------------------------------------------

def _make_row(description, row_type, accounts, acc_fieldnames, amounts=None, total=None):
	"""Build a single data row dict."""
	row = {"description": description, "row_type": row_type}
	grand = 0.0
	for acc in accounts:
		fn = acc_fieldnames[acc.name]
		val = flt((amounts or {}).get(acc.name, 0), 2) if amounts else None
		row[fn] = val
		if val:
			grand += val
	row["total"] = flt(total if total is not None else grand, 2) or None
	return row


def _assemble_rows(accounts, acc_fieldnames, opening_balances, receipts, payments, statement_bals):
	data = []

	# ── Opening Balance ──────────────────────────────────────────────────────
	opening_row = {"description": _("Opening Balance"), "row_type": "opening", "total": 0}
	total_opening = 0.0
	for acc in accounts:
		fn = acc_fieldnames[acc.name]
		val = flt(opening_balances.get(acc.name, 0), 2)
		opening_row[fn] = val
		total_opening += val
	opening_row["total"] = flt(total_opening, 2)
	data.append(opening_row)

	# ── Receipts Section ─────────────────────────────────────────────────────
	data.append({"description": _("Receipts"), "row_type": "section_header", "total": None})

	receipt_totals = defaultdict(float)  # {account.name: total}
	receipt_rows = []
	for label, acct_amounts in receipts.items():
		row = {"description": label, "row_type": "data", "total": 0}
		row_total = 0.0
		for acc in accounts:
			fn = acc_fieldnames[acc.name]
			val = flt(acct_amounts.get(acc.name, 0), 2)
			row[fn] = val if val else None
			row_total += val
			receipt_totals[acc.name] += val
		row["total"] = flt(row_total, 2)
		receipt_rows.append(row)
	receipt_rows.sort(key=lambda r: r["total"] or 0, reverse=True)
	data.extend(receipt_rows)

	total_receipts_row = {"description": _("Total Receipts"), "row_type": "total", "total": 0}
	grand_receipts = 0.0
	for acc in accounts:
		fn = acc_fieldnames[acc.name]
		val = flt(receipt_totals[acc.name], 2)
		total_receipts_row[fn] = val
		grand_receipts += val
	total_receipts_row["total"] = flt(grand_receipts, 2)
	data.append(total_receipts_row)

	data.append({"description": "", "row_type": "spacer", "total": None})

	# ── Payments Section ─────────────────────────────────────────────────────
	data.append({"description": _("Payments"), "row_type": "section_header", "total": None})

	payment_totals = defaultdict(float)
	payment_rows = []
	for label, acct_amounts in payments.items():
		row = {"description": label, "row_type": "data", "total": 0}
		row_total = 0.0
		for acc in accounts:
			fn = acc_fieldnames[acc.name]
			val = flt(acct_amounts.get(acc.name, 0), 2)
			row[fn] = val if val else None
			row_total += val
			payment_totals[acc.name] += val
		row["total"] = flt(row_total, 2)
		payment_rows.append(row)
	payment_rows.sort(key=lambda r: r["total"] or 0, reverse=True)
	data.extend(payment_rows)

	total_payments_row = {"description": _("Total Payments"), "row_type": "total", "total": 0}
	grand_payments = 0.0
	for acc in accounts:
		fn = acc_fieldnames[acc.name]
		val = flt(payment_totals[acc.name], 2)
		total_payments_row[fn] = val
		grand_payments += val
	total_payments_row["total"] = flt(grand_payments, 2)
	data.append(total_payments_row)

	data.append({"description": "", "row_type": "spacer", "total": None})

	# ── Closing Balance ───────────────────────────────────────────────────────
	closing_row = {"description": _("Closing Balance"), "row_type": "closing", "total": 0}
	closing_by_account = {}
	grand_closing = 0.0
	for acc in accounts:
		fn = acc_fieldnames[acc.name]
		closing = flt(
			opening_balances.get(acc.name, 0)
			+ receipt_totals[acc.name]
			- payment_totals[acc.name],
			2,
		)
		closing_row[fn] = closing
		closing_by_account[acc.name] = closing
		grand_closing += closing
	closing_row["total"] = flt(grand_closing, 2)
	data.append(closing_row)

	data.append({"description": "", "row_type": "spacer", "total": None})
	data.append({"description": "", "row_type": "spacer", "total": None})

	# ── Balance as per Statement ──────────────────────────────────────────────
	stmt_row = {"description": _("Balance as per statement"), "row_type": "statement", "total": None}
	grand_stmt = 0.0
	has_any_stmt = False
	for acc in accounts:
		fn = acc_fieldnames[acc.name]
		val = flt(statement_bals.get(acc.name, 0), 2)
		stmt_row[fn] = val if val else None
		if val:
			grand_stmt += val
			has_any_stmt = True
	stmt_row["total"] = flt(grand_stmt, 2) if has_any_stmt else None
	data.append(stmt_row)

	# ── Difference ────────────────────────────────────────────────────────────
	# Only show a difference cell when a statement balance has been entered for that account.
	diff_row = {"description": _("difference"), "row_type": "difference", "total": None}
	grand_diff = 0.0
	has_any_diff = False
	for acc in accounts:
		fn = acc_fieldnames[acc.name]
		stmt_val = flt(statement_bals.get(acc.name, 0), 2)
		closing_val = closing_by_account.get(acc.name, 0)
		if acc.name in statement_bals:
			diff = flt(stmt_val - closing_val, 2)
			diff_row[fn] = diff
			grand_diff += diff
			has_any_diff = True
		else:
			diff_row[fn] = None
	diff_row["total"] = flt(grand_diff, 2) if has_any_diff else None
	data.append(diff_row)

	summary_data = {
		"total_opening": flt(total_opening, 2),
		"grand_receipts": flt(grand_receipts, 2),
		"grand_payments": flt(grand_payments, 2),
		"grand_closing": flt(grand_closing, 2),
	}
	return data, summary_data


# ---------------------------------------------------------------------------
# Report Summary KPI cards
# ---------------------------------------------------------------------------

def _get_report_summary(summary_data, filters):
	closing = summary_data["grand_closing"]
	receipts = summary_data["grand_receipts"]
	payments = summary_data["grand_payments"]
	opening = summary_data["total_opening"]

	week_end = getdate(filters.week_ending_date)
	week_start = add_days(week_end, -6)
	week_label = "{0} — {1}".format(
		formatdate(week_start, "dd MMM yyyy"),
		formatdate(week_end, "dd MMM yyyy"),
	)

	return [
		{
			"value": opening,
			"indicator": "Blue",
			"label": _("Opening Balance"),
			"datatype": "Currency",
			"description": week_label,
		},
		{
			"value": receipts,
			"indicator": "Green",
			"label": _("Total Receipts"),
			"datatype": "Currency",
		},
		{
			"value": payments,
			"indicator": "Orange",
			"label": _("Total Payments"),
			"datatype": "Currency",
		},
		{
			"value": closing,
			"indicator": "Green" if closing >= 0 else "Red",
			"label": _("Closing Balance"),
			"datatype": "Currency",
		},
	]


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def _build_weekly_cash_excel(filters, columns, data):
	"""Build a polished Excel workbook tailored to the matrix row_type system."""
	import io
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
	from openpyxl.utils import get_column_letter
	from frappe.utils import now_datetime

	wb = Workbook()
	ws = wb.active
	ws.title = "Weekly Cash Report"

	# Palette (matches report_export_utils.py)
	_blue    = "305496"
	_grey    = "D6DCE4"
	_zebra   = "F2F2F2"
	_yellow  = "FFF2CC"

	bold_white  = Font(bold=True, color="FFFFFF")
	bold_blue   = Font(bold=True, color=_blue)
	bold_font   = Font(bold=True)
	italic_font = Font(italic=True)
	default_font = Font()
	red_font    = Font(bold=True, color="C00000")
	title_font  = Font(bold=True, size=16)
	sub_font    = Font(bold=True, size=13)
	label_font  = Font(bold=True)

	fill_blue   = PatternFill("solid", fgColor=_blue)
	fill_grey   = PatternFill("solid", fgColor=_grey)
	fill_zebra  = PatternFill("solid", fgColor=_zebra)
	fill_yellow = PatternFill("solid", fgColor=_yellow)
	fill_none   = PatternFill(fill_type=None)

	def _border(top_color=None):
		top = Side(style="thin", color=top_color or "DDDDDD")
		side = Side(style="thin", color="DDDDDD")
		return Border(top=top, left=side, right=side, bottom=side)

	r_align = Alignment(horizontal="right",  vertical="center")
	l_align = Alignment(horizontal="left",   vertical="center")

	SKIP = {"row_type", "indent", "account", "party", "party_type", "voucher_type"}
	xcols = [c for c in columns if c.get("fieldname") not in SKIP]
	ncols = len(xcols)
	cur_fields = {c["fieldname"] for c in xcols if c.get("fieldtype") == "Currency"}

	row_idx = 1

	# ── Company & title ───────────────────────────────────────────────────────
	ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
	c = ws.cell(row=row_idx, column=1, value=filters.get("company", ""))
	c.font = title_font; c.alignment = l_align
	row_idx += 1

	ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
	c = ws.cell(row=row_idx, column=1, value="Weekly Cash Report")
	c.font = sub_font; c.alignment = l_align
	row_idx += 1

	# ── Filters ───────────────────────────────────────────────────────────────
	ws.cell(row=row_idx, column=1, value="Week Ending:").font = label_font
	ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=ncols)
	ws.cell(row=row_idx, column=2, value=formatdate(filters.get("week_ending_date"), "dd MMMM yyyy"))
	row_idx += 1

	ws.cell(row=row_idx, column=1, value="Exported:").font = label_font
	ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=ncols)
	ws.cell(row=row_idx, column=2, value=now_datetime().strftime("%d-%b-%Y %H:%M"))
	row_idx += 1

	# ── Column headers ────────────────────────────────────────────────────────
	header_row = row_idx
	for ci, col in enumerate(xcols, 1):
		c = ws.cell(row=row_idx, column=ci, value=col.get("label", ""))
		c.font = bold_white
		c.fill = fill_blue
		c.border = _border()
		c.alignment = r_align if col.get("fieldtype") == "Currency" else l_align
	row_idx += 1

	# Freeze: lock header row AND description column so you can scroll right
	ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

	# ── Data rows ─────────────────────────────────────────────────────────────
	pending_separator = False
	zebra = 0

	for row_data in data:
		rt = row_data.get("row_type", "data")

		if rt == "spacer":
			pending_separator = True
			zebra = 0
			continue

		# Resolve styles for this row type
		if rt == "section_header":
			r_fill, r_font, do_merge = fill_blue, bold_white, True
			zebra = 0
		elif rt in ("opening", "closing"):
			r_fill, r_font, do_merge = fill_none, bold_font, False
		elif rt == "total":
			r_fill, r_font, do_merge = fill_grey, bold_blue, False
			zebra = 0
		elif rt == "statement":
			r_fill, r_font, do_merge = fill_zebra, italic_font, False
		elif rt == "difference":
			r_fill, r_font, do_merge = fill_yellow, bold_font, False
		else:  # "data"
			zebra += 1
			r_fill = fill_zebra if zebra % 2 == 0 else fill_none
			r_font, do_merge = default_font, False

		top_color = "AAAAAA" if pending_separator else "DDDDDD"
		pending_separator = False

		if do_merge:
			ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
			c = ws.cell(row=row_idx, column=1, value=row_data.get("description", ""))
			c.font = r_font; c.fill = r_fill
			c.alignment = l_align; c.border = _border(top_color)
		else:
			for ci, col in enumerate(xcols, 1):
				fn = col["fieldname"]
				val = row_data.get(fn)
				c = ws.cell(row=row_idx, column=ci)
				c.fill = r_fill
				c.border = _border(top_color if ci == 1 else "DDDDDD")

				if fn in cur_fields:
					c.value = val if isinstance(val, (int, float)) else 0
					c.number_format = "#,##0.00"
					c.alignment = r_align
					# Difference row: red font for non-zero values
					if rt == "difference" and isinstance(val, (int, float)) and val != 0:
						c.font = red_font
					else:
						c.font = r_font
				else:
					c.value = val or ""
					c.alignment = l_align
					c.font = r_font

		row_idx += 1

	# ── Column widths ─────────────────────────────────────────────────────────
	for ci, col in enumerate(xcols, 1):
		ltr = get_column_letter(ci)
		if col.get("fieldname") == "description":
			ws.column_dimensions[ltr].width = 30
		else:
			best = 14
			for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=ci, max_col=ci):
				for cell in row:
					try:
						best = max(best, len(str(cell.value)) if cell.value else 0)
					except Exception:
						pass
			ws.column_dimensions[ltr].width = min(best + 2, 22)

	# ── Presentation ──────────────────────────────────────────────────────────
	ws.sheet_view.showGridLines = False
	ws.sheet_properties.pageSetUpPr.fitToPage = True
	ws.page_setup.orientation = "landscape"
	ws.page_setup.fitToWidth = 1
	ws.page_setup.fitToHeight = 0
	ws.page_setup.paperSize = ws.PAPERSIZE_A4

	out = io.BytesIO()
	wb.save(out)
	out.seek(0)
	return out.getvalue()


@frappe.whitelist()
def export_excel(filters):
	import json
	if isinstance(filters, str):
		filters = json.loads(filters)
	filters = frappe._dict(filters)
	_validate(filters)

	week_end = getdate(filters.week_ending_date)
	week_start = add_days(week_end, -6)

	accounts = _get_accounts(filters.company)
	if not accounts:
		frappe.throw(_("No active Cash or Bank accounts found for this company."))

	acc_fieldnames = _build_fieldname_map(accounts)
	columns = _build_columns(accounts, acc_fieldnames, filters)
	account_name_map = {acc.name: acc.account_name for acc in accounts}
	account_names = set(acc_fieldnames.keys())

	opening_balances = _get_opening_balances(filters.company, week_start, account_names)
	entries = _get_weekly_entries(filters.company, week_start, week_end, account_names)
	inter_account = _identify_inter_account(entries, account_names)

	je_vouchers = [
		e.voucher_no for e in entries
		if e.voucher_type == "Journal Entry"
		and not (e.party or "").strip()
		and e.voucher_no not in inter_account
	]
	je_contra = _get_je_contra_accounts(je_vouchers, account_names)

	receipts = defaultdict(lambda: defaultdict(float))
	payments = defaultdict(lambda: defaultdict(float))
	for entry in entries:
		label = _get_label(entry, inter_account, je_contra, account_name_map)
		if flt(entry.debit, 6) > 0:
			receipts[label][entry.account] += flt(entry.debit, 6)
		if flt(entry.credit, 6) > 0:
			payments[label][entry.account] += flt(entry.credit, 6)

	statement_bals = _get_statement_balances(filters.company, filters.week_ending_date, account_names)
	data, _ = _assemble_rows(accounts, acc_fieldnames, opening_balances, receipts, payments, statement_bals)

	from freightmas.freightmas.report.report_export_utils import send_excel_response
	file_bytes = _build_weekly_cash_excel(filters, columns, data)
	send_excel_response(file_bytes, "Weekly_Cash_Report.xlsx")


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def _get_currency(filters):
	if filters.get("company"):
		return frappe.get_cached_value("Company", filters.company, "default_currency") or "USD"
	return frappe.defaults.get_global_default("currency") or "USD"
