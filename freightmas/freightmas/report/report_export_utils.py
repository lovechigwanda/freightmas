# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd and contributors
# For license information, please see license.txt

"""
Shared export utilities for management account reports.
Provides nicely formatted Excel and PDF exports.
"""

import frappe
from frappe import _
from frappe.utils import flt, formatdate, now_datetime
import io
import re


# ----------------------------------------
# Columns to drop from exports
# ----------------------------------------

# These columns are dropped from BOTH Excel and PDF
ALWAYS_DROP = ["account", "party", "party_type", "voucher_type"]

# Additional columns dropped only from PDF (for page width)
PDF_EXTRA_DROP = []

# Maximum characters for remarks column before truncation
REMARKS_MAX_LEN = 80


def strip_html(text):
    """Remove HTML tags from text."""
    if not text:
        return text or ""
    return re.sub(r"<[^>]+>", "", str(text))


def truncate_remarks(text, max_len=None):
    """Truncate remarks text to a maximum length."""
    if max_len is None:
        max_len = REMARKS_MAX_LEN
    if not text:
        return ""
    text = str(text).strip()
    if len(text) > max_len:
        return text[:max_len] + "..."
    return text


# ============================================================
# EXCEL EXPORT
# ============================================================

def build_excel_file(filters, data, columns, report_title, net_field_label="Net Amount"):
    """
    Build a formatted Excel workbook and return it as bytes.

    Uses the same visual style as the shared ``export_report_to_excel``
    endpoint in ``freightmas.api`` so that all reports look uniform.

    Args:
        filters: report filters dict
        data: list of row dicts from the report
        columns: list of column dicts from the report
        report_title: e.g. "Revenue Detail Report"
        net_field_label: label for the net amount column header
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active

    sheet_title = re.sub(r'[\\/*?:\[\]]', '', report_title)[:31]
    ws.title = sheet_title

    # ---- Styles (matching api.py export_report_to_excel) ----
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=13)
    filter_label_font = Font(bold=True)
    bold_white_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    zebra_fill = PatternFill("solid", fgColor="F2F2F2")
    heading_font = Font(bold=True, size=11, color="305496")
    heading_fill = PatternFill("solid", fgColor="E8ECF1")
    subtotal_font = Font(bold=True, color="305496")
    subtotal_fill = PatternFill("solid", fgColor="D6DCE4")
    grand_total_font = Font(bold=True, color="FFFFFF")
    grand_total_fill = PatternFill("solid", fgColor="305496")
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ---- Filter columns for Excel (drop Account, Party, Party Type, Voucher Type) ----
    excel_columns = [c for c in columns if c.get("fieldname") not in ALWAYS_DROP]
    ncols = len(excel_columns)

    row_idx = 1

    # ---- Company Name (merged) ----
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
    ws.cell(row=row_idx, column=1, value=filters.get("company", "")).font = title_font
    row_idx += 1

    # ---- Report Title (merged) ----
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
    ws.cell(row=row_idx, column=1, value=report_title).font = subtitle_font
    row_idx += 1

    # ---- Filters as label : value rows ----
    display_filters = {
        "from_date": "From Date",
        "to_date": "To Date",
        "fiscal_year": "Fiscal Year",
        "cost_center": "Cost Center",
        "account": "Account",
        "party_type": "Party Type",
        "party": "Party",
        "voucher_type": "Voucher Type",
        "group_by": "Group By",
    }
    for key, label in display_filters.items():
        val = filters.get(key)
        if val:
            # Format date values
            if "date" in key:
                try:
                    val = formatdate(val, "dd-MMM-yy")
                except Exception:
                    pass
            ws.cell(row=row_idx, column=1, value=f"{label}:").font = filter_label_font
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=ncols)
            ws.cell(row=row_idx, column=2, value=val)
            row_idx += 1

    # ---- Exported timestamp ----
    ws.cell(row=row_idx, column=1, value="Exported:").font = filter_label_font
    ws.cell(row=row_idx, column=1).alignment = left_align
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=ncols)
    export_time = now_datetime().strftime("%d-%b-%Y %H:%M")
    ws.cell(row=row_idx, column=2, value=export_time).alignment = left_align
    row_idx += 1

    # ---- Column headers ----
    header_row = row_idx
    for col_idx, col_def in enumerate(excel_columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=strip_html(col_def.get("label", "")))
        cell.font = bold_white_font
        cell.alignment = left_align
        cell.fill = header_fill
        cell.border = border
    row_idx += 1

    # ---- Freeze panes below header ----
    ws.freeze_panes = ws[f"A{header_row + 1}"]

    # ---- Identify currency fields ----
    currency_fields = set()
    for col_def in excel_columns:
        if col_def.get("fieldtype") == "Currency":
            currency_fields.add(col_def["fieldname"])

    # ---- Data rows (zebra striping, headings, subtotals, grand total) ----
    data_row_num = 0

    # Pre-compute text-column boundary for total-row merging
    text_end_col = 0
    for ci, cd in enumerate(excel_columns, 1):
        if cd.get("fieldtype") in ("Currency", "Int", "Float"):
            break
        text_end_col = ci

    for row_data in data:
        if not row_data:
            # Blank separator row
            row_idx += 1
            data_row_num = 0
            continue

        is_heading = row_data.get("is_group_heading", 0)
        is_total = row_data.get("is_group_total", 0)
        is_grand = is_total and "Grand Total" in str(
            row_data.get("remarks", "") or row_data.get("account_name", "")
        )

        # ---- Group heading row (merged across all columns) ----
        if is_heading:
            heading_label = strip_html(
                row_data.get("account_name", "") or row_data.get("remarks", "")
            )
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
            cell = ws.cell(row=row_idx, column=1, value=heading_label)
            cell.font = heading_font
            cell.fill = heading_fill
            cell.alignment = left_align
            cell.border = border
            row_idx += 1
            data_row_num = 0
            continue

        # ---- Subtotal / Grand-total rows (label spans text cols) ----
        if is_total:
            label = strip_html(
                row_data.get("account_name", "") or row_data.get("remarks", "")
            )
            t_font = grand_total_font if is_grand else subtotal_font
            t_fill = grand_total_fill if is_grand else subtotal_fill

            if text_end_col > 1:
                ws.merge_cells(
                    start_row=row_idx, start_column=1,
                    end_row=row_idx, end_column=text_end_col,
                )

            # Label in first cell
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = t_font
            cell.fill = t_fill
            cell.alignment = left_align
            cell.border = border

            # Currency values
            for col_idx, col_def in enumerate(excel_columns, 1):
                if col_idx <= text_end_col:
                    continue
                fieldname = col_def["fieldname"]
                value = row_data.get(fieldname, "")
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = 0
                    cell.number_format = '#,##0.00'
                cell.font = t_font
                cell.fill = t_fill
                cell.alignment = right_align
                cell.border = border

            data_row_num = 0
            row_idx += 1
            continue

        # ---- Normal data row ----
        data_row_num += 1

        for col_idx, col_def in enumerate(excel_columns, 1):
            fieldname = col_def["fieldname"]
            value = row_data.get(fieldname, "")

            # Strip HTML from text values
            if isinstance(value, str):
                value = strip_html(value)

            cell = ws.cell(row=row_idx, column=col_idx)

            # Format numbers and currency
            if col_def.get("fieldtype") in ("Int", "Float", "Currency"):
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = 0
                    cell.number_format = '#,##0.00'
            elif "date" in fieldname and value:
                try:
                    cell.value = formatdate(value, "dd-MMM-yy")
                except Exception:
                    cell.value = value
            else:
                cell.value = value

            cell.border = border

            # Zebra striping and alignment for normal data rows
            if data_row_num % 2 == 0:
                cell.fill = zebra_fill
            cell.alignment = right_align if col_def.get("fieldtype") in ("Int", "Float", "Currency") else left_align

        row_idx += 1

    # ---- Auto-fit column widths (matching api.py approach) ----
    for col_idx in range(1, ncols + 1):
        max_length = 0
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_length + 2, 40))

    # ---- Hide gridlines ----
    ws.sheet_view.showGridLines = False

    # ---- Print setup ----
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# PDF EXPORT
# ============================================================

def build_pdf_file(filters, data, columns, report_title, net_fieldname="net_revenue"):
    """
    Build a formatted PDF and return it as bytes.

    Matches the standard system PDF style (report_pdf_template.html)
    with additional support for group-heading, subtotal and grand-total rows.
    """
    from frappe.utils.pdf import get_pdf

    # Filter columns for PDF (drop technical columns for page width)
    pdf_drop = set(ALWAYS_DROP + PDF_EXTRA_DROP)
    pdf_columns = [c for c in columns if c.get("fieldname") not in pdf_drop]

    company = filters.get("company", "")
    generated = now_datetime().strftime("%d %b %Y %H:%M")

    currency_fields = set()
    for col_def in pdf_columns:
        if col_def.get("fieldtype") in ("Currency", "Int", "Float"):
            currency_fields.add(col_def["fieldname"])

    ncols = len(pdf_columns) + 1  # +1 for row-number column
    text_col_count = sum(1 for c in pdf_columns if c["fieldname"] not in currency_fields)
    curr_cols = [c for c in pdf_columns if c["fieldname"] in currency_fields]

    # ---- Filters ----
    display_filters = {
        "from_date": "From Date",
        "to_date": "To Date",
        "fiscal_year": "Fiscal Year",
        "cost_center": "Cost Center",
        "account": "Account",
        "party_type": "Party Type",
        "party": "Party",
        "voucher_type": "Voucher Type",
        "group_by": "Group By",
    }

    filter_rows = ""
    for key, label in display_filters.items():
        val = filters.get(key)
        if val:
            if "date" in key:
                try:
                    val = formatdate(val, "dd MMM yyyy")
                except Exception:
                    pass
            filter_rows += f"        <tr><td><b>{label}:</b></td><td>{val}</td></tr>\n"

    # ---- Build HTML (matching report_pdf_template.html style) ----
    html = (
        '<!DOCTYPE html>\n<html>\n<head>\n<meta charset="utf-8">\n<style>\n'
        "@page { size: A4 landscape; margin: 20mm 15mm 25mm 15mm; }\n"
        'body { font-family: "Helvetica Neue", Helvetica, Arial, sans-serif; '
        "font-size: 12px; color: #222; }\n"
        ".print-heading { margin-bottom: 10px; }\n"
        ".print-heading-title { font-size: 22px; font-weight: bold; }\n"
        ".report-title { font-size: 15px; font-weight: bold; color: #000; }\n"
        ".filters-table { margin: 10px 0 20px 0; font-size: 11px; }\n"
        ".filters-table td { padding: 2px 8px 2px 0; }\n"
        ".table { width: 100%; border-collapse: collapse; margin-top: 10px; }\n"
        ".table th, .table td { border: 1px solid #bbb; padding: 7px 8px; }\n"
        ".table th { background: #f5f7fa; font-weight: bold; font-size: 13px; text-align: left; }\n"
        ".table td { white-space: nowrap; }\n"
        ".text-right { text-align: right; }\n"
        ".text-left  { text-align: left; }\n"
        ".text-center { text-align: center; }\n"
        "tr.zebra td { background: #f9f9f9; }\n"
        "tr.group-heading td { background: #eef2f7; font-weight: bold; "
        "font-size: 12px; border-bottom: 2px solid #bbb; }\n"
        "tr.subtotal td { font-weight: bold; background: #e8ecf1; border-top: 1px solid #999; }\n"
        "tr.grand-total td { font-weight: bold; background: #d6dce4; "
        "font-size: 12px; border-top: 2px solid #555; }\n"
        "tr.separator td { border-left: none; border-right: none; height: 4px; padding: 0; }\n"
        "thead { display: table-header-group; }\n"
        "tfoot { display: table-row-group; }\n"
        "</style>\n</head>\n<body>\n"
    )

    # Header
    html += (
        f'<div class="print-heading">\n'
        f'  <div class="print-heading-title">{company}</div>\n'
        f'  <div class="report-title">{report_title}</div>\n'
        f"</div>\n"
    )

    # Filters table
    if filter_rows:
        html += f'<table class="filters-table"><tbody>\n{filter_rows}</tbody></table>\n'

    # Exported timestamp
    html += (
        f'<div style="text-align:right; font-size:11px; color:#888; margin-bottom:8px;">'
        f"Exported: {generated}</div>\n"
    )

    # Table header
    html += '<table class="table table-bordered table-striped">\n<thead><tr>\n'
    html += '  <th class="text-center">#</th>\n'
    for col_def in pdf_columns:
        align = "text-right" if col_def["fieldname"] in currency_fields else "text-left"
        html += f'  <th class="{align}">{strip_html(col_def.get("label", ""))}</th>\n'
    html += "</tr></thead>\n<tbody>\n"

    # Table body
    row_num = 0
    alt = False

    for row_data in data:
        if not row_data:
            html += f'<tr class="separator"><td colspan="{ncols}">&nbsp;</td></tr>\n'
            alt = False
            row_num = 0
            continue

        is_heading = row_data.get("is_group_heading", 0)
        is_total = row_data.get("is_group_total", 0)
        is_grand = is_total and "Grand Total" in str(
            row_data.get("remarks", "") or row_data.get("account_name", "")
        )

        # ---- Group heading (spans all columns) ----
        if is_heading:
            label = strip_html(
                row_data.get("account_name", "") or row_data.get("remarks", "")
            )
            html += (
                f'<tr class="group-heading">'
                f'<td colspan="{ncols}" class="text-left">{label}</td></tr>\n'
            )
            alt = False
            row_num = 0
            continue

        # ---- Subtotal / Grand total (label spans text cols) ----
        if is_total:
            row_class = "grand-total" if is_grand else "subtotal"
            label = strip_html(
                row_data.get("account_name", "") or row_data.get("remarks", "")
            )
            span = text_col_count + 1  # +1 for # column
            html += f'<tr class="{row_class}">'
            html += f'<td colspan="{span}" class="text-left">{label}</td>'
            for col_def in curr_cols:
                val = row_data.get(col_def["fieldname"], "")
                try:
                    num = flt(val)
                    if num < 0:
                        formatted = "({:,.2f})".format(abs(num))
                    else:
                        formatted = "{:,.2f}".format(num)
                except (ValueError, TypeError):
                    formatted = val
                html += f'<td class="text-right">{formatted}</td>'
            html += "</tr>\n"
            continue

        # ---- Normal data row ----
        row_num += 1
        row_class = "zebra" if alt else ""
        html += f'<tr class="{row_class}"><td class="text-center">{row_num}</td>'

        for col_def in pdf_columns:
            fieldname = col_def["fieldname"]
            value = row_data.get(fieldname, "")
            if isinstance(value, str):
                value = strip_html(value)

            if fieldname in currency_fields and value is not None and value != "":
                try:
                    num = flt(value)
                    if num == 0:
                        value = ""
                    elif num < 0:
                        value = "({:,.2f})".format(abs(num))
                    else:
                        value = "{:,.2f}".format(num)
                except (ValueError, TypeError):
                    pass
                html += f'<td class="text-right">{value}</td>'
            elif col_def.get("fieldtype") == "Date" and value:
                try:
                    value = formatdate(value, "dd MMM yyyy")
                except Exception:
                    pass
                html += f'<td class="text-left">{value}</td>'
            else:
                html += f'<td class="text-left">{value}</td>'

        html += "</tr>\n"
        alt = not alt

    html += "</tbody>\n</table>\n</body>\n</html>"

    pdf_options = {
        "orientation": "Landscape",
        "page-size": "A4",
        "margin-top": "20mm",
        "margin-bottom": "25mm",
        "margin-left": "15mm",
        "margin-right": "15mm",
        "encoding": "UTF-8",
        "no-outline": None,
        "footer-right": "Page [page] of [topage]",
        "footer-font-size": "10",
    }

    return get_pdf(html, options=pdf_options)


# ============================================================
# Frappe File Response Helpers
# ============================================================

def send_excel_response(file_bytes, filename):
    """Send Excel file as a download response."""
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = file_bytes
    frappe.local.response.type = "download"
    frappe.local.response.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def send_pdf_response(file_bytes, filename):
    """Send PDF file as a download response."""
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = file_bytes
    frappe.local.response.type = "download"
    frappe.local.response.content_type = "application/pdf"
