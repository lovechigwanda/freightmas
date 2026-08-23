# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd and contributors
# For license information, please see license.txt

"""API layer backing the Shipment Dashboard (Vue SPA).

Mostly read-only - every endpoint here reads from Forwarding Job and its
existing child tables / linked Sales & Purchase Invoices - plus a handful of
tracking-report "email this to a recipient" actions that send mail via
frappe.sendmail without writing to any doctype. Kept entirely separate from
the Forwarding Job controller so it can evolve independently of the form's
own whitelisted methods.
"""

import re
import json
from collections import Counter
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate, nowdate, add_days, add_months, get_first_day, get_last_day, formatdate, get_formatted_email

from freightmas.utils.forwarding_dnd_calculator import days_remaining_to_lfd

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from freightmas.utils.permissions import check_freightmas_role, check_doc_read_permission
from freightmas.forwarding_service.utils.operational_phase import (
	OPERATIONAL_PHASES,
	get_overview_bucket_phases,
	get_phase_label,
	build_overview_phase_pipeline,
)
from freightmas.forwarding_service.utils.milestone_progress import forwarding_milestone_progress_map

NOT_ACTIVE_STATUSES = ["Completed", "Closed", "Cancelled"]

MILESTONE_TABLE_FIELDS = [
	"road_freight_milestones",
	"port_clearance_milestones",
	"border_clearance_milestones",
	"warehouse_milestones",
]


def _get_at_risk_containers():
	"""
	Containers whose DND and/or Storage clock is still running (not yet gated-out / returned)
	and is within the configured warning window before its Last Free Day. Containers that have
	already passed their Last Free Day are excluded here - those already show up in the
	"Breaching Free Days" list once accrued cost is recalculated.
	"""
	warning_days = cint(frappe.db.get_single_value("FreightMas Settings", "dnd_storage_warning_days") or 3)

	rows = frappe.db.sql(
		"""
		SELECT
			d.parent, d.container_number, d.container_type,
			d.discharge_date, d.gate_out_date, d.empty_return_date, d.to_be_returned,
			d.dnd_free_days, d.storage_free_days,
			fj.customer, fj.bl_number
		FROM `tabForwarding DND Storage Detail` d
		INNER JOIN `tabForwarding Job` fj ON fj.name = d.parent
		WHERE d.parenttype = 'Forwarding Job' AND fj.docstatus < 2 AND fj.status NOT IN %(statuses)s
		  AND d.discharge_date IS NOT NULL
		""",
		{"statuses": NOT_ACTIVE_STATUSES},
		as_dict=True,
	)

	at_risk = []
	for row in rows:
		dnd_end_date = row.empty_return_date if row.to_be_returned else row.gate_out_date
		candidates = []  # (clock_label, days_remaining, last_free_day)

		if not dnd_end_date:
			dnd_remaining = days_remaining_to_lfd(row.discharge_date, row.dnd_free_days)
			if dnd_remaining is not None and 0 <= dnd_remaining <= warning_days:
				lfd = add_days(getdate(row.discharge_date), cint(row.dnd_free_days))
				candidates.append(("DND", dnd_remaining, lfd))

		if not row.gate_out_date:
			storage_remaining = days_remaining_to_lfd(row.discharge_date, row.storage_free_days)
			if storage_remaining is not None and 0 <= storage_remaining <= warning_days:
				lfd = add_days(getdate(row.discharge_date), cint(row.storage_free_days))
				candidates.append(("Storage", storage_remaining, lfd))

		if not candidates:
			continue

		candidates.sort(key=lambda c: c[1])
		most_urgent = candidates[0]

		at_risk.append({
			"job": row.parent,
			"customer": row.customer,
			"bl_number": row.bl_number,
			"container_number": row.container_number,
			"container_type": row.container_type,
			"clock": " & ".join(c[0] for c in candidates),
			"days_remaining": most_urgent[1],
			"last_free_day": most_urgent[2],
		})

	at_risk.sort(key=lambda r: r["days_remaining"])
	return at_risk


# ============================================================
# BRANDING
# ============================================================

@frappe.whitelist()
def get_branding():
	check_freightmas_role()

	company = frappe.defaults.get_global_default("company") or frappe.defaults.get_user_default("Company")
	if not company:
		return {"company": None, "company_name": "FreightMas", "logo": None}

	info = frappe.db.get_value(
		"Company",
		company,
		["company_name", "company_logo", "phone_no", "email", "tax_id", "default_currency"],
		as_dict=True,
	) or {}

	address_line = ""
	address = frappe.get_all(
		"Address",
		filters={"link_doctype": "Company", "link_name": company},
		fields=["address_line1", "address_line2", "city", "country"],
		limit=1,
	)
	if address:
		a = address[0]
		parts = [p for p in [a.address_line1, a.address_line2, a.city, a.country] if p]
		address_line = ", ".join(parts)

	return {
		"company": company,
		"company_name": info.get("company_name") or company,
		"logo": info.get("company_logo"),
		"phone": info.get("phone_no"),
		"email": info.get("email"),
		"tax_id": info.get("tax_id"),
		"address": address_line,
		"currency": info.get("default_currency"),
	}


# ============================================================
# OVERVIEW
# ============================================================

def _active_status_clause(prefix="", active_only=False):
	col = f"{prefix}status" if prefix else "status"
	if active_only:
		return f"AND {col} NOT IN %(statuses)s"
	return ""


def _build_jobs_summary(active_only=False):
	"""Jobs-by-status, cargo totals, and direction breakdown for the overview widget."""
	params = {"statuses": NOT_ACTIVE_STATUSES} if active_only else {}
	status_clause = _active_status_clause(active_only=active_only)
	fj_status_clause = _active_status_clause(prefix="fj.", active_only=active_only)

	by_status = frappe.db.sql(
		f"""
		SELECT status, COUNT(*) AS count
		FROM `tabForwarding Job`
		WHERE docstatus < 2 {status_clause}
		GROUP BY status
		ORDER BY FIELD(status, 'Draft', 'In Progress', 'Delivered', 'Completed', 'Closed', 'Cancelled')
		""",
		params,
		as_dict=True,
	)

	cargo_totals = frappe.db.sql(
		f"""
		SELECT
			COALESCE(SUM(
				CASE WHEN cpd.cargo_type = 'Containerised'
				THEN COALESCE(cpd.cargo_quantity, 0) ELSE 0 END
			), 0) AS containers,
			COALESCE(SUM(
				CASE WHEN cpd.cargo_type = 'Packages'
				THEN COALESCE(cpd.cargo_quantity, 0) ELSE 0 END
			), 0) AS parcels
		FROM `tabForwarding Job` fj
		LEFT JOIN `tabCargo Parcel Details` cpd
			ON cpd.parent = fj.name AND cpd.parenttype = 'Forwarding Job'
		WHERE fj.docstatus < 2 {fj_status_clause}
		""",
		params,
		as_dict=True,
	)[0]

	by_direction = frappe.db.sql(
		f"""
		SELECT
			fj.direction,
			COUNT(DISTINCT fj.name) AS jobs,
			COALESCE(SUM(
				CASE WHEN cpd.cargo_type = 'Containerised'
				THEN COALESCE(cpd.cargo_quantity, 0) ELSE 0 END
			), 0) AS containers,
			COALESCE(SUM(
				CASE WHEN cpd.cargo_type = 'Packages'
				THEN COALESCE(cpd.cargo_quantity, 0) ELSE 0 END
			), 0) AS parcels
		FROM `tabForwarding Job` fj
		LEFT JOIN `tabCargo Parcel Details` cpd
			ON cpd.parent = fj.name AND cpd.parenttype = 'Forwarding Job'
		WHERE fj.docstatus < 2 {fj_status_clause}
		GROUP BY fj.direction
		ORDER BY FIELD(fj.direction, 'Import', 'Export', 'Local', 'Transit'), jobs DESC
		""",
		params,
		as_dict=True,
	)

	return {
		"by_status": by_status,
		"containers": cint(cargo_totals.containers),
		"parcels": cint(cargo_totals.parcels),
		"by_direction": by_direction,
	}


@frappe.whitelist()
def get_overview():
	check_freightmas_role()

	phase_pipeline = build_overview_phase_pipeline()

	jobs_summary = {
		"active": _build_jobs_summary(active_only=True),
		"all": _build_jobs_summary(active_only=False),
	}
	jobs_by_status = jobs_summary["all"]["by_status"]

	monthly_trend = _monthly_revenue_margin_trend(months=6)

	top_customers = frappe.db.sql(
		"""
		SELECT customer, COUNT(*) AS job_count
		FROM `tabForwarding Job`
		WHERE docstatus < 2 AND status NOT IN %(statuses)s
		GROUP BY customer
		ORDER BY job_count DESC
		LIMIT 5
		""",
		{"statuses": NOT_ACTIVE_STATUSES},
		as_dict=True,
	)

	top_corridors = frappe.db.sql(
		"""
		SELECT port_of_loading, port_of_discharge, COUNT(*) AS job_count
		FROM `tabForwarding Job`
		WHERE docstatus < 2 AND status NOT IN %(statuses)s
		  AND port_of_loading IS NOT NULL AND port_of_discharge IS NOT NULL
		GROUP BY port_of_loading, port_of_discharge
		ORDER BY job_count DESC
		LIMIT 5
		""",
		{"statuses": NOT_ACTIVE_STATUSES},
		as_dict=True,
	)

	blockers = frappe.get_all(
		"Forwarding Job",
		filters={
			"docstatus": ["<", 2],
			"status": ["not in", NOT_ACTIVE_STATUSES],
			"current_comment": ["is", "set"],
		},
		fields=["name", "customer", "status", "current_comment", "last_updated_on"],
		order_by="last_updated_on desc",
		limit_page_length=8,
	)

	return {
		"phase_pipeline": phase_pipeline,
		"jobs_by_status": jobs_by_status,
		"jobs_summary": jobs_summary,
		"monthly_trend": monthly_trend,
		"top_customers": top_customers,
		"top_corridors": top_corridors,
		"recent_blockers": blockers,
	}


def _monthly_revenue_margin_trend(months=6):
	current = nowdate()
	buckets = []
	for i in range(months - 1, -1, -1):
		month_date = add_months(current, -i)
		buckets.append({
			"label": formatdate(get_first_day(month_date), "MMM YYYY"),
			"start": get_first_day(month_date),
			"end": get_last_day(month_date),
		})

	rows = frappe.db.sql(
		"""
		SELECT
			DATE_FORMAT(fj.revenue_recognised_on, '%%Y-%%m') AS month_key,
			COUNT(DISTINCT fj.name) AS shipment_count,
			SUM(COALESCE(si_agg.total_si, 0)) AS revenue,
			SUM(COALESCE(pi_agg.total_pi, 0)) AS cost
		FROM `tabForwarding Job` fj
		LEFT JOIN (
			SELECT forwarding_job_reference, SUM(base_grand_total) AS total_si
			FROM `tabSales Invoice`
			WHERE docstatus = 1 AND forwarding_job_reference IS NOT NULL
			GROUP BY forwarding_job_reference
		) si_agg ON si_agg.forwarding_job_reference = fj.name
		LEFT JOIN (
			SELECT forwarding_job_reference, SUM(base_grand_total) AS total_pi
			FROM `tabPurchase Invoice`
			WHERE docstatus = 1 AND forwarding_job_reference IS NOT NULL
			GROUP BY forwarding_job_reference
		) pi_agg ON pi_agg.forwarding_job_reference = fj.name
		WHERE fj.docstatus = 1 AND fj.revenue_recognised_on IS NOT NULL
		  AND fj.revenue_recognised_on BETWEEN %(start)s AND %(end)s
		GROUP BY month_key
		""",
		{"start": buckets[0]["start"], "end": buckets[-1]["end"]},
		as_dict=True,
	)
	by_month = {r.month_key: r for r in rows}

	trend = []
	for b in buckets:
		key = b["start"].strftime("%Y-%m") if hasattr(b["start"], "strftime") else str(b["start"])[:7]
		row = by_month.get(key)
		revenue = flt(row.revenue) if row else 0
		cost = flt(row.cost) if row else 0
		trend.append({
			"period": b["label"],
			"shipment_count": row.shipment_count if row else 0,
			"revenue": flt(revenue, 2),
			"cost": flt(cost, 2),
			"margin": flt(revenue - cost, 2),
		})
	return trend


# ============================================================
# SHIPMENTS (Forwarding Jobs)
# ============================================================

def _operational_phase_label(phase):
	return get_phase_label(phase)


def _resolve_operational_phase_filter(operational_phase=None, operational_phases=None, overview_bucket=None):
	"""Return a frappe filter value for operational_phase, or None."""
	bucket_phases = get_overview_bucket_phases(overview_bucket) if overview_bucket else []
	if bucket_phases:
		return ["in", bucket_phases] if len(bucket_phases) > 1 else bucket_phases[0]

	phases = []
	if operational_phases:
		if isinstance(operational_phases, str):
			raw = operational_phases.strip()
			if raw.startswith("["):
				try:
					parsed = json.loads(raw)
					phases = [phase for phase in parsed if phase]
				except (json.JSONDecodeError, TypeError):
					phases = [phase.strip() for phase in raw.split(",") if phase.strip()]
			else:
				phases = [phase.strip() for phase in raw.split(",") if phase.strip()]
		elif isinstance(operational_phases, (list, tuple)):
			phases = [phase for phase in operational_phases if phase]

	if not phases and operational_phase:
		phases = [operational_phase]

	if not phases:
		return None
	if len(phases) == 1:
		return phases[0]
	return ["in", phases]


@frappe.whitelist()
def get_operational_phases():
	check_freightmas_role()
	return [{"value": phase, "label": get_phase_label(phase)} for phase in OPERATIONAL_PHASES]


@frappe.whitelist()
def get_jobs(customer=None, status=None, direction=None, operational_phase=None, operational_phases=None, overview_bucket=None, search=None, limit_start=0, limit_page_length=20):
	check_freightmas_role()

	filters = {"docstatus": ["<", 2]}
	if customer:
		filters["customer"] = customer
	if status:
		filters["status"] = status
	if direction:
		filters["direction"] = direction

	phase_filter = _resolve_operational_phase_filter(
		operational_phase=operational_phase,
		operational_phases=operational_phases,
		overview_bucket=overview_bucket,
	)
	if phase_filter is not None:
		filters["operational_phase"] = phase_filter

	or_filters = None
	if search:
		or_filters = [
			["name", "like", f"%{search}%"],
			["customer_reference", "like", f"%{search}%"],
			["bl_number", "like", f"%{search}%"],
			["consignee", "like", f"%{search}%"],
		]

	fields = [
		"name", "customer", "customer_reference", "direction", "shipment_mode",
		"status", "operational_phase", "operational_substage",
		"port_of_loading", "port_of_discharge", "destination",
		"vessel_flight_no", "bl_number", "cargo_count", "eta", "ata", "etd", "atd",
		"discharge_date", "current_comment", "last_updated_on",
	]

	jobs = frappe.get_list(
		"Forwarding Job",
		filters=filters,
		or_filters=or_filters,
		fields=fields,
		order_by="modified desc",
		limit_start=frappe.utils.cint(limit_start),
		limit_page_length=frappe.utils.cint(limit_page_length),
	)

	total_count = frappe.db.count("Forwarding Job", filters=filters)

	progress_map = forwarding_milestone_progress_map([j.name for j in jobs])
	today = getdate(nowdate())
	for j in jobs:
		j["milestone_percent"] = progress_map.get(j.name, 0)
		j["operational_phase_label"] = _operational_phase_label(j.get("operational_phase"))
		j["is_overdue"] = bool(
			(j.direction == "Import" and j.eta and getdate(j.eta) < today and not j.ata)
			or (j.direction == "Export" and j.etd and getdate(j.etd) < today and not j.atd)
		)

	return {"jobs": jobs, "total_count": total_count}


def milestone_stage_rollup(milestones):
	"""Roll a service's milestone list up into its named stages, in order, for
	the summarized (stage-level) client report view. Each milestone dict must
	carry `stage` and `stage_sequence` (snapshotted on Job Milestone Progress).

	Returns [{name, done, total, pct, is_current, missing}] where the *current
	stage* is the first stage (by stage_sequence) that isn't fully complete; if
	every stage is complete none is flagged current. `missing` is the labels of
	that stage's incomplete milestones, in original order - callers decide
	whether to display it (e.g. only for the first stage). Milestones with no
	stage are collected into a trailing "Other" bucket so nothing is silently
	dropped. Callers should first check `has_milestone_stages()` and only use
	this when stages are actually configured for the service."""
	buckets = {}
	order = {}
	for m in milestones:
		key = m.get("stage") or None
		seq = m.get("stage_sequence") or 0
		bucket = buckets.setdefault(key, {"done": 0, "total": 0, "missing": []})
		bucket["total"] += 1
		if m.get("is_completed"):
			bucket["done"] += 1
		else:
			bucket["missing"].append(m.get("label"))
		order[key] = min(order[key], seq) if key in order else seq

	# Order by stage_sequence; the unstaged "Other" bucket always sorts last.
	ordered = sorted(
		buckets.items(),
		key=lambda item: (1, 0) if item[0] is None else (0, order.get(item[0], 0)),
	)

	stages = []
	current_taken = False
	for key, bucket in ordered:
		pct = round(bucket["done"] / bucket["total"] * 100) if bucket["total"] else 0
		is_current = (not current_taken) and bucket["done"] < bucket["total"]
		if is_current:
			current_taken = True
		stages.append({
			"name": key or "Other",
			"done": bucket["done"],
			"total": bucket["total"],
			"pct": pct,
			"is_current": is_current,
			"missing": bucket["missing"],
		})
	return stages


def has_milestone_stages(milestones):
	"""True when at least one milestone in the list carries a stage - i.e. the
	service can be summarized by stage. When False, callers fall back to the
	full milestone checklist even in summary mode."""
	return any(m.get("stage") for m in milestones)


def _build_job_milestone_stages(doc):
	"""Milestone groups (Road Freight / Port Clearance / Border Clearance /
	Warehouse) for a Forwarding Job - only services actually required on the
	job, and only groups that carry milestone rows. Shared by get_job_detail
	(the Vue job drawer) and _build_job_dossier_context (the PDF dossier), so
	both render the exact same milestone shape without re-deriving it.

	Each group also carries a stage rollup (`stages` + `has_stages`) for the
	summarized client-report view - see milestone_stage_rollup()."""
	section_labels = {
		"road_freight_milestones": "Road Freight",
		"port_clearance_milestones": "Port Clearance",
		"border_clearance_milestones": "Border Clearance",
		"warehouse_milestones": "Warehouse",
	}
	requires_map = {
		"road_freight_milestones": True,
		"port_clearance_milestones": doc.requires_port_clearance,
		"border_clearance_milestones": doc.requires_border_clearance,
		"warehouse_milestones": doc.requires_warehousing,
	}
	stages = []
	for fieldname, label in section_labels.items():
		if not requires_map.get(fieldname):
			continue
		rows = doc.get(fieldname) or []
		if not rows:
			continue
		milestones = [
			{
				"label": r.milestone_label,
				"is_completed": bool(r.is_completed),
				"completed_on": r.completed_on,
				"remarks": r.remarks,
				"stage": r.get("stage"),
				"stage_sequence": r.get("stage_sequence") or 0,
			}
			for r in rows
		]
		stages.append({
			"group": label,
			"milestones": milestones,
			"has_stages": has_milestone_stages(milestones),
			"stages": milestone_stage_rollup(milestones),
		})
	return stages


def _build_job_cargo_list(doc):
	"""Cargo/container rows for a Forwarding Job, shaped for the dashboard.
	Shared by get_job_detail (Vue drawer) and _build_job_dossier_context."""
	return [
		{
			"name": r.name,
			"container_number": r.container_number or r.cargo_item_description,
			"container_type": r.container_type,
			"cargo_type": r.cargo_type,
			"to_be_returned": bool(r.to_be_returned),
			"return_by_date": r.return_by_date,
			"is_truck_required": bool(r.is_truck_required),
			"is_booked": bool(r.is_booked),
			"is_loaded": bool(r.is_loaded),
			"is_offloaded": bool(r.is_offloaded),
			"is_returned": bool(r.is_returned),
			"is_completed": bool(r.is_completed),
			"booked_on_date": r.booked_on_date,
			"loaded_on_date": r.loaded_on_date,
			"offloaded_on_date": r.offloaded_on_date,
			"returned_on_date": r.returned_on_date,
			"completed_on_date": r.completed_on_date,
			"discharge_date": r.discharge_date,
			"gate_out_date": r.gate_out_date,
			"empty_return_date": r.empty_return_date,
			"api_container_status": r.api_container_status,
			"api_last_event": r.api_last_event,
			"api_last_event_date": r.api_last_event_date,
		}
		for r in (doc.cargo_parcel_details or [])
	]


@frappe.whitelist()
def get_job_detail(job_name):
	check_freightmas_role()
	check_doc_read_permission("Forwarding Job", job_name)

	doc = frappe.get_doc("Forwarding Job", job_name)

	header = {
		"name": doc.name,
		"customer": doc.customer,
		"customer_reference": doc.customer_reference,
		"shipper": doc.shipper,
		"consignee": doc.consignee,
		"notify_party": doc.notify_party,
		"direction": doc.direction,
		"shipment_mode": doc.shipment_mode,
		"shipment_type": doc.shipment_type,
		"status": doc.status,
		"port_of_loading": doc.port_of_loading,
		"port_of_discharge": doc.port_of_discharge,
		"destination": doc.destination,
		"vessel_flight_no": doc.vessel_flight_no,
		"vessel_flight_date": doc.vessel_flight_date,
		"bl_number": doc.bl_number,
		"is_bl_received": doc.is_bl_received,
		"is_bl_confirmed": doc.is_bl_confirmed,
		"cargo_description": doc.cargo_description,
		"cargo_count": doc.cargo_count,
		"currency": doc.currency,
		"incoterms": doc.incoterms,
		"current_comment": doc.current_comment,
		"last_updated_on": doc.last_updated_on,
		"last_updated_by": doc.last_updated_by,
	}

	shipment_dates = {
		"booking_date": doc.booking_date,
		"cargo_ready_date": doc.cargo_ready_date,
		"etd": doc.etd,
		"atd": doc.atd,
		"eta": doc.eta,
		"ata": doc.ata,
		"discharge_date": doc.discharge_date,
		"completed_on": doc.completed_on,
		"revenue_recognised_on": doc.revenue_recognised_on,
	}

	milestone_stages = _build_job_milestone_stages(doc)

	cargo = _build_job_cargo_list(doc)

	dnd_rows = [
		{
			"container_number": r.container_number,
			"container_type": r.container_type,
			"dnd_free_days": r.dnd_free_days,
			"dnd_rate_per_day": r.dnd_rate_per_day,
			"total_dnd_days": r.total_dnd_days,
			"chargeable_dnd_days": r.chargeable_dnd_days,
			"estimated_dnd_cost": r.estimated_dnd_cost,
			"storage_free_days": r.storage_free_days,
			"storage_rate_per_day": r.storage_rate_per_day,
			"total_storage_days": r.total_storage_days,
			"chargeable_storage_days": r.chargeable_storage_days,
			"estimated_storage_cost": r.estimated_storage_cost,
			"total_container_cost": r.total_container_cost,
		}
		for r in (doc.get("forwarding_dnd_storage_details") or [])
	]

	tracking = [
		{
			"event": r.event,
			"date": r.date,
			"source": r.source,
			"updated_by": r.updated_by,
		}
		for r in sorted(doc.get("tracking_timeline") or [], key=lambda r: r.idx or 0, reverse=True)
	][:15]

	invoices = frappe.get_all(
		"Sales Invoice",
		filters={"forwarding_job_reference": doc.name, "docstatus": ["<", 2]},
		fields=["name", "customer", "posting_date", "due_date", "grand_total", "outstanding_amount", "status"],
	)
	purchase_invoices = frappe.get_all(
		"Purchase Invoice",
		filters={"forwarding_job_reference": doc.name, "docstatus": ["<", 2]},
		fields=["name", "supplier", "posting_date", "due_date", "grand_total", "outstanding_amount", "status"],
	)

	finance = doc.get_job_totals_summary()

	return {
		"header": header,
		"shipment_dates": shipment_dates,
		"milestone_stages": milestone_stages,
		"cargo": cargo,
		"dnd_rows": dnd_rows,
		"dnd_totals": {
			"total_est_dnd_cost": doc.total_est_dnd_cost,
			"total_est_storage_cost": doc.total_est_storage_cost,
			"total_est_dnd_storage_cost": doc.total_est_dnd_storage_cost,
		},
		"tracking": tracking,
		"sales_invoices": invoices,
		"purchase_invoices": purchase_invoices,
		"finance": finance,
	}


# ============================================================
# FINANCE
# ============================================================

def _closed_job_financials():
	"""
	Completed/Closed Forwarding Jobs with quoted vs invoiced (realized) financials.
	Mirrors freightmas/forwarding_service/report/forwarding_closed_job_summary.
	"""
	rows = frappe.db.sql(
		"""
		SELECT
			fj.name, fj.customer, fj.status, fj.completed_on,
			fj.total_quoted_profit_base AS quoted_profit,
			fj.quoted_margin_percent AS quoted_margin_percent,
			COALESCE(si.invoiced_revenue, 0) AS invoiced_revenue,
			COALESCE(pi.invoiced_cost, 0) AS invoiced_cost
		FROM `tabForwarding Job` fj
		LEFT JOIN (
			SELECT forwarding_job_reference, SUM(grand_total) AS invoiced_revenue
			FROM `tabSales Invoice`
			WHERE docstatus = 1 AND forwarding_job_reference IS NOT NULL
			GROUP BY forwarding_job_reference
		) si ON si.forwarding_job_reference = fj.name
		LEFT JOIN (
			SELECT forwarding_job_reference, SUM(grand_total) AS invoiced_cost
			FROM `tabPurchase Invoice`
			WHERE docstatus = 1 AND forwarding_job_reference IS NOT NULL
			GROUP BY forwarding_job_reference
		) pi ON pi.forwarding_job_reference = fj.name
		WHERE fj.status IN ('Completed', 'Closed')
		""",
		as_dict=True,
	)

	for r in rows:
		r["invoiced_profit"] = flt(flt(r.invoiced_revenue) - flt(r.invoiced_cost), 2)
		r["invoiced_margin_percent"] = flt(
			r.invoiced_profit / r.invoiced_revenue * 100, 2
		) if r.invoiced_revenue else 0
		r["profit_variance"] = flt(r.invoiced_profit - flt(r.quoted_profit or 0), 2)
		# Positive gap = invoiced margin fell short of the quoted margin (percentage points).
		r["margin_percent_gap"] = flt(flt(r.quoted_margin_percent or 0) - r["invoiced_margin_percent"], 2)

	return rows


def _loss_making_customers(closed_rows, limit=10):
	agg = {}
	for r in closed_rows:
		if not r.customer:
			continue
		bucket = agg.setdefault(r.customer, {"customer": r.customer, "job_count": 0, "invoiced_profit": 0.0})
		bucket["job_count"] += 1
		bucket["invoiced_profit"] += r.invoiced_profit

	losers = [v for v in agg.values() if v["invoiced_profit"] < 0]
	for v in losers:
		v["invoiced_profit"] = flt(v["invoiced_profit"], 2)
	losers.sort(key=lambda v: v["invoiced_profit"])
	return losers[:limit]


def _margin_at_risk_jobs(threshold_pct, limit=20):
	"""
	Open jobs where working-so-far margin is already trailing quoted margin by more than
	the configured threshold - a chance to intervene before the job closes with a loss.
	"""
	rows = frappe.db.sql(
		"""
		SELECT name, customer, status, quoted_margin_percent, profit_margin_percent
		FROM `tabForwarding Job`
		WHERE docstatus < 2 AND status NOT IN %(statuses)s
		  AND IFNULL(total_working_revenue_base, 0) > 0
		""",
		{"statuses": NOT_ACTIVE_STATUSES},
		as_dict=True,
	)

	at_risk = []
	for r in rows:
		gap = flt(r.quoted_margin_percent or 0) - flt(r.profit_margin_percent or 0)
		if gap >= threshold_pct:
			r["margin_gap"] = flt(gap, 2)
			at_risk.append(r)

	at_risk.sort(key=lambda r: -r["margin_gap"])
	return at_risk[:limit]


def _aging_bucket(days_overdue):
	if days_overdue <= 30:
		return "0-30"
	if days_overdue <= 60:
		return "31-60"
	if days_overdue <= 90:
		return "61-90"
	return "90+"


def _overdue_invoices_with_aging(doctype, display_limit=20):
	party_field = "customer" if doctype == "Sales Invoice" else "supplier"

	rows = frappe.db.sql(
		f"""
		SELECT name, {party_field} AS party, posting_date, due_date, grand_total, outstanding_amount,
			DATEDIFF(CURDATE(), due_date) AS days_overdue
		FROM `tab{doctype}`
		WHERE docstatus = 1 AND outstanding_amount > 0
		  AND forwarding_job_reference IS NOT NULL AND due_date < CURDATE()
		ORDER BY due_date ASC
		LIMIT 500
		""",
		as_dict=True,
	)

	buckets = {b: {"bucket": b, "count": 0, "value": 0.0} for b in ("0-30", "31-60", "61-90", "90+")}
	for r in rows:
		bucket = buckets[_aging_bucket(r.days_overdue)]
		bucket["count"] += 1
		bucket["value"] += flt(r.outstanding_amount)
	for b in buckets.values():
		b["value"] = flt(b["value"], 2)

	return rows[:display_limit], list(buckets.values())


@frappe.whitelist()
def get_finance_summary(from_date=None, to_date=None, customer=None):
	check_freightmas_role()

	conditions = ["fj.docstatus = 1", "fj.status != 'Cancelled'"]
	values = {}
	if from_date:
		conditions.append("fj.creation >= %(from_date)s")
		values["from_date"] = from_date
	if to_date:
		conditions.append("fj.creation <= %(to_date)s")
		values["to_date"] = to_date
	if customer:
		conditions.append("fj.customer = %(customer)s")
		values["customer"] = customer

	where = " AND ".join(conditions)

	rows = frappe.db.sql(
		f"""
		SELECT
			fj.name, fj.customer, fj.status,
			fj.total_quoted_revenue_base AS quoted_revenue,
			fj.total_quoted_cost_base AS quoted_cost,
			fj.total_quoted_profit_base AS quoted_profit,
			fj.total_working_revenue_base AS working_revenue,
			fj.total_working_cost AS working_cost,
			fj.total_working_profit_base AS working_profit,
			COALESCE(si.invoiced_revenue, 0) AS invoiced_revenue,
			COALESCE(pi.invoiced_cost, 0) AS invoiced_cost
		FROM `tabForwarding Job` fj
		LEFT JOIN (
			SELECT forwarding_job_reference, SUM(grand_total) AS invoiced_revenue
			FROM `tabSales Invoice`
			WHERE docstatus = 1 AND forwarding_job_reference IS NOT NULL
			GROUP BY forwarding_job_reference
		) si ON si.forwarding_job_reference = fj.name
		LEFT JOIN (
			SELECT forwarding_job_reference, SUM(grand_total) AS invoiced_cost
			FROM `tabPurchase Invoice`
			WHERE docstatus = 1 AND forwarding_job_reference IS NOT NULL
			GROUP BY forwarding_job_reference
		) pi ON pi.forwarding_job_reference = fj.name
		WHERE {where}
		ORDER BY fj.creation DESC
		LIMIT 200
		""",
		values,
		as_dict=True,
	)

	totals = {"quoted_revenue": 0, "quoted_cost": 0, "working_revenue": 0, "working_cost": 0,
			  "invoiced_revenue": 0, "invoiced_cost": 0}
	for r in rows:
		r["invoiced_profit"] = flt(r.invoiced_revenue) - flt(r.invoiced_cost)
		r["invoiced_margin_percent"] = flt(r.invoiced_profit / r.invoiced_revenue * 100, 2) if r.invoiced_revenue else 0
		for k in totals:
			totals[k] += flt(r.get(k) or 0)

	totals["quoted_profit"] = flt(totals["quoted_revenue"] - totals["quoted_cost"], 2)
	totals["working_profit"] = flt(totals["working_revenue"] - totals["working_cost"], 2)
	totals["invoiced_profit"] = flt(totals["invoiced_revenue"] - totals["invoiced_cost"], 2)
	totals["invoiced_margin_percent"] = flt(
		totals["invoiced_profit"] / totals["invoiced_revenue"] * 100, 2
	) if totals["invoiced_revenue"] else 0

	overdue_sales, sales_aging = _overdue_invoices_with_aging("Sales Invoice")
	overdue_purchase, purchase_aging = _overdue_invoices_with_aging("Purchase Invoice")

	margin_variance_threshold = flt(
		frappe.db.get_single_value("FreightMas Settings", "margin_variance_threshold_pct") or 5
	)
	closed_jobs = _closed_job_financials()
	loss_making_jobs = sorted(
		(r for r in closed_jobs if r.invoiced_profit < 0), key=lambda r: r["invoiced_profit"]
	)
	margin_variance_jobs = sorted(closed_jobs, key=lambda r: -r["margin_percent_gap"])[:20]
	for r in margin_variance_jobs:
		r["breaches_threshold"] = r["margin_percent_gap"] >= margin_variance_threshold
	loss_making_customers = _loss_making_customers(closed_jobs)
	margin_at_risk_jobs = _margin_at_risk_jobs(margin_variance_threshold)

	top_customers_by_revenue = frappe.db.sql(
		"""
		SELECT forwarding_job_reference AS job, customer,
		       SUM(grand_total) AS revenue
		FROM `tabSales Invoice`
		WHERE docstatus = 1 AND forwarding_job_reference IS NOT NULL
		GROUP BY customer
		ORDER BY revenue DESC
		LIMIT 5
		""",
		as_dict=True,
	)

	return {
		"jobs": rows,
		"totals": totals,
		"overdue_sales_invoices": overdue_sales,
		"overdue_purchase_invoices": overdue_purchase,
		"aging_summary": {"sales": sales_aging, "purchase": purchase_aging},
		"loss_making_jobs": loss_making_jobs,
		"margin_variance_jobs": margin_variance_jobs,
		"loss_making_customers": loss_making_customers,
		"margin_at_risk_jobs": margin_at_risk_jobs,
		"top_customers_by_revenue": top_customers_by_revenue,
		"monthly_trend": _monthly_revenue_margin_trend(months=12),
	}


# ============================================================
# DND & ADDITIONAL COSTS
# ============================================================

@frappe.whitelist()
def get_dnd_overview():
	check_freightmas_role()

	jobs = frappe.db.sql(
		"""
		SELECT
			fj.name, fj.customer, fj.status, fj.bl_number,
			fj.total_est_dnd_cost, fj.total_est_storage_cost, fj.total_est_dnd_storage_cost
		FROM `tabForwarding Job` fj
		WHERE fj.docstatus < 2 AND fj.status NOT IN %(statuses)s
		  AND IFNULL(fj.total_est_dnd_storage_cost, 0) > 0
		ORDER BY fj.total_est_dnd_storage_cost DESC
		""",
		{"statuses": NOT_ACTIVE_STATUSES},
		as_dict=True,
	)

	job_names = [j.name for j in jobs]
	containers = []
	if job_names:
		containers = frappe.db.sql(
			"""
			SELECT
				parent, container_number, container_type,
				dnd_free_days, dnd_rate_per_day, total_dnd_days, chargeable_dnd_days, estimated_dnd_cost,
				storage_free_days, storage_rate_per_day, total_storage_days, chargeable_storage_days, estimated_storage_cost,
				total_container_cost
			FROM `tabForwarding DND Storage Detail`
			WHERE parenttype = 'Forwarding Job' AND parent IN %(names)s
			ORDER BY total_container_cost DESC
			""",
			{"names": job_names},
			as_dict=True,
		)

	totals = frappe.db.sql(
		"""
		SELECT
			COALESCE(SUM(total_est_dnd_cost), 0) AS total_dnd,
			COALESCE(SUM(total_est_storage_cost), 0) AS total_storage,
			COALESCE(SUM(total_est_dnd_storage_cost), 0) AS total_combined,
			COUNT(*) AS job_count
		FROM `tabForwarding Job`
		WHERE docstatus < 2 AND status NOT IN %(statuses)s AND IFNULL(total_est_dnd_storage_cost, 0) > 0
		""",
		{"statuses": NOT_ACTIVE_STATUSES},
		as_dict=True,
	)[0]

	overdue_returns = frappe.db.sql(
		"""
		SELECT cpd.parent AS job, cpd.container_number, cpd.return_by_date, fj.customer
		FROM `tabCargo Parcel Details` cpd
		INNER JOIN `tabForwarding Job` fj ON fj.name = cpd.parent
		WHERE cpd.parenttype = 'Forwarding Job' AND cpd.cargo_type = 'Containerised'
		  AND IFNULL(cpd.to_be_returned, 0) = 1 AND IFNULL(cpd.is_returned, 0) = 0
		  AND cpd.empty_return_date IS NULL
		  AND cpd.return_by_date IS NOT NULL AND cpd.return_by_date < %(today)s
		  AND fj.status NOT IN %(statuses)s
		ORDER BY cpd.return_by_date ASC
		""",
		{"today": nowdate(), "statuses": NOT_ACTIVE_STATUSES},
		as_dict=True,
	)

	at_risk_containers = _get_at_risk_containers()

	return {
		"jobs": jobs,
		"containers": containers,
		"totals": totals,
		"overdue_returns": overdue_returns,
		"at_risk_containers": at_risk_containers,
		"at_risk_count": len(at_risk_containers),
	}


# ============================================================
# EXCEL EXPORTS
# ============================================================

_HEADER_FILL = PatternFill("solid", fgColor="17406B")
_ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_SUBTITLE_FONT = Font(bold=True, size=11, color="667085")
_BORDER = Border(
	left=Side(style="thin", color="DDDDDD"),
	right=Side(style="thin", color="DDDDDD"),
	top=Side(style="thin", color="DDDDDD"),
	bottom=Side(style="thin", color="DDDDDD"),
)
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
NUMERIC_TYPES = ("Currency", "Int", "Float", "Percent")


def _write_sheet(ws, sheet_title, columns, rows):
	"""columns: list of dicts {label, fieldname, fieldtype}."""
	ws.title = sheet_title[:31]
	ncols = len(columns)

	company = frappe.defaults.get_global_default("company") or frappe.defaults.get_user_default("Company") or "FreightMas"
	row_idx = 1
	ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
	ws.cell(row=row_idx, column=1, value=company).font = _TITLE_FONT
	row_idx += 1

	ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
	export_time = frappe.utils.now_datetime().strftime("%d-%b-%Y %H:%M")
	ws.cell(row=row_idx, column=1, value=f"{sheet_title} \u2014 Exported {export_time}").font = _SUBTITLE_FONT
	row_idx += 2

	header_row = row_idx
	for col_idx, col in enumerate(columns, start=1):
		cell = ws.cell(row=header_row, column=col_idx, value=col["label"])
		cell.font = _HEADER_FONT
		cell.fill = _HEADER_FILL
		cell.alignment = _LEFT
		cell.border = _BORDER
	row_idx += 1
	ws.freeze_panes = ws[f"A{header_row + 1}"]

	for i, row in enumerate(rows, start=1):
		fill = _ZEBRA_FILL if i % 2 == 0 else None
		for col_idx, col in enumerate(columns, start=1):
			value = row.get(col["fieldname"], "")
			cell = ws.cell(row=row_idx, column=col_idx)
			fieldtype = col.get("fieldtype", "Data")
			if fieldtype in NUMERIC_TYPES:
				cell.value = flt(value) if value not in (None, "") else 0
				cell.number_format = "0.0%" if fieldtype == "Percent" else "#,##0.00"
				cell.alignment = _RIGHT
			elif fieldtype in ("Date", "Datetime") and value:
				try:
					cell.value = formatdate(value, "dd-MMM-yy")
				except Exception:
					cell.value = str(value)
				cell.alignment = _LEFT
			else:
				cell.value = value if value not in (None,) else ""
				cell.alignment = _LEFT
			cell.border = _BORDER
			if fill:
				cell.fill = fill
		row_idx += 1

	for col_idx, col in enumerate(columns, start=1):
		max_length = len(col["label"])
		for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
			for cell in row:
				try:
					max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
				except Exception:
					pass
		ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_length + 2, 40))

	ws.sheet_view.showGridLines = False


def _workbook_bytes(wb):
	output = BytesIO()
	wb.save(output)
	output.seek(0)
	return output.read()


def _send_workbook(wb, filename):
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = _workbook_bytes(wb)
	frappe.local.response.type = "binary"


def _timestamped(name):
	return f"{name}_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M')}.xlsx"


@frappe.whitelist()
def export_jobs(customer=None, status=None, direction=None, operational_phase=None, operational_phases=None, overview_bucket=None, search=None):
	"""Export the (unpaginated, filtered) Shipments list to Excel."""
	check_freightmas_role()

	res = get_jobs(
		customer=customer, status=status, direction=direction,
		operational_phase=operational_phase, operational_phases=operational_phases,
		overview_bucket=overview_bucket, search=search,
		limit_start=0, limit_page_length=5000,
	)

	columns = [
		{"label": "Job", "fieldname": "name"},
		{"label": "Customer", "fieldname": "customer"},
		{"label": "Reference", "fieldname": "customer_reference"},
		{"label": "Direction", "fieldname": "direction"},
		{"label": "Mode", "fieldname": "shipment_mode"},
		{"label": "Origin", "fieldname": "port_of_loading"},
		{"label": "Discharge", "fieldname": "port_of_discharge"},
		{"label": "Destination", "fieldname": "destination"},
		{"label": "Vessel / Flight", "fieldname": "vessel_flight_no"},
		{"label": "BL Number", "fieldname": "bl_number"},
		{"label": "ETA", "fieldname": "eta", "fieldtype": "Date"},
		{"label": "ATA", "fieldname": "ata", "fieldtype": "Date"},
		{"label": "ETD", "fieldname": "etd", "fieldtype": "Date"},
		{"label": "ATD", "fieldname": "atd", "fieldtype": "Date"},
		{"label": "Status", "fieldname": "status"},
		{"label": "Operational Phase", "fieldname": "operational_phase_label"},
		{"label": "Substage", "fieldname": "operational_substage"},
		{"label": "Milestone %", "fieldname": "milestone_percent", "fieldtype": "Float"},
	]

	wb = openpyxl.Workbook()
	_write_sheet(wb.active, "Shipments", columns, res["jobs"])
	_send_workbook(wb, _timestamped("Shipments"))


@frappe.whitelist()
def export_finance(from_date=None, to_date=None, customer=None):
	"""Export the Finance job profitability table to Excel."""
	check_freightmas_role()

	res = get_finance_summary(from_date=from_date, to_date=to_date, customer=customer)

	columns = [
		{"label": "Job", "fieldname": "name"},
		{"label": "Customer", "fieldname": "customer"},
		{"label": "Status", "fieldname": "status"},
		{"label": "Quoted Revenue", "fieldname": "quoted_revenue", "fieldtype": "Currency"},
		{"label": "Quoted Cost", "fieldname": "quoted_cost", "fieldtype": "Currency"},
		{"label": "Working Revenue", "fieldname": "working_revenue", "fieldtype": "Currency"},
		{"label": "Working Cost", "fieldname": "working_cost", "fieldtype": "Currency"},
		{"label": "Invoiced Revenue", "fieldname": "invoiced_revenue", "fieldtype": "Currency"},
		{"label": "Invoiced Cost", "fieldname": "invoiced_cost", "fieldtype": "Currency"},
		{"label": "Invoiced Profit", "fieldname": "invoiced_profit", "fieldtype": "Currency"},
		{"label": "Invoiced Margin %", "fieldname": "invoiced_margin_percent", "fieldtype": "Float"},
	]

	wb = openpyxl.Workbook()
	_write_sheet(wb.active, "Finance", columns, res["jobs"])
	_send_workbook(wb, _timestamped("Finance"))


@frappe.whitelist()
def export_dnd():
	"""Export the DND & Additional Costs overview (jobs + container detail) to Excel."""
	check_freightmas_role()

	res = get_dnd_overview()

	job_columns = [
		{"label": "Job", "fieldname": "name"},
		{"label": "Customer", "fieldname": "customer"},
		{"label": "Status", "fieldname": "status"},
		{"label": "BL Number", "fieldname": "bl_number"},
		{"label": "DND Cost", "fieldname": "total_est_dnd_cost", "fieldtype": "Currency"},
		{"label": "Storage Cost", "fieldname": "total_est_storage_cost", "fieldtype": "Currency"},
		{"label": "Total Exposure", "fieldname": "total_est_dnd_storage_cost", "fieldtype": "Currency"},
	]
	container_columns = [
		{"label": "Job", "fieldname": "parent"},
		{"label": "Container", "fieldname": "container_number"},
		{"label": "Type", "fieldname": "container_type"},
		{"label": "DND Free Days", "fieldname": "dnd_free_days", "fieldtype": "Int"},
		{"label": "DND Rate/Day", "fieldname": "dnd_rate_per_day", "fieldtype": "Currency"},
		{"label": "Chargeable DND Days", "fieldname": "chargeable_dnd_days", "fieldtype": "Int"},
		{"label": "DND Cost", "fieldname": "estimated_dnd_cost", "fieldtype": "Currency"},
		{"label": "Storage Free Days", "fieldname": "storage_free_days", "fieldtype": "Int"},
		{"label": "Storage Rate/Day", "fieldname": "storage_rate_per_day", "fieldtype": "Currency"},
		{"label": "Chargeable Storage Days", "fieldname": "chargeable_storage_days", "fieldtype": "Int"},
		{"label": "Storage Cost", "fieldname": "estimated_storage_cost", "fieldtype": "Currency"},
		{"label": "Total Cost", "fieldname": "total_container_cost", "fieldtype": "Currency"},
	]

	wb = openpyxl.Workbook()
	_write_sheet(wb.active, "DND Jobs", job_columns, res["jobs"])
	_write_sheet(wb.create_sheet("Containers"), "Containers", container_columns, res["containers"])
	_send_workbook(wb, _timestamped("DND_Additional_Costs"))


# ============================================================
# TRACKING REPORT (color-coded milestone grid)
# ============================================================
# Same column/section/shading design as freightmas.portal.api.shipments's
# export_tracking_report - ported as a separate copy rather than imported,
# per this module's existing deliberate independence from the portal (see
# that module's own docstring: it's a re-implementation, not a delegation,
# so the internal-staff and customer-portal trust boundaries never mix).
# The only structural difference here is a multi-customer `customers` filter
# instead of being implicitly locked to one caller's own customer.

TRACKING_AMBER_FILL = "FDEBD0"
TRACKING_GREEN_FILL = "D5F5E3"
TRACKING_NA_FILL = "D9D9D9"
TRACKING_PERCENT_FILL = "E8EDF7"

# Sentinel for a Port/Border Clearance column when that service isn't
# required (ticked) on the job at all - rendered as "-" with a neutral fill,
# excluded from the % Complete applicable/completed counts.
TRACKING_NOT_APPLICABLE = object()

TRACKING_SHIPMENT_DETAILS_SECTION = ("Shipment Details", "595959", [
	("Job ID", "job_id", "identity"),
	("Consignee", "consignee", "identity"),
	("Customer Reference", "customer_reference", "identity"),
	("BL Number", "bl_number", "identity"),
	("Container Number", "container_number", "identity"),
	("Type", "container_type", "identity"),
])

TRACKING_SEA_AIR_SECTION = ("Sea / Air Freight", "2E5C8A", [
	("Departed Origin (ATD)", "atd", "milestone"),
	("ETA/ATA", "eta_ata", "milestone"),
	("Discharged (Shipment)", "discharge_date", "milestone"),
	("Container Discharged", "container_discharge_date", "milestone"),
	("Gate Out", "gate_out_date", "milestone"),
	("Empty Returned", "empty_return_date", "milestone"),
])

TRACKING_ROAD_TRANSPORT_SECTION = ("Road Transport", "1E7A6F", [
	("Booked", "booked_on_date", "milestone"),
	("Loaded", "loaded_on_date", "milestone"),
	("Offloaded", "offloaded_on_date", "milestone"),
	("Returned", "returned_on_date", "milestone"),
	("Completed", "trucking_completed_on_date", "milestone"),
])

TRACKING_OVERVIEW_SECTION = ("Overview", "6B6B6B", [
	("Completed", "job_completed_on", "milestone"),
	("Status", "status_comment", "status"),
	("% Complete", "percent_complete", "percent"),
])

# service_module -> band color, for the dynamic milestone sections built live
# from Milestone Definition (sequence order).
TRACKING_DYNAMIC_MILESTONE_COLORS = {
	"Port Clearance": "5B3A8E",
	"Border Clearance": "A04B2E",
}


def _tracking_milestone_definition_columns(service_module):
	"""One (label, fieldname) column per active Milestone Definition for this
	service_module, in sequence order, deduped by milestone_code (used as the
	fieldname - already globally unique, e.g. "PC_VESSEL_ARRIVED")."""
	rows = frappe.get_all(
		"Milestone Definition",
		filters={"service_module": service_module, "is_active": 1},
		fields=["milestone_code", "milestone_label"],
		order_by="sequence asc",
	)
	seen = set()
	columns = []
	for r in rows:
		if r.milestone_code in seen:
			continue
		seen.add(r.milestone_code)
		columns.append((r.milestone_label, r.milestone_code, "milestone"))
	return columns


def _tracking_dynamic_section(service_module):
	return (
		service_module,
		TRACKING_DYNAMIC_MILESTONE_COLORS[service_module],
		_tracking_milestone_definition_columns(service_module),
	)


def _build_tracking_report_sections():
	"""Full ordered section list: Shipment Details, Sea/Air Freight, Port
	Clearance, Road Transport, Border Clearance, Overview."""
	return [
		TRACKING_SHIPMENT_DETAILS_SECTION,
		TRACKING_SEA_AIR_SECTION,
		_tracking_dynamic_section("Port Clearance"),
		TRACKING_ROAD_TRANSPORT_SECTION,
		_tracking_dynamic_section("Border Clearance"),
		TRACKING_OVERVIEW_SECTION,
	]


def _build_tracking_workbook(sections, rows):
	"""One flat table: every individual milestone gets its own date column,
	color-banded into sections, with each milestone cell shaded amber (blank/
	outstanding), green (date present/achieved), or grey with a "-"
	(TRACKING_NOT_APPLICABLE), and Shipment Details cells shaded amber too
	(matching the outstanding-milestone color, per the portal version's own
	final styling). The 3 header rows are frozen. `sections` is the output of
	_build_tracking_report_sections(); `rows` is a list of flat
	{fieldname: value} dicts, one per report row."""
	columns = []  # (label, fieldname, kind)
	band_spans = []  # (title, color, start_col, end_col)
	col_idx = 1
	for title, color, cols in sections:
		start = col_idx
		for label, fieldname, kind in cols:
			columns.append((label, fieldname, kind))
			col_idx += 1
		band_spans.append((title, color, start, col_idx - 1))
	ncols = len(columns)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Tracking Report"

	title_font = Font(bold=True, size=14, color="1F2A44")
	band_font = Font(bold=True, color="FFFFFF")
	header_font = Font(bold=True, color="FFFFFF")
	center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
	left_align = Alignment(horizontal="left", vertical="center")
	percent_align = Alignment(horizontal="center", vertical="center")

	ws.cell(row=1, column=1, value="Shipment Tracking Report").font = title_font
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

	for title, color, start, end in band_spans:
		cell = ws.cell(row=2, column=start, value=title)
		cell.fill = PatternFill("solid", fgColor=color)
		cell.font = band_font
		cell.alignment = center_wrap
		if end > start:
			ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)

	band_color_by_col = {}
	for title, color, start, end in band_spans:
		for c in range(start, end + 1):
			band_color_by_col[c] = color

	header_row = 3
	for idx, (label, fieldname, kind) in enumerate(columns, 1):
		cell = ws.cell(row=header_row, column=idx, value=label)
		cell.fill = PatternFill("solid", fgColor=band_color_by_col[idx])
		cell.font = header_font
		cell.alignment = center_wrap

	data_start = header_row + 1
	for row_offset, row_data in enumerate(rows):
		row_idx = data_start + row_offset
		for col_idx, (label, fieldname, kind) in enumerate(columns, 1):
			value = row_data.get(fieldname)
			cell = ws.cell(row=row_idx, column=col_idx)

			if kind == "milestone":
				cell.value = "-" if value is TRACKING_NOT_APPLICABLE else (formatdate(value, "dd-MMM-yy") if value else "")
				cell.fill = _milestone_cell_fill(value)
				cell.alignment = percent_align
			elif kind == "percent":
				cell.value = value or 0
				cell.number_format = "0%"
				cell.fill = PatternFill("solid", fgColor=TRACKING_PERCENT_FILL)
				cell.alignment = percent_align
			elif kind == "identity":
				cell.value = value or ""
				cell.alignment = left_align
				cell.fill = PatternFill("solid", fgColor=TRACKING_AMBER_FILL)
			else:  # status
				cell.value = value or ""
				cell.alignment = left_align

	ws.freeze_panes = f"A{data_start}"  # pins the 3 header rows only, no column freeze

	# Milestone (date) and percent columns get a fixed, compact width
	# regardless of their (often long) header label - "01-Jul-26"/"-"/blank
	# never need more than this, and the header wraps (center_wrap, set
	# above) to fit instead of forcing the column wide. Percent needs the
	# same fixed treatment for a different reason: auto-fit measures each
	# cell's raw (unformatted) value, and a raw float like
	# 0.037037037037037035 is far longer than what "0%" actually displays.
	# Other columns still auto-fit, but sized off the data rows only, not
	# the header label.
	for col_idx, (label, fieldname, kind) in enumerate(columns, 1):
		if kind in ("milestone", "percent"):
			ws.column_dimensions[get_column_letter(col_idx)].width = 11
			continue
		max_length = 0
		for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
			for cell in row:
				length = len(str(cell.value)) if cell.value else 0
				max_length = max(max_length, length)
		ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_length + 2, 30))

	ws.sheet_view.showGridLines = False

	return wb


@frappe.whitelist()
def get_customers():
	"""Distinct customers with at least one Forwarding Job - for populating
	the tracking report's customer filter (not the entire Customer master,
	which may include customers unrelated to freight forwarding)."""
	check_freightmas_role()

	return frappe.db.sql(
		"""
		SELECT DISTINCT c.name, c.customer_name
		FROM `tabCustomer` c
		INNER JOIN `tabForwarding Job` fj ON fj.customer = c.name
		WHERE fj.docstatus < 2
		ORDER BY c.customer_name
		""",
		as_dict=True,
	)


def _gather_tracking_report_rows(customers=None, status=None, direction=None, search=None):
	"""Shared data layer for the Full Tracking Report and the Master Tracking
	Report. Returns (sections, report_rows, containers_by_job):
	  - sections: output of _build_tracking_report_sections().
	  - report_rows: flat {fieldname: value} dicts, one per (job, container) -
	    always at least one row per job (a containerless job gets a single
	    synthetic placeholder row, same as before).
	  - containers_by_job: {job_id: [row, ...]}, the same row dicts grouped by
	    job - lets a caller aggregate per-job (e.g. the Master report's "At a
	    Glance" sheet) without re-scanning report_rows.
	"""
	if customers and isinstance(customers, str):
		customers = frappe.parse_json(customers)

	# Tracking is for jobs still being worked - once a job is submitted it's
	# considered finalized and drops out of this report (unlike get_jobs/
	# export_jobs, which show submitted jobs too).
	filters = {"docstatus": 0}
	if customers:
		filters["customer"] = ["in", customers]
	if status:
		filters["status"] = status
	if direction:
		filters["direction"] = direction

	or_filters = None
	if search:
		or_filters = [
			["name", "like", f"%{search}%"],
			["customer_reference", "like", f"%{search}%"],
			["bl_number", "like", f"%{search}%"],
		]

	jobs = frappe.get_all(
		"Forwarding Job",
		filters=filters,
		or_filters=or_filters,
		fields=[
			"name", "consignee", "customer_reference", "bl_number", "current_comment",
			"atd", "eta", "ata", "discharge_date", "completed_on",
			"requires_port_clearance", "requires_border_clearance", "is_trucking_required",
		],
		order_by="modified desc",
	)
	job_names = [j.name for j in jobs]

	customer_name_map = {}
	consignee_names = list({j.consignee for j in jobs if j.consignee})
	if consignee_names:
		customer_name_map = {
			c.name: c.customer_name
			for c in frappe.get_all(
				"Customer", filters={"name": ["in", consignee_names]}, fields=["name", "customer_name"]
			)
		}

	sections = _build_tracking_report_sections()
	dynamic_milestone_fields = {
		title: [fieldname for _label, fieldname, _kind in cols]
		for title, _color, cols in sections
		if title in ("Port Clearance", "Border Clearance")
	}
	section_to_table = {
		"Port Clearance": "port_clearance_milestones",
		"Border Clearance": "border_clearance_milestones",
	}
	section_requires_field = {
		"Port Clearance": "requires_port_clearance",
		"Border Clearance": "requires_border_clearance",
	}

	cargo_by_job = {}
	milestone_rows_by_job = {table: {} for table in section_to_table.values()}

	if job_names:
		cargo_rows = frappe.get_all(
			"Cargo Parcel Details",
			filters={"parenttype": "Forwarding Job", "parent": ["in", job_names]},
			fields=[
				"parent", "container_number", "container_type",
				"discharge_date", "gate_out_date", "empty_return_date",
				"booked_on_date", "loaded_on_date", "offloaded_on_date",
				"returned_on_date", "completed_on_date",
			],
			order_by="parent, idx",
		)
		for r in cargo_rows:
			cargo_by_job.setdefault(r.parent, []).append(r)

		for table in section_to_table.values():
			rows = frappe.get_all(
				"Job Milestone Progress",
				filters={"parenttype": "Forwarding Job", "parentfield": table, "parent": ["in", job_names]},
				fields=["parent", "milestone_code", "is_completed", "completed_on"],
				order_by="parent, idx",
			)
			for r in rows:
				milestone_rows_by_job[table].setdefault(r.parent, []).append(r)

	report_rows = []
	containers_by_job = {}
	for job in jobs:
		real_containers = cargo_by_job.get(job.name, [])

		job_fields = {
			"job_id": job.name,
			"consignee": customer_name_map.get(job.consignee, job.consignee),
			"customer_reference": job.customer_reference,
			"bl_number": job.bl_number,
			"atd": job.atd,
			"eta_ata": job.ata or job.eta,
			"discharge_date": job.discharge_date,
			"job_completed_on": job.completed_on,
			"status_comment": job.current_comment,
			"is_trucking_required": bool(job.is_trucking_required),
			"parcel_count": len(real_containers),
		}
		sea_air_job_applicable = ["atd", "eta_ata", "discharge_date"]

		# Per-job milestone maps: {milestone_code: completed_on}, plus counts
		# for % Complete. A service not required (ticked) on this job at all
		# gets TRACKING_NOT_APPLICABLE on every one of its milestone columns
		# (rendered as "-") and contributes nothing to applicable/completed -
		# explicitly gated on the job's own requires_* flag, not inferred
		# from row presence (populate_mode_milestones() never removes rows
		# if a requires_* flag is later unticked).
		section_counts = {}  # title -> (applicable, completed)
		for title, table in section_to_table.items():
			required = bool(job.get(section_requires_field[title]))
			rows = milestone_rows_by_job[table].get(job.name, []) if required else []
			value_map = {r.milestone_code: r.completed_on for r in rows if r.is_completed}
			for fieldname in dynamic_milestone_fields[title]:
				job_fields[fieldname] = value_map.get(fieldname) if required else TRACKING_NOT_APPLICABLE
			section_counts[title] = (len(rows), sum(1 for r in rows if r.is_completed))

		containers = real_containers or [None]

		for c in containers:
			row = dict(job_fields)
			if c is None:
				row.update({
					"container_number": None, "container_type": None,
					"container_discharge_date": None, "gate_out_date": None, "empty_return_date": None,
					"booked_on_date": None, "loaded_on_date": None, "offloaded_on_date": None,
					"returned_on_date": None, "trucking_completed_on_date": None,
				})
			else:
				row.update({
					"container_number": c.container_number,
					"container_type": c.container_type,
					"container_discharge_date": c.discharge_date,
					"gate_out_date": c.gate_out_date,
					"empty_return_date": c.empty_return_date,
					"booked_on_date": c.booked_on_date,
					"loaded_on_date": c.loaded_on_date,
					"offloaded_on_date": c.offloaded_on_date,
					"returned_on_date": c.returned_on_date,
					"trucking_completed_on_date": c.completed_on_date,
				})

			road_transport_fields = [
				"booked_on_date", "loaded_on_date", "offloaded_on_date",
				"returned_on_date", "trucking_completed_on_date",
			]
			sea_air_container_fields = ["container_discharge_date", "gate_out_date", "empty_return_date"]

			applicable = len(sea_air_job_applicable) + len(sea_air_container_fields) + len(road_transport_fields) + 1
			completed = (
				sum(1 for f in sea_air_job_applicable if row.get(f))
				+ sum(1 for f in sea_air_container_fields if row.get(f))
				+ sum(1 for f in road_transport_fields if row.get(f))
				+ (1 if row.get("job_completed_on") else 0)
			)
			for title in section_to_table:
				a, done = section_counts[title]
				applicable += a
				completed += done

			row["percent_complete"] = (completed / applicable) if applicable else 0
			report_rows.append(row)
			containers_by_job.setdefault(job.name, []).append(row)

	return sections, report_rows, containers_by_job


@frappe.whitelist()
def export_tracking_report(customers=None, status=None, direction=None, search=None):
	check_freightmas_role()
	sections, report_rows, _containers_by_job = _gather_tracking_report_rows(
		customers=customers, status=status, direction=direction, search=search
	)
	wb = _build_tracking_workbook(sections, report_rows)
	_send_workbook(wb, _timestamped("Tracking_Report"))


def _milestone_cell_fill(value):
	"""Shared by the Full Tracking Report's flat grid and the Master Tracking
	Report's Detailed Tracking sheet: green = achieved, amber = outstanding,
	grey "-" = not applicable (service not required on this job)."""
	if value is TRACKING_NOT_APPLICABLE:
		return PatternFill("solid", fgColor=TRACKING_NA_FILL)
	return PatternFill("solid", fgColor=TRACKING_GREEN_FILL if value else TRACKING_AMBER_FILL)


# ============================================================
# MASTER TRACKING REPORT (stage ladder + Summary / At a Glance /
# Detailed Tracking / Exceptions / Config workbook)
# ============================================================
# Adds one unifying concept on top of _gather_tracking_report_rows()'s
# existing flat rows: a single ordered "stage ladder" walking every
# milestone/date field in sequence (mirrors a client-supplied reference
# workbook's "Config" sheet), so each row reduces to one "Current Stage",
# an ageing/exception layer, and 5 stage-buckets. Everything below is
# computed once in Python at export time and written as static values/
# fills - there are no live Excel formulas.

MASTER_STAGE_BUCKETS = [
	# (bucket_name, fill_color, font_color) - pastel per-row "status pill"
	# used on Current Stage/Stage No. cells. Distinct from the bold
	# section-header band colors in MASTER_DETAIL_BAND_COLORS below.
	("NOT YET DEPARTED", "F2F2F2", "595959"),
	("ON WATER / AIR", "E8EFF8", "1F3864"),
	("AT PORT / IN CLEARANCE", "FFF4E0", "8A5200"),
	("ON ROAD / AT BORDER", "E9F2F2", "0F5C5C"),
	("DELIVERED / COMPLETED", "EAF3E6", "3D6B29"),
]

MASTER_DETAIL_BAND_COLORS = {
	"Shipment Details": "17406B",
	"Sea / Air Freight": "178A8A",
	"Port Clearance": "4A2D7F",
	"Road Transport & Delivery": "F58220",
	"Border Clearance": "0F5C5C",
	"Overview": "17406B",
	"Progress & Status": "17406B",
}

MASTER_AT_PORT_DAYS_THRESHOLD = 14
MASTER_ARRIVING_SOON_DAYS = 14


def _build_stage_ladder(sections):
	"""Builds the ordered "stage ladder" _annotate_stage_fields() walks to
	derive each row's Current Stage. Dynamic: the Port/Border Clearance
	rungs are pulled straight from `sections` (itself built live from active
	Milestone Definition rows), so the ladder stays correct if those change.
	Returns {"rungs": [...], "by_field": {field: rung}, "by_seq": {seq: rung}}.
	"""
	section_cols = {title: cols for title, _color, cols in sections}

	def dynamic_rungs(section_title, block):
		return [
			{"label": label, "field": fieldname, "block": block}
			for label, fieldname, _kind in section_cols.get(section_title, [])
		]

	rungs = [{"seq": 0, "label": "Booked - awaiting departure", "field": None, "block": "Overview"}]
	fixed_rungs = (
		[
			{"label": "Departed origin", "field": "atd", "block": "Sea / Air Freight"},
			{"label": "Vessel discharged", "field": "discharge_date", "block": "Sea / Air Freight"},
			{"label": "Container discharged at port", "field": "container_discharge_date", "block": "Sea / Air Freight"},
		]
		+ dynamic_rungs("Port Clearance", "Port Clearance")
		+ [
			{"label": "Gated out of port", "field": "gate_out_date", "block": "Sea / Air Freight"},
			{"label": "Transport booked", "field": "booked_on_date", "block": "Road Transport & Delivery"},
			{"label": "Loaded for road transport", "field": "loaded_on_date", "block": "Road Transport & Delivery"},
		]
		+ dynamic_rungs("Border Clearance", "Border Clearance")
		+ [
			{"label": "Offloaded at destination", "field": "offloaded_on_date", "block": "Road Transport & Delivery"},
			{"label": "Truck returned", "field": "returned_on_date", "block": "Road Transport & Delivery"},
			{"label": "Road leg completed", "field": "trucking_completed_on_date", "block": "Road Transport & Delivery"},
			{"label": "Empty container returned", "field": "empty_return_date", "block": "Sea / Air Freight"},
			{"label": "Job completed", "field": "job_completed_on", "block": "Overview"},
		]
	)
	for i, rung in enumerate(fixed_rungs, start=1):
		rung["seq"] = i
		rungs.append(rung)

	by_field = {r["field"]: r for r in rungs if r["field"]}
	by_seq = {r["seq"]: r for r in rungs}
	return {"rungs": rungs, "by_field": by_field, "by_seq": by_seq}


def _stage_bucket_bounds(ladder):
	"""Resolves the 5 bucket boundaries by looking up known rungs by field
	name rather than counting, so it stays correct however many Port/Border
	Clearance milestones exist. Returns [(name, fill, font, start_seq, end_seq)]."""
	by_field = ladder["by_field"]
	last_seq = ladder["rungs"][-1]["seq"]

	vessel_discharged_seq = by_field["discharge_date"]["seq"]
	gate_out_seq = by_field["gate_out_date"]["seq"]
	loaded_seq = by_field["loaded_on_date"]["seq"]
	border_seqs = [r["seq"] for r in ladder["rungs"] if r["field"] and r["field"].startswith("BC_")]
	border_end_seq = max(border_seqs) if border_seqs else loaded_seq

	seq_bounds = [
		(0, 0),
		(1, vessel_discharged_seq),
		(vessel_discharged_seq + 1, gate_out_seq),
		(gate_out_seq + 1, border_end_seq),
		(border_end_seq + 1, last_seq),
	]
	return [
		(name, fill, font, start, end)
		for (name, fill, font), (start, end) in zip(MASTER_STAGE_BUCKETS, seq_bounds)
	]


def _bucket_for_seq(seq, bounds):
	for name, fill, font, start, end in bounds:
		if start <= seq <= end:
			return name, fill, font
	return bounds[-1][0], bounds[-1][1], bounds[-1][2]


def _pred_at_port_over_threshold(row):
	days = row.get("_days_at_port")
	return bool(days and days > MASTER_AT_PORT_DAYS_THRESHOLD)


def _pred_eta_passed_no_discharge(row, today):
	eta = row.get("eta_ata")
	return bool(
		eta and getdate(eta) < today
		and not row.get("discharge_date") and not row.get("container_discharge_date")
	)


def _pred_awaiting_departure(row):
	return row.get("_stage_seq") == 0


def _pred_arriving_soon(row, days_ahead=MASTER_ARRIVING_SOON_DAYS):
	days = row.get("_days_to_eta")
	return days is not None and 0 <= days <= days_ahead


def _annotate_stage_fields(report_rows, ladder):
	"""Mutates each row in report_rows, adding the fields the Master Tracking
	Report needs on top of what _gather_tracking_report_rows already
	produces: _stage_seq/_stage_label/_bucket_name/_bucket_fill/_bucket_font,
	_days_to_eta, _days_at_port. `percent_complete` (already on every row) is
	reused as-is, not recomputed."""
	today = getdate(nowdate())
	bounds = _stage_bucket_bounds(ladder)
	rungs = ladder["rungs"][1:]  # skip the seq-0 baseline, it has no field
	discharge_seq = ladder["by_field"]["discharge_date"]["seq"]

	for row in report_rows:
		achieved_seq = 0
		for rung in rungs:
			value = row.get(rung["field"])
			if value in (None, "", TRACKING_NOT_APPLICABLE):
				continue
			if getdate(value) <= today:
				achieved_seq = rung["seq"]  # ascending order - last match wins (highest)
		row["_stage_seq"] = achieved_seq
		row["_stage_label"] = ladder["by_seq"][achieved_seq]["label"]

		bucket_name, bucket_fill, bucket_font = _bucket_for_seq(achieved_seq, bounds)
		row["_bucket_name"] = bucket_name
		row["_bucket_fill"] = bucket_fill
		row["_bucket_font"] = bucket_font

		eta = row.get("eta_ata")
		row["_days_to_eta"] = (
			(getdate(eta) - today).days if eta and achieved_seq < discharge_seq else None
		)

		if bucket_name == "AT PORT / IN CLEARANCE":
			disch_dates = [
				getdate(d) for d in (row.get("discharge_date"), row.get("container_discharge_date")) if d
			]
			row["_days_at_port"] = (today - max(disch_dates)).days if disch_dates else None
		else:
			row["_days_at_port"] = None


def _master_sheet_banner(ws, ncols, subtitle):
	"""Company name + subtitle banner, reused by the At a Glance and Detailed
	Tracking sheets (the Summary sheet has its own richer title block).
	Returns the first free row index for the sheet's own content (banner
	occupies rows 1-2, row 3 left blank)."""
	company = frappe.defaults.get_global_default("company") or frappe.defaults.get_user_default("Company") or "FreightMas"
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
	ws.cell(row=1, column=1, value=company).font = _TITLE_FONT
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
	ws.cell(row=2, column=1, value=subtitle).font = _SUBTITLE_FONT
	ws.sheet_view.showGridLines = False
	return 4


def _apply_stage_fill(cell, row):
	cell.fill = PatternFill("solid", fgColor=row["_bucket_fill"])
	cell.font = Font(color=row["_bucket_font"], bold=True)
	cell.alignment = Alignment(horizontal="left", vertical="center")


_SUMMARY_CARD_BORDER = Border(
	left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
	top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)


def _write_summary_section_banner(ws, row_idx, title, ncols):
	cell = ws.cell(row=row_idx, column=1, value=title)
	ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
	cell.font = Font(bold=True, color="FFFFFF", size=11)
	cell.fill = _HEADER_FILL
	cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
	ws.row_dimensions[row_idx].height = 20
	return row_idx + 1


def _write_summary_table_header(ws, row_idx, label_text, ncols):
	header_fill = PatternFill("solid", fgColor="F2F2F2")
	bottom_border = Border(bottom=Side(style="thin", color="BFBFBF"))
	ws.cell(row=row_idx, column=1, value=label_text).font = Font(bold=True, color="404040")
	lines_cell = ws.cell(row=row_idx, column=4, value="Lines")
	lines_cell.font = Font(bold=True, color="404040")
	lines_cell.alignment = _RIGHT
	share_cell = ws.cell(row=row_idx, column=5, value="Share")
	share_cell.font = Font(bold=True, color="404040")
	share_cell.alignment = Alignment(horizontal="center")
	for c in range(1, ncols + 1):
		cell = ws.cell(row=row_idx, column=c)
		cell.fill = header_fill
		cell.border = bottom_border
	return row_idx + 1


def _write_summary_data_row(ws, row_idx, label, count, total, zebra, alert=False):
	"""Writes label (column A, left to overflow visually across the blank
	B/C spacer columns if long - no merge needed) + Lines (D) + Share (E).
	`alert=True` colors the count/share red when >0, green when 0 - used for
	the Requiring Attention table."""
	ws.cell(row=row_idx, column=1, value=label)
	count_cell = ws.cell(row=row_idx, column=4, value=count)
	count_cell.alignment = _RIGHT
	share_cell = ws.cell(row=row_idx, column=5, value=(count / total) if total else 0)
	share_cell.number_format = "0%"
	share_cell.alignment = Alignment(horizontal="center")
	if alert:
		alert_fill = PatternFill("solid", fgColor="FBE7E7" if count else "EAF3E6")
		count_cell.fill = alert_fill
		count_cell.font = Font(bold=True, color="9C0006" if count else "3D6B29")
		share_cell.fill = alert_fill
	elif zebra:
		for c in (1, 4, 5):
			ws.cell(row=row_idx, column=c).fill = _ZEBRA_FILL
	return row_idx + 1


def _write_master_summary_sheet(ws, report_rows, ladder, total_jobs):
	"""Cover-page style summary: big title block, 5 colored KPI "cards" (one
	per stage bucket - count + share, bordered like tiles), then two banded
	tables (Consignments by Stage, Requiring Attention) with zebra striping
	and traffic-light coloring on the attention counts."""
	ws.title = "Summary"
	ncols = 5
	today = getdate(nowdate())
	ws.sheet_view.showGridLines = False

	company = frappe.defaults.get_global_default("company") or frappe.defaults.get_user_default("Company") or "FreightMas"
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
	ws.cell(row=1, column=1, value=company).font = Font(bold=True, size=18, color="17406B")
	ws.row_dimensions[1].height = 26

	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
	ws.cell(row=2, column=1, value="Master Tracking Report").font = Font(bold=True, size=13, color="178A8A")
	ws.row_dimensions[2].height = 20

	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
	ws.cell(
		row=3, column=1,
		value=(
			f"Position as at {formatdate(today, 'dd MMMM yyyy')}  |  "
			f"{len(report_rows)} consignment lines across {total_jobs} forwarding jobs"
		),
	).font = Font(size=10, italic=True, color="7F7F7F")

	bounds = _stage_bucket_bounds(ladder)
	bucket_counts = Counter(r["_bucket_name"] for r in report_rows)
	total = len(report_rows) or 1

	label_row, value_row, share_row = 5, 6, 7
	ws.row_dimensions[label_row].height = 28
	ws.row_dimensions[value_row].height = 34
	for col_idx, (name, fill, font_color, _start, _end) in enumerate(bounds, start=1):
		count = bucket_counts.get(name, 0)
		for r, value, size, bold in (
			(label_row, name, 9, True),
			(value_row, count, 20, True),
			(share_row, count / total, 9, False),
		):
			cell = ws.cell(row=r, column=col_idx, value=value)
			cell.fill = PatternFill("solid", fgColor=fill)
			cell.font = Font(bold=bold, size=size, color=font_color)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(r == label_row))
			cell.border = _SUMMARY_CARD_BORDER
		ws.cell(row=share_row, column=col_idx).number_format = "0%"

	row_idx = share_row + 2

	row_idx = _write_summary_section_banner(ws, row_idx, "CONSIGNMENTS BY STAGE", ncols)
	row_idx = _write_summary_table_header(ws, row_idx, "Stage", ncols)
	stage_counts = Counter(r["_stage_label"] for r in report_rows)
	zebra = False
	for rung in ladder["rungs"]:
		count = stage_counts.get(rung["label"], 0)
		if not count:
			continue
		row_idx = _write_summary_data_row(ws, row_idx, rung["label"], count, total, zebra)
		zebra = not zebra
	total_row = row_idx
	row_idx = _write_summary_data_row(ws, row_idx, "Total", len(report_rows), total, False)
	for c in (1, 4, 5):
		cell = ws.cell(row=total_row, column=c)
		cell.font = Font(bold=True)
		cell.border = Border(top=Side(style="thin", color="808080"))
	row_idx += 1

	row_idx = _write_summary_section_banner(ws, row_idx, "REQUIRING ATTENTION", ncols)
	row_idx = _write_summary_table_header(ws, row_idx, "Metric", ncols)
	attention_rows = [
		("Containers standing at port more than 14 days without gate-out",
			sum(1 for r in report_rows if _pred_at_port_over_threshold(r))),
		("Consignments where the ETA has passed with no discharge recorded",
			sum(1 for r in report_rows if _pred_eta_passed_no_discharge(r, today))),
		("Consignments still awaiting departure from origin",
			sum(1 for r in report_rows if _pred_awaiting_departure(r))),
		("Consignments arriving within the next 14 days",
			sum(1 for r in report_rows if _pred_arriving_soon(r))),
	]
	for label, count in attention_rows:
		row_idx = _write_summary_data_row(ws, row_idx, label, count, total, False, alert=True)

	ws.column_dimensions["A"].width = 26
	ws.column_dimensions["B"].width = 20
	ws.column_dimensions["C"].width = 20
	ws.column_dimensions["D"].width = 13
	ws.column_dimensions["E"].width = 13


AT_A_GLANCE_HEADERS = [
	"Job ID", "Consignee", "Reference", "BL / AWB", "Container", "Parcels", "Type",
	"ETA / ATA", "Current Stage", "Progress", "Days to ETA", "Days at Port",
	"Next Step", "Remark",
]


def _write_at_a_glance_row(ws, row_idx, values, source_row, is_parent):
	for col_idx, value in enumerate(values, start=1):
		cell = ws.cell(row=row_idx, column=col_idx)
		if col_idx == 8:  # ETA / ATA
			cell.value = formatdate(value, "dd-MMM-yy") if value else ""
		elif col_idx == 10:  # Progress
			cell.value = value or 0
			cell.number_format = "0%"
			cell.alignment = Alignment(horizontal="center")
		else:
			cell.value = value if value not in (None,) else ""
		if is_parent:
			cell.font = Font(bold=True)

	_apply_stage_fill(ws.cell(row=row_idx, column=9), source_row)

	if _pred_at_port_over_threshold(source_row):
		port_cell = ws.cell(row=row_idx, column=12)
		port_cell.fill = PatternFill("solid", fgColor="FFF4E0")
		port_cell.font = Font(color="8A5200", bold=True)

	days_to_eta = source_row.get("_days_to_eta")
	if days_to_eta is not None and days_to_eta < 0:
		eta_cell = ws.cell(row=row_idx, column=11)
		eta_cell.fill = PatternFill("solid", fgColor="FBE7E7")
		eta_cell.font = Font(color="9C0006", bold=True)


def _write_at_a_glance_sheet(ws, report_rows, containers_by_job, ladder):
	ws.title = "At a Glance"
	ncols = len(AT_A_GLANCE_HEADERS)
	subtitle = "Master Tracking Report — At a Glance (one line per BL / AWB - expand for containers)"
	row_idx = _master_sheet_banner(ws, ncols, subtitle)

	header_row = row_idx
	for col_idx, label in enumerate(AT_A_GLANCE_HEADERS, start=1):
		cell = ws.cell(row=header_row, column=col_idx, value=label)
		cell.font = _HEADER_FONT
		cell.fill = _HEADER_FILL
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	row_idx = header_row + 1
	ws.freeze_panes = f"A{row_idx}"

	max_seq = ladder["rungs"][-1]["seq"]

	def next_step_label(seq):
		if seq >= max_seq:
			return "Completed"
		return "Next: " + ladder["by_seq"][seq + 1]["label"]

	seen_jobs = set()
	for row in report_rows:
		job_id = row["job_id"]
		if job_id in seen_jobs:
			continue
		seen_jobs.add(job_id)
		group = containers_by_job[job_id]

		lead = min(group, key=lambda r: r["_stage_seq"])
		avg_progress = sum(r["percent_complete"] for r in group) / len(group)

		parent_values = [
			job_id, row["consignee"], row["customer_reference"], row["bl_number"],
			"", row["parcel_count"], "",
			lead.get("eta_ata"), lead["_stage_label"], avg_progress,
			lead.get("_days_to_eta"), lead.get("_days_at_port"),
			next_step_label(lead["_stage_seq"]), lead.get("status_comment") or "",
		]
		_write_at_a_glance_row(ws, row_idx, parent_values, lead, is_parent=True)
		row_idx += 1

		for container_row in group:
			child_values = [
				"", "", "", "",
				container_row.get("container_number") or "(no container)", "",
				container_row.get("container_type") or "",
				container_row.get("eta_ata"), container_row["_stage_label"], container_row["percent_complete"],
				container_row.get("_days_to_eta"), container_row.get("_days_at_port"),
				next_step_label(container_row["_stage_seq"]), container_row.get("status_comment") or "",
			]
			_write_at_a_glance_row(ws, row_idx, child_values, container_row, is_parent=False)
			ws.row_dimensions[row_idx].outlineLevel = 1
			ws.row_dimensions[row_idx].hidden = True
			row_idx += 1

	ws.sheet_properties.outlinePr.summaryBelow = False

	widths = [14, 26, 18, 20, 16, 9, 8, 12, 24, 10, 10, 10, 26, 30]
	for col_idx, width in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_master_detailed_sheet(ws, sections, report_rows, ladder):
	ws.title = "Detailed Tracking"
	section_by_title = {title: cols for title, _color, cols in sections}

	blocks = [
		("Shipment Details", [
			("Job ID", "job_id", "text"), ("Consignee", "consignee", "text"),
			("Reference", "customer_reference", "text"), ("BL / AWB", "bl_number", "text"),
			("Container", "container_number", "text"), ("Type", "container_type", "text"),
		]),
		("Sea / Air Freight", [
			("Departed Origin (ATD)", "atd", "date"), ("ETA / ATA", "eta_ata", "date"),
			("Discharged (Shipment)", "discharge_date", "date"),
			("Container Discharged", "container_discharge_date", "date"),
			("Gate Out", "gate_out_date", "date"), ("Empty Returned", "empty_return_date", "date"),
		]),
		("Port Clearance", [
			(label, fieldname, "date") for label, fieldname, _kind in section_by_title.get("Port Clearance", [])
		]),
		("Road Transport & Delivery", [
			("Booked", "booked_on_date", "date"), ("Loaded", "loaded_on_date", "date"),
			("Offloaded", "offloaded_on_date", "date"), ("Returned", "returned_on_date", "date"),
			("Completed", "trucking_completed_on_date", "date"),
		]),
		("Border Clearance", [
			(label, fieldname, "date") for label, fieldname, _kind in section_by_title.get("Border Clearance", [])
		]),
		("Overview", [("Completed", "job_completed_on", "date")]),
		("Progress & Status", [
			("Milestones", "_milestones_achieved", "int"), ("Progress", "percent_complete", "percent"),
			("Days to ETA", "_days_to_eta", "int"), ("Days at Port", "_days_at_port", "int"),
			("Current Stage", "_stage_label", "stage"), ("Remark", "status_comment", "text"),
			("Stage No.", "_stage_seq", "stageno"),
		]),
	]

	ncols = sum(len(cols) for _title, cols in blocks)
	subtitle = f"Master Tracking Report — Detailed Tracking ({len(report_rows)} lines)"
	band_row = _master_sheet_banner(ws, ncols, subtitle)
	header_row = band_row + 1

	col = 1
	columns = []  # flattened (label, fieldname, kind)
	for title, cols in blocks:
		start = col
		for label, fieldname, kind in cols:
			columns.append((label, fieldname, kind))
			col += 1
		end = col - 1
		band_cell = ws.cell(row=band_row, column=start, value=title.upper())
		band_cell.fill = PatternFill("solid", fgColor=MASTER_DETAIL_BAND_COLORS.get(title, "17406B"))
		band_cell.font = Font(bold=True, color="FFFFFF")
		band_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		if end > start:
			ws.merge_cells(start_row=band_row, start_column=start, end_row=band_row, end_column=end)

	for col_idx, (label, _fieldname, _kind) in enumerate(columns, start=1):
		cell = ws.cell(row=header_row, column=col_idx, value=label)
		cell.font = _HEADER_FONT
		cell.fill = _HEADER_FILL
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

	data_start = header_row + 1
	for offset, row in enumerate(report_rows):
		row_idx = data_start + offset
		row["_milestones_achieved"] = sum(
			1 for r in ladder["rungs"][1:]
			if row.get(r["field"]) not in (None, "", TRACKING_NOT_APPLICABLE)
		)
		for col_idx, (_label, fieldname, kind) in enumerate(columns, start=1):
			value = row.get(fieldname)
			cell = ws.cell(row=row_idx, column=col_idx)
			if kind == "date":
				if value is TRACKING_NOT_APPLICABLE:
					cell.value = "-"
				else:
					cell.value = formatdate(value, "dd-MMM-yy") if value else ""
				cell.fill = _milestone_cell_fill(value)
				cell.alignment = Alignment(horizontal="center")
			elif kind == "percent":
				cell.value = value or 0
				cell.number_format = "0%"
				cell.alignment = Alignment(horizontal="center")
			elif kind in ("int", "stageno"):
				cell.value = value if value is not None else ""
				cell.alignment = Alignment(horizontal="center")
				if kind == "stageno":
					_apply_stage_fill(cell, row)
			elif kind == "stage":
				cell.value = value or ""
				_apply_stage_fill(cell, row)
			else:  # text
				cell.value = value or ""
				cell.alignment = Alignment(horizontal="left")

	ws.freeze_panes = f"A{data_start}"
	for col_idx, (_label, _fieldname, kind) in enumerate(columns, start=1):
		if kind == "stage":
			ws.column_dimensions[get_column_letter(col_idx)].width = 26
		elif kind == "text":
			ws.column_dimensions[get_column_letter(col_idx)].width = 20
		else:
			ws.column_dimensions[get_column_letter(col_idx)].width = 12


def _build_master_tracking_workbook(sections, report_rows, containers_by_job, ladder):
	wb = openpyxl.Workbook()

	ws_summary = wb.active
	_write_master_summary_sheet(ws_summary, report_rows, ladder, len(containers_by_job))

	_write_at_a_glance_sheet(wb.create_sheet("At a Glance"), report_rows, containers_by_job, ladder)
	_write_master_detailed_sheet(wb.create_sheet("Detailed Tracking"), sections, report_rows, ladder)

	return wb


@frappe.whitelist()
def export_master_tracking_report(customers=None, status=None, direction=None, search=None):
	check_freightmas_role()
	sections, report_rows, containers_by_job = _gather_tracking_report_rows(
		customers=customers, status=status, direction=direction, search=search
	)
	ladder = _build_stage_ladder(sections)
	_annotate_stage_fields(report_rows, ladder)
	wb = _build_master_tracking_workbook(sections, report_rows, containers_by_job, ladder)
	_send_workbook(wb, _timestamped("Master_Tracking_Report"))


# ============================================================
# SHIPMENT TRACKING REPORT (per-customer PDF dossier)
# ============================================================
# A single-customer PDF export of the same compact, color-coded dossier
# layout used for the per-job tracking view, covering all of one customer's
# shipments in one document. Renders a Jinja template to PDF via
# frappe.utils.pdf.get_pdf (the report_pdf_template.html pattern), reusing
# the milestone/cargo assembly get_job_detail already produces
# (_build_job_milestone_stages / _build_job_cargo_list) rather than
# re-deriving it. Each job reduces to one color-coded block:
#   orange = in progress, green = completed/delivered, red = delayed
# (overdue - ETA/ETD passed with no actual arrival/departure recorded).

# status_key -> the exact hex values the dossier template paints with, from
# the approved mock (Shipment Tracking Report.dc.html).
DOSSIER_STATUS_COLORS = {
	"orange": "#dd6b20",
	"green": "#1f8a4a",
	"red": "#c0392b",
	"gray": "#98a1ac",
}


def _dossier_status_key(doc, today):
	"""One of orange/green/red for a job's color band. Completed/Closed/
	Delivered are green; an import past its ETA with no ATA (or an export past
	its ETD with no ATD) is a red 'delayed'; everything else in progress is
	orange."""
	if doc.status in ("Completed", "Closed", "Delivered"):
		return "green"
	is_overdue = (
		(doc.direction == "Import" and doc.eta and getdate(doc.eta) < today and not doc.ata)
		or (doc.direction == "Export" and doc.etd and getdate(doc.etd) < today and not doc.atd)
	)
	return "red" if is_overdue else "orange"


def _dossier_date(value, fmt="MMM d, yyyy"):
	return formatdate(value, fmt) if value else None


def _dossier_fraction(done, total):
	"""A '{done}/{total} · {pct}%' progress chip plus the color it prints
	in: green when complete, orange when partial, gray when nothing's done."""
	pct = round(done / total * 100) if total else 0
	if total and done >= total:
		color = "green"
	elif done > 0:
		color = "orange"
	else:
		color = "gray"
	return {
		"done": done,
		"total": total,
		"pct": pct,
		"color": DOSSIER_STATUS_COLORS[color],
		"text": f"{done}/{total} · {pct}%",
	}


def _dossier_stage_row(stage, is_first=False):
	"""One row of a Port/Border Clearance stage-summary card: the stage name,
	its '· {pct}%' chip (colored like _dossier_fraction, but percentage-only -
	stage rows don't show raw done/total counts), whether it's the stage
	currently in progress (rendered emphasized), and - only for the first
	stage by sequence - the labels of its outstanding milestones."""
	frac = _dossier_fraction(stage["done"], stage["total"])
	return {
		"name": stage["name"],
		"text": f"· {frac['pct']}%",
		"color": frac["color"],
		"is_current": stage["is_current"],
		"is_first": is_first,
		"missing": stage.get("missing") or [] if is_first else [],
	}


def _dossier_latest_line(doc, status_key):
	"""The 'LATEST: ...' status line on a job block. Prefers the job's own
	current_comment (the same free-text note the Command Centre shows); falls
	back to a status/ETA-derived line when there isn't one."""
	if doc.current_comment:
		return doc.current_comment
	if status_key == "green":
		return "Delivered · Job closed" if doc.status in ("Completed", "Closed") else "Delivered"
	eta = _dossier_date(doc.eta, "dd-MMM-yy")
	if status_key == "red":
		return f"Delayed · Revised ETA {eta}" if eta else "Delayed"
	return f"In transit · ETA {eta}" if eta else "In progress"


# status_key -> the label shown in the "Shipments at a Glance" status cell.
DOSSIER_STATUS_LABELS = {
	"orange": "In Progress",
	"green": "Completed",
	"red": "Delayed",
	"gray": "In Progress",
}


def _build_job_dossier_context(job_name):
	"""Reshapes one Forwarding Job into the dossier block the PDF template
	renders: ref, status color/line, an at-a-glance summary row, a 5-field
	info bar (BL Number, Reference, Cargo, Mode, Last Update) and 4 section
	cards (Sea/Air Freight + Road Transport on the left, Port Clearance +
	Completion on the right)."""
	doc = frappe.get_doc("Forwarding Job", job_name)
	today = getdate(nowdate())

	cargo = _build_job_cargo_list(doc)
	groups_by_name = {s["group"]: s for s in _build_job_milestone_stages(doc)}
	status_key = _dossier_status_key(doc, today)

	# Container-count summary (e.g. "3×40HC") for the at-a-glance Cargo column -
	# distinct from the info bar's Cargo field, which is the description only.
	type_counts = Counter(c["container_type"] for c in cargo if c.get("container_type"))
	cargo_units = " + ".join(f"{n}×{t}" for t, n in type_counts.items()) or "—"
	mode_line = " · ".join(p for p in [doc.direction, doc.shipment_mode] if p) or "—"

	# --- Sea / Air Freight card (container milestone dates) ---
	sea_rows = []
	sea_done = sea_total = 0
	for c in cargo:
		checks = [c.get("discharge_date"), c.get("gate_out_date"), c.get("empty_return_date")]
		sea_total += 3
		sea_done += sum(1 for v in checks if v)
		sea_rows.append({
			"container": c.get("container_number") or "—",
			"type": c.get("container_type") or "—",
			"discharge": _dossier_date(c.get("discharge_date"), "MMM d") or "×",
			"gate_out": _dossier_date(c.get("gate_out_date"), "MMM d") or "×",
			"ret": _dossier_date(c.get("empty_return_date"), "MMM d") or "×",
			"status": c.get("api_container_status")
				or ("Delivered" if c.get("is_completed") else "In Transit"),
		})
	sea_section = {
		"kind": "containers", "column": "left", "title": "Sea / Air Freight",
		"frac": _dossier_fraction(sea_done, sea_total),
		"atd": _dossier_date(doc.atd), "ata": _dossier_date(doc.ata),
		"headers": ["Container", "Type", "Disch.", "Gate Out", "Ret.", "Status"],
		"rows": [
			[r["container"], r["type"], r["discharge"], r["gate_out"], r["ret"], r["status"]]
			for r in sea_rows
		],
	}

	# --- Road Transport card (per-container trucking dates) ---
	road_rows = []
	road_done = road_total = 0
	for c in cargo:
		returned = c.get("is_returned") or (c.get("to_be_returned") and c.get("empty_return_date"))
		flags = [c.get("is_booked"), c.get("is_loaded"), c.get("is_offloaded"),
				 returned, c.get("is_completed")]
		dates = [c.get("booked_on_date"), c.get("loaded_on_date"), c.get("offloaded_on_date"),
				 c.get("returned_on_date"), c.get("completed_on_date")]
		road_total += 5
		road_done += sum(1 for v in flags if v)
		road_rows.append([
			c.get("container_number") or "—",
			*[_dossier_date(d, "MMM d") or "×" for d in dates],
		])
	road_section = {
		"kind": "containers", "column": "left", "title": "Road Transport",
		"frac": _dossier_fraction(road_done, road_total),
		"headers": ["Container", "Booked", "Loaded", "Offloaded", "Returned", "Comp."],
		"rows": road_rows,
	}

	# --- Port Clearance card ---
	# Always a compact stage rollup (Documentation / Shipping Line / Customs /
	# Port Release) with the first stage's outstanding items called out, plus
	# an overall progress row - clearer for clients than the full 16-item
	# checklist. Falls back to the full checklist only if the milestones
	# somehow carry no stage data (shouldn't happen; Milestone Definition
	# assigns a stage to every Port Clearance item).
	pc_group = groups_by_name.get("Port Clearance", {})
	pc_milestones = pc_group.get("milestones", [])
	pc_done = sum(1 for m in pc_milestones if m["is_completed"])
	pc_frac = _dossier_fraction(pc_done, len(pc_milestones))
	# Percentage-only display for the Port Clearance card - raw done/total
	# counts are dropped here (unlike Sea/Air Freight, Road Transport and
	# Completion, which keep the "X/Y · Z%" format).
	pc_frac_pct = dict(pc_frac, text=f"· {pc_frac['pct']}%")
	if pc_group.get("has_stages"):
		port_section = {
			"kind": "stages", "column": "right", "title": "Port Clearance",
			"frac": pc_frac_pct,
			"stages": [
				_dossier_stage_row(s, is_first=(i == 0))
				for i, s in enumerate(pc_group.get("stages", []))
			],
		}
	else:
		port_section = {
			"kind": "checklist", "column": "right", "title": "Port Clearance",
			"frac": pc_frac_pct,
			"entries": [
				{
					"label": m["label"],
					"done": m["is_completed"],
					"value": (_dossier_date(m["completed_on"], "MMM d") or "Done")
						if m["is_completed"] else "Pend",
				}
				for m in pc_milestones
			],
		}

	# --- Completion card (single "Completed" row - no Rev. Recognised) ---
	completed = doc.status in ("Completed", "Closed") or bool(doc.completed_on)
	completion_section = {
		"kind": "checklist", "column": "right", "title": "Completion",
		"frac": _dossier_fraction(1 if completed else 0, 1),
		"entries": [{
			"label": "Completed",
			"done": completed,
			"value": (_dossier_date(doc.completed_on) or "Done") if completed else "Pending",
		}],
	}

	# Overall completion across every section - drives the at-a-glance
	# Status cell's percentage (the same done/total roll-up the job 3 mock
	# uses: 6+2 of 12+15+15+2 = 18%).
	overall_done = sea_done + road_done + pc_done + (1 if completed else 0)
	overall_total = sea_total + road_total + len(pc_milestones) + 1
	overall = _dossier_fraction(overall_done, overall_total)

	latest_line = _dossier_latest_line(doc, status_key)

	# Flat per-container rows for the Excel Container Detail sheet - the same
	# cargo the PDF's per-container tables draw from, unpivoted for filtering.
	containers = [
		{
			"container": c.get("container_number") or "—",
			"type": c.get("container_type") or "",
			"booked": bool(c.get("is_booked")),
			"loaded": bool(c.get("is_loaded")),
			"discharged": bool(c.get("discharge_date")),
			"offloaded": bool(c.get("is_offloaded")),
			"returned": bool(c.get("is_returned") or (c.get("to_be_returned") and c.get("empty_return_date"))),
			"completed": bool(c.get("is_completed")),
			"api_status": c.get("api_container_status") or "",
		}
		for c in cargo
	]

	return {
		"ref": doc.name,
		"status_key": status_key,
		"status_color": DOSSIER_STATUS_COLORS[status_key],
		"status_label": DOSSIER_STATUS_LABELS[status_key],
		"progress_percent": overall["pct"],
		"latest_line": latest_line,
		"eta": doc.eta,
		"cargo_count": doc.cargo_count,
		# One row of the "Shipments at a Glance" table.
		"glance": {
			"ref": doc.name,
			"bl_no": doc.bl_number or "—",
			"cargo_units": cargo_units,
			"latest_update": latest_line,
			"status_label": DOSSIER_STATUS_LABELS[status_key],
			"status_pct": overall["pct"],
			"status_color": DOSSIER_STATUS_COLORS[status_key],
		},
		# The 5-column info bar on the detail block.
		"fields": {
			"bl_number": doc.bl_number or "—",
			"reference": doc.customer_reference or "—",
			"cargo": doc.cargo_description or "—",
			"mode": mode_line,
			"last_update": _dossier_date(doc.last_updated_on) or "—",
		},
		"sections": [sea_section, road_section, port_section, completion_section],
		"containers": containers,
	}


def _summarize_statuses(dossier_jobs):
	"""Counts for the banner strip: total shipments and how many are in
	transit / completed / delayed, plus a ready-to-print summary line."""
	counts = Counter(j["status_key"] for j in dossier_jobs)
	total = len(dossier_jobs)
	parts = [f"{total} SHIPMENT" + ("S" if total != 1 else "")]
	if counts.get("orange"):
		parts.append(f"{counts['orange']} IN TRANSIT")
	if counts.get("green"):
		parts.append(f"{counts['green']} COMPLETED")
	if counts.get("red"):
		parts.append(f"{counts['red']} DELAYED")
	return {
		"total": total,
		"in_transit": counts.get("orange", 0),
		"completed": counts.get("green", 0),
		"delayed": counts.get("red", 0),
		"line": " · ".join(parts),
	}


def _company_logo_data_uri(company):
	"""Best-effort base64 data URI for the Company's logo, embedded inline so
	the PDF renderer never has to fetch it over the network (which is flaky
	during server-side PDF generation). Returns None if there's no logo or it
	can't be read - the template then falls back to a text monogram."""
	logo = frappe.db.get_value("Company", company, "company_logo")
	if not logo or not isinstance(logo, str) or not logo.startswith("/"):
		return None
	try:
		import base64
		import os

		is_private = "/private/" in logo
		relative = logo.split("/files/", 1)[-1]
		file_path = frappe.utils.get_files_path(relative, is_private=is_private)
		if not os.path.exists(file_path):
			return None
		with open(file_path, "rb") as fh:
			encoded = base64.b64encode(fh.read()).decode()
		ext = logo.rsplit(".", 1)[-1].lower()
		mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
				"gif": "image/gif", "svg": "image/svg+xml"}.get(ext, "image/png")
		return f"data:{mime};base64,{encoded}"
	except Exception:
		return None


def _dossier_jobs_for_customer(customer, numbered=False):
	"""Shared by export_shipment_tracking_report(_excel) and their email
	counterparts: fetches this customer's open Forwarding Jobs and builds
	each one's dossier context. numbered=True stamps ctx["num"] for the
	PDF's at-a-glance table (not needed by the Excel workbook)."""
	jobs = frappe.get_all(
		"Forwarding Job",
		filters={
			"customer": customer,
			"docstatus": ["in", [0, 1]],
			"status": ["in", ["Draft", "In Progress", "Delivered"]],
		},
		fields=["name"],
		order_by="status asc, eta asc",
	)
	dossier_jobs = []
	for i, j in enumerate(jobs, start=1):
		ctx = _build_job_dossier_context(j.name)
		if numbered:
			ctx["num"] = i  # cross-references the at-a-glance table's "#" column
		dossier_jobs.append(ctx)
	return dossier_jobs


def _build_shipment_tracking_pdf(customer):
	"""Builds the compact, color-coded Shipment Tracking dossier (one block
	per Forwarding Job) as PDF bytes, scoped to a single customer. Shared by
	export_shipment_tracking_report (download) and email_shipment_tracking_report
	(email attachment) - identical output either way."""
	if not customer:
		frappe.throw(_("Select a customer to generate this report."))

	company = (
		frappe.defaults.get_user_default("Company")
		or frappe.defaults.get_global_default("company")
		or "FreightMas"
	)

	dossier_jobs = _dossier_jobs_for_customer(customer, numbered=True)

	generated_on = frappe.utils.now_datetime().strftime("%d %b %Y")
	customer_name = frappe.db.get_value("Customer", customer, "customer_name") or customer
	company_name = frappe.db.get_value("Company", company, "company_name") or company

	html = frappe.render_template(
		"freightmas/templates/shipment_tracking_report.html",
		{
			"company": company_name,
			"customer": customer_name,
			"logo": _company_logo_data_uri(company),
			"jobs": dossier_jobs,
			"generated_on": generated_on,
			"summary": _summarize_statuses(dossier_jobs),
		},
	)

	from frappe.utils.pdf import get_pdf

	# footer-left / footer-right are wkhtmltopdf running-footer options;
	# [page]/[topage] are its built-in per-page counters (CSS can't compute
	# real page totals), giving genuine "Page X of Y" numbering.
	pdf = get_pdf(
		html,
		options={
			"orientation": "Portrait",
			"page-size": "A4",
			"margin-top": "12mm",
			"margin-bottom": "16mm",
			"margin-left": "10mm",
			"margin-right": "10mm",
			"footer-left": f"Prepared for {customer_name} — Confidential",
			"footer-right": "Page [page] of [topage]",
			"footer-font-size": "8",
			"footer-spacing": "4",
		},
	)

	return pdf, customer_name


@frappe.whitelist()
def export_shipment_tracking_report(customer):
	"""Renders the compact, color-coded Shipment Tracking dossier (one block
	per Forwarding Job) as a PDF, scoped to a single customer."""
	check_freightmas_role()

	pdf, _customer_name = _build_shipment_tracking_pdf(customer)

	frappe.local.response.filename = f"Shipment-Tracking-{frappe.scrub(customer)}.pdf"
	frappe.local.response.filecontent = pdf
	frappe.local.response.type = "download"


# ============================================================
# SHIPMENT TRACKING REPORT (per-customer Excel workbook)
# ============================================================
# Companion to the PDF dossier: the same single-customer data, reshaped into
# a filterable/pivotable 3-sheet workbook (Summary, Milestone Detail,
# Container Detail). Reuses _build_job_dossier_context() as the single source
# of per-job data - no milestone/container logic is re-derived here.

DOSSIER_XL_PURPLE = "2C2560"
DOSSIER_XL_PURPLE_LIGHT = "4B3F8F"
DOSSIER_XL_STATUS_FILLS = {
	"In Progress": ("FDEEE0", "9A4A10"),
	"Completed": ("E5F3EA", "186429"),
	"Delayed": ("FCE9E8", "A3241F"),
}


def _xl_style_header_row(ws, row, ncols):
	"""Purple header row (light purple fill, white bold text) used on all
	three sheets."""
	for col in range(1, ncols + 1):
		cell = ws.cell(row=row, column=col)
		cell.font = Font(bold=True, color="FFFFFF", size=10.5)
		cell.fill = PatternFill("solid", fgColor=DOSSIER_XL_PURPLE_LIGHT)
		cell.alignment = Alignment(horizontal="left", vertical="center")


def _build_shipment_tracking_workbook(customer, jobs):
	from openpyxl.formatting.rule import DataBarRule

	wb = openpyxl.Workbook()

	# ---- Sheet 1: Summary ----
	ws = wb.active
	ws.title = "Summary"
	ws.sheet_properties.tabColor = DOSSIER_XL_PURPLE

	ncols = 11
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
	title_cell = ws.cell(row=1, column=1, value=f"SHIPMENT TRACKING REPORT — {customer.upper()}")
	title_cell.font = Font(bold=True, size=13, color="FFFFFF")
	title_cell.fill = PatternFill("solid", fgColor=DOSSIER_XL_PURPLE)
	title_cell.alignment = Alignment(vertical="center", indent=1)
	ws.row_dimensions[1].height = 26

	headers = ["#", "Job Ref", "BL Number", "Reference", "Cargo", "Qty",
			   "Mode", "ETA", "Status", "Progress %", "Latest Update"]
	for col, label in enumerate(headers, start=1):
		ws.cell(row=2, column=col, value=label)
	_xl_style_header_row(ws, 2, ncols)
	for col, label in enumerate(headers, start=1):
		if label in ("#", "Progress %"):
			ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")

	for i, job in enumerate(jobs, start=1):
		row = 2 + i
		values = [
			i, job["ref"], job["fields"]["bl_number"], job["fields"]["reference"],
			job["fields"]["cargo"], job.get("cargo_count"), job["fields"]["mode"],
			job.get("eta"), job["status_label"], job["progress_percent"],
			job["glance"]["latest_update"],
		]
		for col, val in enumerate(values, start=1):
			ws.cell(row=row, column=col, value=val)

		ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
		eta_cell = ws.cell(row=row, column=8)
		if job.get("eta"):
			eta_cell.number_format = "DD-MMM-YY"
		progress_cell = ws.cell(row=row, column=10)
		progress_cell.alignment = Alignment(horizontal="center")

		status_cell = ws.cell(row=row, column=9)
		fill, text = DOSSIER_XL_STATUS_FILLS.get(job["status_label"], ("F3F4F6", "5F6368"))
		status_cell.fill = PatternFill("solid", fgColor=fill)
		status_cell.font = Font(bold=True, color=text)
		status_cell.alignment = Alignment(horizontal="center")

		# Job Ref as a real hyperlink back into Desk.
		ref_cell = ws.cell(row=row, column=2)
		ref_cell.hyperlink = frappe.utils.get_url(f"/app/forwarding-job/{job['ref']}")
		ref_cell.font = Font(color="1155CC", underline="single")

	last_row = 2 + len(jobs)

	if jobs:
		# Native Excel data bar on the raw 0-100 progress number.
		ws.conditional_formatting.add(
			f"J3:J{last_row}",
			DataBarRule(
				start_type="num", start_value=0, end_type="num", end_value=100,
				color="F6C199", showValue=True,
			),
		)

	widths = [4, 16, 15, 15, 20, 9, 13, 12, 13, 11, 34]
	for col, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(col)].width = w

	ws.freeze_panes = "C3"
	ws.auto_filter.ref = f"A2:K{max(last_row, 2)}"

	ws.page_setup.orientation = "landscape"
	ws.page_setup.fitToWidth = 1
	ws.page_setup.fitToHeight = 0
	ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
	ws.print_title_rows = "1:2"
	ws.print_area = f"A1:K{max(last_row, 2)}"

	# ---- Sheet 2: Milestone Detail ----
	_fill_milestone_sheet(wb.create_sheet("Milestone Detail"), jobs)

	# ---- Sheet 3: Container Detail ----
	_fill_container_sheet(wb.create_sheet("Container Detail"), jobs)

	return wb


def _fill_milestone_sheet(ws, jobs):
	"""One row per (job, section) - Job Ref, Milestone Group, Completed, Total,
	Percent - the same fractions the PDF section heads show, unpivoted."""
	ws.sheet_properties.tabColor = DOSSIER_XL_PURPLE_LIGHT
	headers = ["Job Ref", "Milestone Group", "Completed", "Total", "Percent"]
	for col, label in enumerate(headers, start=1):
		ws.cell(row=1, column=col, value=label)
	_xl_style_header_row(ws, 1, len(headers))

	row = 2
	for job in jobs:
		for sec in job["sections"]:
			frac = sec["frac"]
			ws.cell(row=row, column=1, value=job["ref"])
			ws.cell(row=row, column=2, value=sec["title"])
			ws.cell(row=row, column=3, value=frac["done"])
			ws.cell(row=row, column=4, value=frac["total"])
			pct_cell = ws.cell(row=row, column=5, value=(frac["pct"] or 0) / 100)
			pct_cell.number_format = "0%"
			row += 1

	last_row = max(row - 1, 1)
	widths = [16, 22, 12, 10, 10]
	for col, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(col)].width = w
	ws.freeze_panes = "A2"
	ws.auto_filter.ref = f"A1:E{last_row}"


def _fill_container_sheet(ws, jobs):
	"""One row per (job, container) - the per-container milestone booleans the
	PDF's Sea/Air + Road tables render, as filterable TRUE/FALSE columns."""
	ws.sheet_properties.tabColor = DOSSIER_XL_PURPLE_LIGHT
	headers = ["Job Ref", "Container / Item", "Type", "Booked", "Loaded",
			   "Discharged", "Offloaded", "Returned", "Completed", "API Status"]
	for col, label in enumerate(headers, start=1):
		ws.cell(row=1, column=col, value=label)
	_xl_style_header_row(ws, 1, len(headers))

	row = 2
	for job in jobs:
		for c in job["containers"]:
			ws.cell(row=row, column=1, value=job["ref"])
			ws.cell(row=row, column=2, value=c["container"])
			ws.cell(row=row, column=3, value=c["type"])
			ws.cell(row=row, column=4, value=c["booked"])
			ws.cell(row=row, column=5, value=c["loaded"])
			ws.cell(row=row, column=6, value=c["discharged"])
			ws.cell(row=row, column=7, value=c["offloaded"])
			ws.cell(row=row, column=8, value=c["returned"])
			ws.cell(row=row, column=9, value=c["completed"])
			ws.cell(row=row, column=10, value=c["api_status"])
			row += 1

	last_row = max(row - 1, 1)
	widths = [16, 20, 10, 9, 9, 11, 11, 10, 11, 18]
	for col, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(col)].width = w
	ws.freeze_panes = "C2"
	ws.auto_filter.ref = f"A1:J{last_row}"


@frappe.whitelist()
def export_shipment_tracking_report_excel(customer):
	"""Renders the Shipment Tracking data for a single customer as a 3-sheet
	Excel workbook (Summary, Milestone Detail, Container Detail)."""
	check_freightmas_role()

	if not customer:
		frappe.throw(_("Select a customer to generate this report."))

	dossier_jobs = _dossier_jobs_for_customer(customer)

	customer_name = frappe.db.get_value("Customer", customer, "customer_name") or customer
	wb = _build_shipment_tracking_workbook(customer_name, dossier_jobs)
	_send_workbook(wb, f"Shipment-Tracking-{frappe.scrub(customer)}.xlsx")


# --- Tracking report email actions ------------------------------------------------
#
# Mirror the export_* download endpoints above 1:1 (same param shapes, same
# report-building helpers) but send the generated file as an email
# attachment via frappe.sendmail instead of writing to frappe.local.response.

_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _validate_email_address(email):
	if not email or not _EMAIL_RE.match(email):
		frappe.throw(_("Enter a valid recipient email address."))


def _parse_cc_emails(cc_emails):
	if not cc_emails:
		return []
	parts = cc_emails.split(",") if isinstance(cc_emails, str) else list(cc_emails)
	parts = [e.strip() for e in parts if e and e.strip()]
	for e in parts:
		if not _EMAIL_RE.match(e):
			frappe.throw(_("Invalid CC email address: {0}").format(e))
	return parts


def _single_customer_reference(customers):
	"""Best-effort Communication linkage - only meaningful when the report
	was scoped to exactly one customer (Full/Master allow 0..N)."""
	if customers and isinstance(customers, str):
		customers = frappe.parse_json(customers)
	if customers and len(customers) == 1:
		return {"reference_doctype": "Customer", "reference_name": customers[0]}
	return {}


@frappe.whitelist()
def get_customer_tracking_info(customer):
	"""Lightweight lookup for the Command Center's Email modal To/CC prefill.
	Bypasses standard Customer doctype permissions the same way get_customers()
	does above, since Command Center users are gated by check_freightmas_role()
	alone and may not hold standard Customer read permission."""
	check_freightmas_role()
	if not customer:
		return {}
	return (
		frappe.db.get_value(
			"Customer",
			customer,
			["customer_name", "tracking_email", "tracking_cc_emails", "tracking_email_enabled", "email_id"],
			as_dict=True,
		)
		or {}
	)


@frappe.whitelist()
def email_tracking_report(to_email, subject, message, customers=None, status=None, direction=None, search=None, cc_emails=None):
	check_freightmas_role()
	_validate_email_address(to_email)
	cc = _parse_cc_emails(cc_emails)

	sections, report_rows, _containers_by_job = _gather_tracking_report_rows(
		customers=customers, status=status, direction=direction, search=search
	)
	wb = _build_tracking_workbook(sections, report_rows)
	attachment = {"fname": _timestamped("Tracking_Report"), "fcontent": _workbook_bytes(wb)}

	frappe.sendmail(
		recipients=[to_email],
		sender=get_formatted_email(frappe.session.user),
		cc=cc,
		subject=subject,
		message=message,
		attachments=[attachment],
		**_single_customer_reference(customers),
		delayed=False,
	)
	return {"success": True, "message": f"Email sent to {to_email}"}


@frappe.whitelist()
def email_master_tracking_report(to_email, subject, message, customers=None, status=None, direction=None, search=None, cc_emails=None):
	check_freightmas_role()
	_validate_email_address(to_email)
	cc = _parse_cc_emails(cc_emails)

	sections, report_rows, containers_by_job = _gather_tracking_report_rows(
		customers=customers, status=status, direction=direction, search=search
	)
	ladder = _build_stage_ladder(sections)
	_annotate_stage_fields(report_rows, ladder)
	wb = _build_master_tracking_workbook(sections, report_rows, containers_by_job, ladder)
	attachment = {"fname": _timestamped("Master_Tracking_Report"), "fcontent": _workbook_bytes(wb)}

	frappe.sendmail(
		recipients=[to_email],
		sender=get_formatted_email(frappe.session.user),
		cc=cc,
		subject=subject,
		message=message,
		attachments=[attachment],
		**_single_customer_reference(customers),
		delayed=False,
	)
	return {"success": True, "message": f"Email sent to {to_email}"}


@frappe.whitelist()
def email_shipment_tracking_report(to_email, subject, message, customer, cc_emails=None):
	check_freightmas_role()
	_validate_email_address(to_email)
	cc = _parse_cc_emails(cc_emails)

	pdf, _customer_name = _build_shipment_tracking_pdf(customer)
	attachment = {"fname": f"Shipment-Tracking-{frappe.scrub(customer)}.pdf", "fcontent": pdf}

	frappe.sendmail(
		recipients=[to_email],
		sender=get_formatted_email(frappe.session.user),
		cc=cc,
		subject=subject,
		message=message,
		attachments=[attachment],
		reference_doctype="Customer",
		reference_name=customer,
		delayed=False,
	)
	return {"success": True, "message": f"Email sent to {to_email}"}


@frappe.whitelist()
def email_shipment_tracking_report_excel(to_email, subject, message, customer, cc_emails=None):
	check_freightmas_role()
	if not customer:
		frappe.throw(_("Select a customer to generate this report."))
	_validate_email_address(to_email)
	cc = _parse_cc_emails(cc_emails)

	dossier_jobs = _dossier_jobs_for_customer(customer)
	customer_name = frappe.db.get_value("Customer", customer, "customer_name") or customer
	wb = _build_shipment_tracking_workbook(customer_name, dossier_jobs)
	attachment = {"fname": f"Shipment-Tracking-{frappe.scrub(customer)}.xlsx", "fcontent": _workbook_bytes(wb)}

	frappe.sendmail(
		recipients=[to_email],
		sender=get_formatted_email(frappe.session.user),
		cc=cc,
		subject=subject,
		message=message,
		attachments=[attachment],
		reference_doctype="Customer",
		reference_name=customer,
		delayed=False,
	)
	return {"success": True, "message": f"Email sent to {to_email}"}
