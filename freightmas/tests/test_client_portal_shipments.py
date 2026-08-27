# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd and contributors
# For license information, please see license.txt

"""
Phase 1 cross-tenant tests for the Client Portal shipments/dashboard API.

Two customers, two Forwarding Jobs, two provisioned portal users - every
endpoint must return customer A's own data and reject any attempt to read
customer B's, whether by an explicit docname or via an unfiltered list.

Also includes a regression guard: the internal (Desk-facing) shipment
dashboard must still return unfiltered, cross-customer data after the
portal's changes land - proving Phase 1 never touched internal access.

Each test builds its own fixtures with a suffix unique to that test method
(rather than sharing a setUp()): this suite observed that IntegrationTestCase
does not roll back between individual test methods in this environment, only
at class teardown, so identically-named fixtures across methods collide.
"""

import io

import frappe
from frappe.tests import IntegrationTestCase
from frappe.utils import add_days, formatdate, nowdate
from openpyxl import load_workbook

from freightmas.freightmas.page.shipment_dashboard import shipment_dashboard
from freightmas.portal.api import dashboard as portal_dashboard
from freightmas.portal.api import shipments as portal_shipments


def _make_customer(suffix):
	customer = frappe.get_doc(
		{
			"doctype": "Customer",
			"customer_name": f"Portal Shipments Test Customer {suffix}",
			"customer_type": "Company",
			"customer_group": "Commercial",
			"territory": "Zimbabwe",
		}
	)
	customer.insert(ignore_permissions=True)
	return customer


def _make_user_and_contact(suffix, customer):
	user = frappe.get_doc(
		{
			"doctype": "User",
			"email": f"portal.shp.{suffix}@example.com",
			"first_name": f"Portal Shipments {suffix}",
			"send_welcome_email": 0,
		}
	)
	user.insert(ignore_permissions=True)

	contact = frappe.get_doc(
		{
			"doctype": "Contact",
			"first_name": user.first_name,
			"user": user.name,
			"email_ids": [{"email_id": user.email, "is_primary": 1}],
		}
	)
	contact.append("links", {"link_doctype": "Customer", "link_name": customer.name})
	contact.insert(ignore_permissions=True)

	return user


def _make_forwarding_job(customer, suffix, status="In Progress"):
	# Inserted as Draft: the controller's ensure_planned_charges_before_status_
	# change() blocks leaving Draft without planned revenue/cost charges, which
	# is irrelevant to these read-path scoping tests. Force the status directly
	# at the DB level afterwards instead of wiring up charges just to satisfy it.
	job = frappe.get_doc(
		{
			"doctype": "Forwarding Job",
			"company": "Maita",
			"created_by": "Administrator",
			"naming_series": "FWJB-.#####.-.YY.",
			"shipment_mode": "Sea",
			"incoterms": "CIF",
			"direction": "Import",
			"shipment_type": "FCL",
			"customer": customer.name,
			"customer_reference": f"PO-PORTAL-TEST-{suffix}",
			"consignee": customer.name,
			"port_of_loading": "Beira",
			"port_of_discharge": "Harare",
			"destination": "Harare",
			"eta": add_days(nowdate(), 5),
			"status": "Draft",
		}
	)
	job.insert(ignore_permissions=True)
	if status != "Draft":
		frappe.db.set_value("Forwarding Job", job.name, "status", status)
		job.status = status
	return job


def _add_container(job, suffix, container_type="20SD", **milestone_kwargs):
	"""Append a Cargo Parcel Details row and set its trucking-milestone
	booleans/dates directly at the DB level, bypassing the controller's
	milestone-sequence validation (same technique _make_forwarding_job uses
	for status) since these read-path tests don't care about that business
	rule.

	Reloads from the DB first rather than trusting the caller's `job`
	reference is still current - callers append multiple containers across
	several calls and a stale in-memory `modified` timestamp trips
	TimestampMismatchError on save.
	"""
	job = frappe.get_doc("Forwarding Job", job.name)
	job.append("cargo_parcel_details", {
		"cargo_type": "Containerised",
		"container_number": f"CONT-{suffix}",
		"container_type": container_type,
		"cargo_quantity": 1,
	})
	job.save(ignore_permissions=True)
	row = job.cargo_parcel_details[-1]
	if milestone_kwargs:
		frappe.db.set_value("Cargo Parcel Details", row.name, milestone_kwargs)
	job.reload()
	return job


def _complete_milestone(job, fieldname, completed_on=None):
	"""Mark the first row of a Job Milestone Progress checklist as completed,
	directly at the DB level."""
	rows = job.get(fieldname) or []
	if not rows:
		return
	frappe.db.set_value(
		"Job Milestone Progress",
		rows[0].name,
		{"is_completed": 1, "completed_on": completed_on or nowdate()},
	)


def _make_pair(suffix):
	"""Two customers + a job each + one portal user provisioned for A."""
	customer_a = _make_customer(f"{suffix}a")
	customer_b = _make_customer(f"{suffix}b")
	user_a = _make_user_and_contact(f"{suffix}a", customer_a)
	job_a = _make_forwarding_job(customer_a, f"{suffix}a")
	job_b = _make_forwarding_job(customer_b, f"{suffix}b")
	return customer_a, customer_b, user_a, job_a, job_b


class TestPortalShipmentsCrossTenant(IntegrationTestCase):
	def setUp(self):
		frappe.set_user("Administrator")

	def test_get_jobs_returns_only_own_customer_jobs(self):
		_customer_a, _customer_b, user_a, job_a, job_b = _make_pair("E1")

		frappe.set_user(user_a.name)
		try:
			result = portal_shipments.get_jobs()
		finally:
			frappe.set_user("Administrator")

		names = [j.name for j in result["jobs"]]
		self.assertIn(job_a.name, names)
		self.assertNotIn(job_b.name, names)

	def test_get_job_detail_allows_own_job(self):
		_customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E2")

		frappe.set_user(user_a.name)
		try:
			result = portal_shipments.get_job_detail(job_a.name)
		finally:
			frappe.set_user("Administrator")

		self.assertEqual(result["header"]["name"], job_a.name)
		self.assertIn("tracking_view", result)
		self.assertIn("banner", result["tracking_view"])
		# Financial/margin fields must never appear in a portal response.
		for leaky_key in ("finance", "dnd_totals", "purchase_invoices", "sales_invoices"):
			self.assertNotIn(leaky_key, result)

	def test_get_job_detail_denies_other_customers_job(self):
		_customer_a, _customer_b, user_a, _job_a, job_b = _make_pair("E3")

		frappe.set_user(user_a.name)
		try:
			with self.assertRaises(frappe.PermissionError):
				portal_shipments.get_job_detail(job_b.name)
		finally:
			frappe.set_user("Administrator")

	def test_get_overview_scopes_kpis_to_own_customer(self):
		customer_a, _customer_b, user_a, job_a, job_b = _make_pair("E4")
		completed_job = _make_forwarding_job(customer_a, "E4a2", status="Completed")
		frappe.db.set_value("Forwarding Job", job_a.name, "operational_phase", "planning")
		frappe.db.set_value("Forwarding Job", completed_job.name, "operational_phase", "closed")
		frappe.db.set_value("Forwarding Job", job_b.name, "operational_phase", "in_transit")

		frappe.set_user(user_a.name)
		try:
			result = portal_dashboard.get_overview()
		finally:
			frappe.set_user("Administrator")

		self.assertEqual(len(result["phase_pipeline"]), 7)
		at_origin = next(b for b in result["phase_pipeline"] if b["key"] == "at_origin")
		in_transit = next(b for b in result["phase_pipeline"] if b["key"] == "in_transit")
		self.assertEqual(at_origin["count"], 1)  # job_a only; completed job excluded from pipeline phases
		self.assertEqual(in_transit["count"], 0)  # job_b belongs to another customer
		in_motion_names = [j.name for j in result["in_motion_jobs"]]
		self.assertIn(job_a.name, in_motion_names)
		self.assertNotIn(job_b.name, in_motion_names)
		self.assertIn("active_count", result)
		self.assertIn("paid_ytd", result)
		self.assertIn("attention_items", result)

	def test_get_jobs_filters_by_operational_phases(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E4b")
		job_origin = _make_forwarding_job(customer_a, "E4b-origin")
		frappe.db.set_value("Forwarding Job", job_a.name, "operational_phase", "in_transit")
		frappe.db.set_value("Forwarding Job", job_origin.name, "operational_phase", "planning")

		frappe.set_user(user_a.name)
		try:
			result = portal_shipments.get_jobs(operational_phases="planning,awaiting_departure")
		finally:
			frappe.set_user("Administrator")

		names = [j.name for j in result["jobs"]]
		self.assertIn(job_origin.name, names)
		self.assertNotIn(job_a.name, names)

	def test_portal_endpoints_reject_guest(self):
		frappe.set_user("Guest")
		try:
			with self.assertRaises(frappe.PermissionError):
				portal_shipments.get_jobs()
		finally:
			frappe.set_user("Administrator")

	def test_internal_dashboard_still_sees_all_customers_unfiltered(self):
		# Regression guard: Phase 1 must not have touched the internal
		# (Desk-facing) shipment dashboard's own access or data scope.
		_customer_a, _customer_b, _user_a, job_a, job_b = _make_pair("E6")

		frappe.set_user("Administrator")
		result_a = shipment_dashboard.get_jobs(search=job_a.customer_reference[:14])
		self.assertIn(job_a.name, [j.name for j in result_a["jobs"]])

		result_b = shipment_dashboard.get_jobs(search=job_b.customer_reference[:14])
		self.assertIn(job_b.name, [j.name for j in result_b["jobs"]])

	def _export_as_user(self, user, **kwargs):
		"""Runs export_tracking_report and parses the resulting workbook."""
		frappe.set_user(user.name)
		try:
			portal_shipments.export_tracking_report(**kwargs)
		finally:
			frappe.set_user("Administrator")
		wb = load_workbook(io.BytesIO(frappe.local.response.filecontent))
		self.assertEqual(wb.sheetnames, ["Shipments", "Detail", "Containers"])

		ws = wb["Shipments"]
		rows = list(ws.iter_rows(values_only=True))
		self.assertIn("SHIPMENT TRACKING", str(rows[0][0]))
		header = rows[1]
		data_rows = [r for r in rows[2:] if r[0] is not None]
		return header, data_rows, wb

	def test_export_tracking_report_scopes_to_own_customer(self):
		customer_a, _customer_b, user_a, job_a, job_b = _make_pair("E7")

		header, data_rows, _wb = self._export_as_user(user_a)

		self.assertEqual(header[0], "Reference")
		self.assertEqual(header[1], "Job Ref")
		job_refs = {r[1] for r in data_rows}
		self.assertIn(job_a.name, job_refs)
		self.assertNotIn(job_b.name, job_refs)

	def test_export_tracking_report_one_row_per_job_on_shipments_sheet(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E8")
		_add_container(job_a, "E8-1", booked_on_date=nowdate())
		_add_container(job_a, "E8-2", booked_on_date=nowdate(), loaded_on_date=nowdate())

		header, data_rows, wb = self._export_as_user(user_a)

		job_rows = [r for r in data_rows if r[1] == job_a.name]
		self.assertEqual(len(job_rows), 1)

		container_rows = list(wb["Containers"].iter_rows(min_row=3, values_only=True))
		container_rows = [r for r in container_rows if r and r[0] == job_a.name]
		self.assertEqual(len(container_rows), 2)
		container_numbers = {r[2] for r in container_rows}
		self.assertEqual(container_numbers, {"CONT-E8-1", "CONT-E8-2"})

	def test_export_tracking_report_job_with_no_containers_gets_one_shipment_row(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E9")

		header, data_rows, wb = self._export_as_user(user_a)

		job_rows = [r for r in data_rows if r[1] == job_a.name]
		self.assertEqual(len(job_rows), 1)

		container_rows = list(wb["Containers"].iter_rows(min_row=3, values_only=True))
		container_rows = [r for r in container_rows if r and r[0]]
		self.assertEqual(container_rows, [])

	def test_export_tracking_report_detail_sheet_has_journey_columns(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E10")
		frappe.db.set_value("Forwarding Job", job_a.name, "requires_port_clearance", 1)
		job_a.reload()
		job_a.save(ignore_permissions=True)
		job_a.reload()
		_complete_milestone(job_a, "port_clearance_milestones", completed_on=nowdate())

		_header, _data_rows, wb = self._export_as_user(user_a)
		detail_header = list(wb["Detail"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
		self.assertIn("Port Clearance Summary", detail_header)
		self.assertNotIn("Consignee", detail_header)

	def test_export_tracking_report_uses_client_progress_not_milestone_grid(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E12")
		frappe.db.set_value("Forwarding Job", job_a.name, "atd", nowdate())

		header, data_rows, wb = self._export_as_user(user_a)
		job_rows = [r for r in data_rows if r[1] == job_a.name]
		self.assertEqual(len(job_rows), 1)

		progress_col = header.index("Progress %")
		self.assertIsNotNone(job_rows[0][progress_col])
		detail_header = list(wb["Detail"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
		self.assertNotIn("% Complete", detail_header)

	def test_export_tracking_report_omits_port_clearance_when_not_required(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E13")
		frappe.db.set_value("Forwarding Job", job_a.name, "requires_border_clearance", 1)
		job_a.reload()
		job_a.save(ignore_permissions=True)
		job_a.reload()
		self.assertTrue(job_a.border_clearance_milestones)
		_complete_milestone(job_a, "border_clearance_milestones", completed_on=nowdate())

		_header, data_rows, wb = self._export_as_user(user_a)
		detail_rows = list(wb["Detail"].iter_rows(min_row=3, values_only=True))
		job_detail = next(r for r in detail_rows if r[0] == job_a.name)
		detail_header = list(wb["Detail"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
		port_summary_col = detail_header.index("Port Clearance Summary")
		self.assertIsNone(job_detail[port_summary_col])

	def test_export_tracking_report_reflects_status_filter(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E11")
		_make_forwarding_job(customer_a, "E11a2", status="Completed")

		_header, data_rows, _wb = self._export_as_user(user_a, status="Completed")
		job_refs = {r[1] for r in data_rows}
		self.assertNotIn(job_a.name, job_refs)  # job_a is "In Progress", filtered out

	def test_export_tracking_report_excludes_submitted_jobs(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E14")

		_header, data_rows, _wb = self._export_as_user(user_a)
		self.assertIn(job_a.name, {r[1] for r in data_rows})  # still Draft (docstatus 0)

		# Force submitted at the DB level - same bypass technique already used
		# for `status` in _make_forwarding_job, avoids triggering the
		# controller's full submit-time validation just for this scoping test.
		frappe.db.set_value("Forwarding Job", job_a.name, "docstatus", 1)

		_header, data_rows, _wb = self._export_as_user(user_a)
		self.assertNotIn(job_a.name, {r[1] for r in data_rows})

	def test_get_tracking_summary_returns_active_counts(self):
		customer_a, _customer_b, user_a, job_a, _job_b = _make_pair("E15")
		frappe.db.set_value("Forwarding Job", job_a.name, "operational_phase", "at_terminal")

		frappe.set_user(user_a.name)
		try:
			result = portal_shipments.get_tracking_summary(status="In Progress")
		finally:
			frappe.set_user("Administrator")

		for key in ("active_count", "delayed_count", "at_port_count", "arriving_soon_count", "filtered_count"):
			self.assertIn(key, result)
		self.assertGreaterEqual(result["active_count"], 1)
		self.assertGreaterEqual(result["filtered_count"], 1)

	def test_export_tracking_report_pdf_sets_download_disposition(self):
		_customer_a, _customer_b, user_a, _job_a, _job_b = _make_pair("E16")

		frappe.set_user(user_a.name)
		try:
			portal_shipments.export_tracking_report_pdf(status="In Progress")
		finally:
			frappe.set_user("Administrator")

		self.assertEqual(frappe.local.response.type, "download")
		self.assertTrue(frappe.local.response.filename.endswith(".pdf"))
		self.assertTrue(frappe.local.response.filecontent.startswith(b"%PDF"))
