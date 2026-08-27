# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd and contributors
# For license information, please see license.txt

import io

from frappe.tests import IntegrationTestCase
from openpyxl import load_workbook

from freightmas.forwarding_service.utils.client_tracking_export import build_client_tracking_workbook
from freightmas.forwarding_service.utils.client_tracking_view import build_pdf_job_context
from freightmas.tests.test_client_tracking_view import _make_customer, _make_job


class TestClientTrackingExport(IntegrationTestCase):
	def setUp(self):
		import frappe

		frappe.set_user("Administrator")

	def test_workbook_has_three_client_sheets(self):
		customer = _make_customer("XL1")
		job = _make_job(customer, "XL1", customer_reference="PO-XL-001")

		file_bytes = build_client_tracking_workbook(
			customer.customer_name,
			[build_pdf_job_context(job)],
		)
		wb = load_workbook(io.BytesIO(file_bytes))
		self.assertEqual(wb.sheetnames, ["Shipments", "Detail", "Containers"])

		shipment_rows = list(wb["Shipments"].iter_rows(min_row=3, values_only=True))
		self.assertEqual(len([r for r in shipment_rows if r[0]]), 1)
		self.assertEqual(shipment_rows[0][0], "PO-XL-001")
