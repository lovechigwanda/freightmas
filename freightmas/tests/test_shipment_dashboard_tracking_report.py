# Copyright (c) 2026, Zvomaita Technologies (Pvt) Ltd and contributors
# For license information, please see license.txt

"""
Tests for the Command Center's tracking-report Excel export
(freightmas.freightmas.page.shipment_dashboard.shipment_dashboard.export_tracking_report),
the internal multi-customer counterpart to the Client Portal's tracking
report. Column shape, amber/green/NA shading, and % Complete math are
intentionally mirrored from test_client_portal_shipments.py's tests for
that version, since the two exporters share the same design - the only
real difference under test here is the multi-customer `customers` filter
and the FreightMas-role staff gate instead of the portal gate.

Each test builds its own fixtures with a suffix unique to that test method:
IntegrationTestCase does not roll back between individual test methods in
this environment, only at class teardown, so identically-named fixtures
across methods collide (see test_client_portal_shipments.py for the same
observation).
"""

import io

import frappe
from frappe.tests import IntegrationTestCase
from frappe.utils import add_days, formatdate, nowdate
from openpyxl import load_workbook

from freightmas.freightmas.page.shipment_dashboard import shipment_dashboard


def _make_customer(suffix):
	customer = frappe.get_doc({
		"doctype": "Customer",
		"customer_name": f"Dashboard Tracking Test Customer {suffix}",
		"customer_type": "Company",
		"customer_group": "Commercial",
		"territory": "Zimbabwe",
	})
	customer.insert(ignore_permissions=True)
	return customer


def _make_staff_user(suffix, roles=("FreightMas User",)):
	user = frappe.get_doc({
		"doctype": "User",
		"email": f"dashboard.trpt.{suffix}@example.com",
		"first_name": f"Dashboard Tracking {suffix}",
		"send_welcome_email": 0,
	})
	for role in roles:
		user.append("roles", {"role": role})
	user.insert(ignore_permissions=True)
	return user


def _make_forwarding_job(customer, suffix, status="In Progress", **extra):
	# requires_port_clearance/requires_border_clearance can be passed in
	# `extra` - set before insert() so populate_mode_milestones() (runs in
	# validate(), including on insert) populates the right checklist rows
	# on the very first save, no reload/re-save cycle needed.
	job = frappe.get_doc({
		"doctype": "Forwarding Job",
		"company": "Maita",
		"created_by": "Administrator",
		"naming_series": "FWJB-.#####.-.YY.",
		"shipment_mode": "Sea",
		"incoterms": "CIF",
		"direction": "Import",
		"shipment_type": "FCL",
		"customer": customer.name,
		"customer_reference": f"PO-DASH-TEST-{suffix}",
		"consignee": customer.name,
		"port_of_loading": "Beira",
		"port_of_discharge": "Harare",
		"destination": "Harare",
		"eta": add_days(nowdate(), 5),
		"status": "Draft",
		**extra,
	})
	job.insert(ignore_permissions=True)
	if status != "Draft":
		frappe.db.set_value("Forwarding Job", job.name, "status", status)
		job.status = status
	return job


def _add_container(job, suffix, container_type="20SD", **milestone_kwargs):
	"""Reloads from the DB first rather than trusting the caller's `job`
	reference is still current - callers append multiple containers across
	several calls and a stale in-memory `modified` timestamp trips
	TimestampMismatchError on save."""
	job = frappe.get_doc("Forwarding Job", job.name)
	job.append("cargo_parcel_details", {
		"cargo_type": "Containerised",
		"container_number": f"CONT-{suffix}",
		"container_type": container_type,
		"cargo_quantity": 1,
	})
	job.save(ignore_permissions=True)
	row = job.cargo_parcel_details[-1]
	if milestone_kwargs:
		frappe.db.set_value("Cargo Parcel Details", row.name, milestone_kwargs)
	job.reload()
	return job


def _complete_milestone(job, fieldname, completed_on=None):
	rows = job.get(fieldname) or []
	if not rows:
		return
	frappe.db.set_value(
		"Job Milestone Progress",
		rows[0].name,
		{"is_completed": 1, "completed_on": completed_on or nowdate()},
	)


class TestShipmentDashboardTrackingReport(IntegrationTestCase):
	def setUp(self):
		frappe.set_user("Administrator")

	def _export_as(self, user, **kwargs):
		frappe.set_user(user.name)
		try:
			shipment_dashboard.export_tracking_report(**kwargs)
		finally:
			frappe.set_user("Administrator")
		wb = load_workbook(io.BytesIO(frappe.local.response.filecontent))
		self.assertEqual(wb.sheetnames, ["Tracking Report"])
		ws = wb["Tracking Report"]
		rows = list(ws.iter_rows(values_only=True))
		self.assertEqual(rows[0][0], "Shipment Tracking Report")
		header = rows[2]
		data_rows = [r for r in rows[3:] if r[0] is not None]
		return header, data_rows

	def test_rejects_non_freightmas_user(self):
		user = _make_staff_user("E1", roles=[])
		frappe.set_user(user.name)
		try:
			with self.assertRaises(frappe.PermissionError):
				shipment_dashboard.export_tracking_report()
		finally:
			frappe.set_user("Administrator")

	def test_no_customer_filter_includes_multiple_customers(self):
		customer_a = _make_customer("E2a")
		customer_b = _make_customer("E2b")
		staff = _make_staff_user("E2")
		job_a = _make_forwarding_job(customer_a, "E2a")
		job_b = _make_forwarding_job(customer_b, "E2b")

		_header, data_rows = self._export_as(staff)
		job_ids = {r[0] for r in data_rows}
		self.assertIn(job_a.name, job_ids)
		self.assertIn(job_b.name, job_ids)

	def test_customer_filter_restricts_to_selected_customers(self):
		customer_a = _make_customer("E3a")
		customer_b = _make_customer("E3b")
		staff = _make_staff_user("E3")
		job_a = _make_forwarding_job(customer_a, "E3a")
		job_b = _make_forwarding_job(customer_b, "E3b")

		_header, data_rows = self._export_as(staff, customers=frappe.as_json([customer_a.name]))
		job_ids = {r[0] for r in data_rows}
		self.assertIn(job_a.name, job_ids)
		self.assertNotIn(job_b.name, job_ids)

	def test_one_row_per_container(self):
		customer_a = _make_customer("E4")
		staff = _make_staff_user("E4")
		job_a = _make_forwarding_job(customer_a, "E4")
		_add_container(job_a, "E4-1", booked_on_date=nowdate())
		_add_container(job_a, "E4-2", booked_on_date=nowdate(), loaded_on_date=nowdate())

		header, data_rows = self._export_as(staff, customers=frappe.as_json([customer_a.name]))
		job_rows = [r for r in data_rows if r[0] == job_a.name]
		self.assertEqual(len(job_rows), 2)

		booked_col = header.index("Booked")
		loaded_col = header.index("Loaded")
		self.assertEqual({r[booked_col] for r in job_rows}, {formatdate(nowdate(), "dd-MMM-yy")})
		# openpyxl round-trips a blank cell (we write "") back as None.
		self.assertEqual(
			{r[loaded_col] for r in job_rows},
			{None, formatdate(nowdate(), "dd-MMM-yy")},
		)

	def test_shows_dash_for_not_required_service(self):
		customer_a = _make_customer("E5")
		staff = _make_staff_user("E5")
		job_a = _make_forwarding_job(customer_a, "E5", requires_border_clearance=1)
		self.assertTrue(job_a.border_clearance_milestones)  # checklist populated
		_complete_milestone(job_a, "border_clearance_milestones", completed_on=nowdate())

		header, data_rows = self._export_as(staff, customers=frappe.as_json([customer_a.name]))
		job_rows = [r for r in data_rows if r[0] == job_a.name]
		self.assertEqual(len(job_rows), 1)

		# requires_port_clearance was never set -> every Port Clearance
		# column (between Sea/Air's last column and Road Transport's first)
		# shows "-", not blank.
		port_clearance_cols = range(header.index("Empty Returned") + 1, header.index("Booked"))
		self.assertTrue(port_clearance_cols)  # sanity: some Port Clearance definitions exist
		for col in port_clearance_cols:
			self.assertEqual(job_rows[0][col], "-")

		completed_label = job_a.border_clearance_milestones[0].milestone_label
		border_col = header.index(completed_label)
		self.assertEqual(job_rows[0][border_col], formatdate(nowdate(), "dd-MMM-yy"))

	def test_percent_complete_ignores_inapplicable_milestones(self):
		customer_a = _make_customer("E6")
		staff = _make_staff_user("E6")
		job_a = _make_forwarding_job(customer_a, "E6")
		frappe.db.set_value("Forwarding Job", job_a.name, "atd", nowdate())

		header, data_rows = self._export_as(staff, customers=frappe.as_json([customer_a.name]))
		job_rows = [r for r in data_rows if r[0] == job_a.name]
		self.assertEqual(len(job_rows), 1)

		percent_col = header.index("% Complete")
		# 2/12: ATD (just set) + ETA/ATA (the fixture's own `eta` falls back
		# into this column) - neither Port nor Border Clearance is required,
		# so those columns aren't counted in the denominator at all.
		self.assertAlmostEqual(job_rows[0][percent_col], 2 / 12, places=4)
