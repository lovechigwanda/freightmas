# Client Portal read API: Forwarding Job list + detail + tracking.
#
# Mirrors the query idiom of freightmas.freightmas.page.shipment_dashboard.
# shipment_dashboard (SQL join / milestone-percent aggregation style) but
# is re-implemented rather than called directly: that module is gated by
# check_freightmas_role() and returns every customer's data (including
# WIP/margin fields), and there is no safe way to parameterize it to one
# Customer without changing its gate - which would weaken the internal
# dashboard. Every field returned here has been hand-picked to exclude
# costing/margin data.

import frappe
from frappe import _
from frappe.utils import formatdate, getdate, now_datetime, nowdate

from freightmas.freightmas.report.report_export_utils import send_excel_response
from freightmas.portal.security import (
	assert_customer_scope,
	check_portal_access,
	get_portal_customer_names,
	log_portal_access,
)

NOT_ACTIVE_STATUSES = ["Completed", "Closed", "Cancelled"]

JOB_LIST_FIELDS = [
	"name", "customer_reference", "direction", "shipment_mode", "shipment_type",
	"status", "port_of_loading", "port_of_discharge", "destination",
	"vessel_flight_no", "bl_number", "cargo_count", "eta", "ata", "etd", "atd",
	"discharge_date", "current_comment", "last_updated_on",
]

MILESTONE_TABLE_FIELDS = [
	"road_freight_milestones",
	"port_clearance_milestones",
	"border_clearance_milestones",
	"warehouse_milestones",
]

# ============================================================
# Tracking report (Excel export) - one flat table, one row per container
# (a job with none gets one row, container columns blank). Job/BL-level
# facts (Sea/Air dates, Port/Border/Warehouse Milestones) repeat identically
# across a job's container rows - by design, since they genuinely are one
# BL-level fact. Readability instead comes from giving every individual
# milestone its own date column (blank = outstanding, amber; date entered =
# achieved, green) and color-banding columns into sections, following the
# reference tracking-sheet template the user supplied, rather than any
# structural trick to avoid the repetition (two earlier versions of this
# report tried splitting into blocks / Excel row-grouping instead - both
# were reworked after real-file review).
# ============================================================

AMBER_FILL = "FDEBD0"
GREEN_FILL = "D5F5E3"
NA_FILL = "D9D9D9"
PERCENT_FILL = "E8EDF7"

# Sentinel written into a job's field dict for a dynamic (Port/Border
# Clearance) milestone column when that service isn't required (ticked) on
# the job at all - distinct from "required but not yet done" (None). The
# workbook builder renders this as "-" with a neutral fill, and it's
# excluded entirely from the % Complete applicable/completed counts.
NOT_APPLICABLE = object()

# (section title, band color, [(label, fieldname, kind), ...]) - kind is
# "identity" (data cells shaded amber, header white-on-dark-grey), "milestone"
# (date, amber/green/NA shaded), "status" (plain free text), or "percent"
# (computed ratio, own fill).
SHIPMENT_DETAILS_SECTION = ("Shipment Details", "595959", [
	("Job ID", "job_id", "identity"),
	("Consignee", "consignee", "identity"),
	("Customer Reference", "customer_reference", "identity"),
	("BL Number", "bl_number", "identity"),
	("Container Number", "container_number", "identity"),
	("Type", "container_type", "identity"),
])

SEA_AIR_SECTION = ("Sea / Air Freight", "2E5C8A", [
	("Departed Origin (ATD)", "atd", "milestone"),
	("ETA/ATA", "eta_ata", "milestone"),
	("Discharged (Shipment)", "discharge_date", "milestone"),
	("Container Discharged", "container_discharge_date", "milestone"),
	("Gate Out", "gate_out_date", "milestone"),
	("Empty Returned", "empty_return_date", "milestone"),
])

ROAD_TRANSPORT_SECTION = ("Road Transport", "1E7A6F", [
	("Booked", "booked_on_date", "milestone"),
	("Loaded", "loaded_on_date", "milestone"),
	("Offloaded", "offloaded_on_date", "milestone"),
	("Returned", "returned_on_date", "milestone"),
	("Completed", "trucking_completed_on_date", "milestone"),
])

OVERVIEW_SECTION = ("Overview", "6B6B6B", [
	("Completed", "job_completed_on", "milestone"),
	("Status", "status_comment", "status"),
	("% Complete", "percent_complete", "percent"),
])

# service_module -> band color, for the dynamic milestone sections built live
# from Milestone Definition (sequence order) - the system's real milestone
# set, not hardcoded column names. Warehouse dropped for now (per user
# request) - re-add here (and to _build_report_sections()'s ordered list)
# if it comes back.
DYNAMIC_MILESTONE_COLORS = {
	"Port Clearance": "5B3A8E",
	"Border Clearance": "A04B2E",
}


def _milestone_definition_columns(service_module):
	"""One (label, fieldname) column per active Milestone Definition for this
	service_module, in sequence order, deduped by milestone_code (used as the
	fieldname - already globally unique, e.g. "PC_VESSEL_ARRIVED")."""
	rows = frappe.get_all(
		"Milestone Definition",
		filters={"service_module": service_module, "is_active": 1},
		fields=["milestone_code", "milestone_label"],
		order_by="sequence asc",
	)
	seen = set()
	columns = []
	for r in rows:
		if r.milestone_code in seen:
			continue
		seen.add(r.milestone_code)
		columns.append((r.milestone_label, r.milestone_code, "milestone"))
	return columns


def _dynamic_section(service_module):
	return (service_module, DYNAMIC_MILESTONE_COLORS[service_module], _milestone_definition_columns(service_module))


def _build_report_sections():
	"""Full ordered section list for this request, in the specific display
	order requested: Shipment Details, Sea/Air Freight, Port Clearance, Road
	Transport, Border Clearance, Overview."""
	return [
		SHIPMENT_DETAILS_SECTION,
		SEA_AIR_SECTION,
		_dynamic_section("Port Clearance"),
		ROAD_TRANSPORT_SECTION,
		_dynamic_section("Border Clearance"),
		OVERVIEW_SECTION,
	]


def _caller_customer_filter():
	customers = get_portal_customer_names()
	if not customers:
		frappe.throw(
			_("Your account is not linked to a customer profile. Contact your account manager."),
			frappe.PermissionError,
		)
	return customers


def _job_list_filters(customers, status=None, direction=None, search=None):
	filters = {"docstatus": ["<", 2], "customer": ["in", customers]}
	if status:
		filters["status"] = status
	if direction:
		filters["direction"] = direction

	or_filters = None
	if search:
		or_filters = [
			["name", "like", f"%{search}%"],
			["customer_reference", "like", f"%{search}%"],
			["bl_number", "like", f"%{search}%"],
		]
	return filters, or_filters


def _milestone_progress_map(job_names):
	if not job_names:
		return {}

	counts = {}
	for fieldname in MILESTONE_TABLE_FIELDS:
		rows = frappe.db.sql(
			"""
			SELECT parent, COUNT(*) AS total, SUM(IFNULL(is_completed, 0)) AS done
			FROM `tabJob Milestone Progress`
			WHERE parenttype = 'Forwarding Job' AND parentfield = %(field)s AND parent IN %(names)s
			GROUP BY parent
			""",
			{"field": fieldname, "names": job_names},
			as_dict=True,
		)
		for r in rows:
			bucket = counts.setdefault(r.parent, {"total": 0, "done": 0})
			bucket["total"] += r.total or 0
			bucket["done"] += int(r.done or 0)

	return {
		name: (round(v["done"] / v["total"] * 100) if v["total"] else 0)
		for name, v in counts.items()
	}


@frappe.whitelist()
def get_jobs(status=None, direction=None, search=None, limit_start=0, limit_page_length=20):
	check_portal_access()
	customers = _caller_customer_filter()

	filters, or_filters = _job_list_filters(customers, status, direction, search)

	# get_all(), not get_list(): Customer Portal User holds zero DocType
	# permissions by design (see freightmas/portal/security.py) - the
	# explicit `customer` filter above is the actual access boundary, not
	# Frappe's own permission system, so it must not be re-checked here.
	jobs = frappe.get_all(
		"Forwarding Job",
		filters=filters,
		or_filters=or_filters,
		fields=JOB_LIST_FIELDS,
		order_by="modified desc",
		limit_start=frappe.utils.cint(limit_start),
		limit_page_length=frappe.utils.cint(limit_page_length),
	)

	total_count = frappe.db.count("Forwarding Job", filters=filters)

	progress_map = _milestone_progress_map([j.name for j in jobs])
	today = getdate(nowdate())
	for j in jobs:
		j["milestone_percent"] = progress_map.get(j.name, 0)
		j["is_overdue"] = bool(
			(j.direction == "Import" and j.eta and getdate(j.eta) < today and not j.ata)
			or (j.direction == "Export" and j.etd and getdate(j.etd) < today and not j.atd)
		)

	log_portal_access("list_shipments", doctype="Forwarding Job")

	return {"jobs": jobs, "total_count": total_count}


@frappe.whitelist()
def get_job_detail(job_name):
	check_portal_access()
	customer = assert_customer_scope("Forwarding Job", job_name, "customer")

	doc = frappe.get_doc("Forwarding Job", job_name)

	header = {
		"name": doc.name,
		"customer_reference": doc.customer_reference,
		"consignee": doc.consignee,
		"direction": doc.direction,
		"shipment_mode": doc.shipment_mode,
		"shipment_type": doc.shipment_type,
		"status": doc.status,
		"port_of_loading": doc.port_of_loading,
		"port_of_discharge": doc.port_of_discharge,
		"destination": doc.destination,
		"vessel_flight_no": doc.vessel_flight_no,
		"vessel_flight_date": doc.vessel_flight_date,
		"bl_number": doc.bl_number,
		"is_bl_received": doc.is_bl_received,
		"cargo_description": doc.cargo_description,
		"cargo_count": doc.cargo_count,
		"incoterms": doc.incoterms,
		"current_comment": doc.current_comment,
		"last_updated_on": doc.last_updated_on,
	}

	shipment_dates = {
		"booking_date": doc.booking_date,
		"cargo_ready_date": doc.cargo_ready_date,
		"etd": doc.etd,
		"atd": doc.atd,
		"eta": doc.eta,
		"ata": doc.ata,
		"discharge_date": doc.discharge_date,
		"completed_on": doc.completed_on,
	}

	milestone_stages = []
	section_labels = {
		"road_freight_milestones": "Road Freight",
		"port_clearance_milestones": "Port Clearance",
		"border_clearance_milestones": "Border Clearance",
		"warehouse_milestones": "Warehouse",
	}
	requires_map = {
		"road_freight_milestones": True,
		"port_clearance_milestones": doc.requires_port_clearance,
		"border_clearance_milestones": doc.requires_border_clearance,
		"warehouse_milestones": doc.requires_warehousing,
	}
	for fieldname, label in section_labels.items():
		if not requires_map.get(fieldname):
			continue
		rows = doc.get(fieldname) or []
		if not rows:
			continue
		milestone_stages.append(
			{
				"group": label,
				"milestones": [
					{
						"label": r.milestone_label,
						"is_completed": bool(r.is_completed),
						"completed_on": r.completed_on,
					}
					for r in rows
				],
			}
		)

	cargo = [
		{
			"name": r.name,
			"container_number": r.container_number or r.cargo_item_description,
			"container_type": r.container_type,
			"cargo_type": r.cargo_type,
			"cargo_quantity": r.cargo_quantity,
			"cargo_uom": r.cargo_uom,
			"is_hazardous": bool(r.is_hazardous),
			"is_booked": bool(r.is_booked),
			"is_loaded": bool(r.is_loaded),
			"is_offloaded": bool(r.is_offloaded),
			"is_returned": bool(r.is_returned),
			"is_completed": bool(r.is_completed),
			"truck_location": r.truck_location,
			"tracking_comment": r.tracking_comment,
			"updated_on": r.updated_on,
		}
		for r in (doc.cargo_parcel_details or [])
	]

	tracking = [
		{
			"event": r.event,
			"date": r.date,
			"source": r.source,
		}
		for r in sorted(doc.get("tracking_timeline") or [], key=lambda r: r.idx or 0, reverse=True)
	][:15]

	log_portal_access("view_shipment", doctype="Forwarding Job", docname=job_name, customer=customer)

	return {
		"header": header,
		"shipment_dates": shipment_dates,
		"milestone_stages": milestone_stages,
		"cargo": cargo,
		"tracking": tracking,
	}


def _build_tracking_workbook(sections, rows):
	"""One flat table: every individual milestone gets its own date column,
	color-banded into sections (Shipment Details / Sea-Air / Port Clearance /
	Road Transport / Border Clearance / Overview), with each milestone cell
	shaded amber (blank/outstanding), green (date present/achieved), or grey
	with a "-" (NOT_APPLICABLE - that service isn't required on this job at
	all), and Shipment Details cells shaded the same amber as an outstanding
	milestone (the user's explicit color choice) - so a table that still
	repeats job/BL-level values on every container row (that repetition is a
	genuine BL-level fact) stays readable via color rather than a structural
	trick. The 3 header rows are frozen.
	`sections` is the output of _build_report_sections(); `rows` is a list of
	flat {fieldname: value} dicts, one per report row.
	"""
	import io

	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from openpyxl.utils import get_column_letter

	# Flatten sections into one ordered column list, remembering section
	# spans for the merged band row.
	columns = []  # (label, fieldname, kind)
	band_spans = []  # (title, color, start_col, end_col)
	col_idx = 1
	for title, color, cols in sections:
		start = col_idx
		for label, fieldname, kind in cols:
			columns.append((label, fieldname, kind))
			col_idx += 1
		band_spans.append((title, color, start, col_idx - 1))
	ncols = len(columns)

	wb = Workbook()
	ws = wb.active
	ws.title = "Tracking Report"

	title_font = Font(bold=True, size=14, color="1F2A44")
	band_font = Font(bold=True, color="FFFFFF")
	header_font = Font(bold=True, color="FFFFFF")
	center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
	left_align = Alignment(horizontal="left", vertical="center")
	percent_align = Alignment(horizontal="center", vertical="center")

	ws.cell(row=1, column=1, value="Shipment Tracking Report").font = title_font
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

	for title, color, start, end in band_spans:
		cell = ws.cell(row=2, column=start, value=title)
		cell.fill = PatternFill("solid", fgColor=color)
		cell.font = band_font
		cell.alignment = center_wrap
		if end > start:
			ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)

	band_color_by_col = {}
	for title, color, start, end in band_spans:
		for c in range(start, end + 1):
			band_color_by_col[c] = color

	header_row = 3
	for idx, (label, fieldname, kind) in enumerate(columns, 1):
		cell = ws.cell(row=header_row, column=idx, value=label)
		cell.fill = PatternFill("solid", fgColor=band_color_by_col[idx])
		cell.font = header_font
		cell.alignment = center_wrap

	data_start = header_row + 1
	for row_offset, row_data in enumerate(rows):
		row_idx = data_start + row_offset
		for col_idx, (label, fieldname, kind) in enumerate(columns, 1):
			value = row_data.get(fieldname)
			cell = ws.cell(row=row_idx, column=col_idx)

			if kind == "milestone":
				if value is NOT_APPLICABLE:
					cell.value = "-"
					cell.fill = PatternFill("solid", fgColor=NA_FILL)
				else:
					cell.value = formatdate(value, "dd-MMM-yy") if value else ""
					cell.fill = PatternFill("solid", fgColor=GREEN_FILL if value else AMBER_FILL)
				cell.alignment = percent_align
			elif kind == "percent":
				cell.value = value or 0
				cell.number_format = "0%"
				cell.fill = PatternFill("solid", fgColor=PERCENT_FILL)
				cell.alignment = percent_align
			elif kind == "identity":
				cell.value = value or ""
				cell.alignment = left_align
				cell.fill = PatternFill("solid", fgColor=AMBER_FILL)
			else:  # status
				cell.value = value or ""
				cell.alignment = left_align

	ws.freeze_panes = f"A{data_start}"  # pins the 3 header rows only, no column freeze

	for col_idx in range(1, ncols + 1):
		max_length = 0
		for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
			for cell in row:
				length = len(str(cell.value)) if cell.value else 0
				max_length = max(max_length, length)
		ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_length + 2, 40))

	ws.sheet_view.showGridLines = False

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)
	return output.getvalue()


@frappe.whitelist()
def export_tracking_report(status=None, direction=None, search=None):
	check_portal_access()
	customers = _caller_customer_filter()

	filters, or_filters = _job_list_filters(customers, status, direction, search)

	jobs = frappe.get_all(
		"Forwarding Job",
		filters=filters,
		or_filters=or_filters,
		fields=[
			"name", "customer_reference", "consignee", "bl_number", "current_comment",
			"atd", "eta", "ata", "discharge_date", "completed_on",
			"requires_port_clearance", "requires_border_clearance",
		],
		order_by="modified desc",
	)
	job_names = [j.name for j in jobs]

	customer_name_map = {}
	consignee_names = list({j.consignee for j in jobs if j.consignee})
	if consignee_names:
		customer_name_map = {
			c.name: c.customer_name
			for c in frappe.get_all(
				"Customer", filters={"name": ["in", consignee_names]}, fields=["name", "customer_name"]
			)
		}

	sections = _build_report_sections()
	dynamic_milestone_fields = {
		title: [fieldname for _label, fieldname, _kind in cols]
		for title, _color, cols in sections
		if title in ("Port Clearance", "Border Clearance")
	}
	section_to_table = {
		"Port Clearance": "port_clearance_milestones",
		"Border Clearance": "border_clearance_milestones",
	}
	section_requires_field = {
		"Port Clearance": "requires_port_clearance",
		"Border Clearance": "requires_border_clearance",
	}

	cargo_by_job = {}
	milestone_rows_by_job = {table: {} for table in section_to_table.values()}

	if job_names:
		cargo_rows = frappe.get_all(
			"Cargo Parcel Details",
			filters={"parenttype": "Forwarding Job", "parent": ["in", job_names]},
			fields=[
				"parent", "container_number", "container_type",
				"discharge_date", "gate_out_date", "empty_return_date",
				"booked_on_date", "loaded_on_date", "offloaded_on_date",
				"returned_on_date", "completed_on_date",
			],
			order_by="parent, idx",
		)
		for r in cargo_rows:
			cargo_by_job.setdefault(r.parent, []).append(r)

		for table in section_to_table.values():
			rows = frappe.get_all(
				"Job Milestone Progress",
				filters={"parenttype": "Forwarding Job", "parentfield": table, "parent": ["in", job_names]},
				fields=["parent", "milestone_code", "is_completed", "completed_on"],
				order_by="parent, idx",
			)
			for r in rows:
				milestone_rows_by_job[table].setdefault(r.parent, []).append(r)

	report_rows = []
	for job in jobs:
		containers = cargo_by_job.get(job.name, [])

		job_fields = {
			"job_id": job.name,
			"consignee": customer_name_map.get(job.consignee, job.consignee),
			"customer_reference": job.customer_reference,
			"bl_number": job.bl_number,
			"atd": job.atd,
			"eta_ata": job.ata or job.eta,
			"discharge_date": job.discharge_date,
			"job_completed_on": job.completed_on,
			"status_comment": job.current_comment,
		}
		sea_air_job_applicable = ["atd", "eta_ata", "discharge_date"]

		# Per-job milestone maps: {milestone_code: completed_on}, plus counts
		# for % Complete. A service not required (ticked) on this job at all
		# gets NOT_APPLICABLE on every one of its milestone columns (rendered
		# as "-", not amber) and contributes nothing to applicable/completed -
		# explicitly gated on the job's own requires_* flag, not inferred from
		# row presence, so a job where the flag was later unticked (rows can
		# linger - populate_mode_milestones() never removes them) still reads
		# as not-applicable rather than "outstanding".
		section_counts = {}  # title -> (applicable, completed)
		for title, table in section_to_table.items():
			required = bool(job.get(section_requires_field[title]))
			rows = milestone_rows_by_job[table].get(job.name, []) if required else []
			value_map = {r.milestone_code: r.completed_on for r in rows if r.is_completed}
			for fieldname in dynamic_milestone_fields[title]:
				job_fields[fieldname] = value_map.get(fieldname) if required else NOT_APPLICABLE
			section_counts[title] = (len(rows), sum(1 for r in rows if r.is_completed))

		if not containers:
			containers = [None]

		for c in containers:
			row = dict(job_fields)
			if c is None:
				row.update({
					"container_number": None, "container_type": None,
					"container_discharge_date": None, "gate_out_date": None, "empty_return_date": None,
					"booked_on_date": None, "loaded_on_date": None, "offloaded_on_date": None,
					"returned_on_date": None, "trucking_completed_on_date": None,
				})
			else:
				row.update({
					"container_number": c.container_number,
					"container_type": c.container_type,
					"container_discharge_date": c.discharge_date,
					"gate_out_date": c.gate_out_date,
					"empty_return_date": c.empty_return_date,
					"booked_on_date": c.booked_on_date,
					"loaded_on_date": c.loaded_on_date,
					"offloaded_on_date": c.offloaded_on_date,
					"returned_on_date": c.returned_on_date,
					"trucking_completed_on_date": c.completed_on_date,
				})

			road_transport_fields = [
				"booked_on_date", "loaded_on_date", "offloaded_on_date",
				"returned_on_date", "trucking_completed_on_date",
			]
			sea_air_container_fields = ["container_discharge_date", "gate_out_date", "empty_return_date"]

			applicable = len(sea_air_job_applicable) + len(sea_air_container_fields) + len(road_transport_fields) + 1
			completed = (
				sum(1 for f in sea_air_job_applicable if row.get(f))
				+ sum(1 for f in sea_air_container_fields if row.get(f))
				+ sum(1 for f in road_transport_fields if row.get(f))
				+ (1 if row.get("job_completed_on") else 0)
			)
			for title in section_to_table:
				a, done = section_counts[title]
				applicable += a
				completed += done

			row["percent_complete"] = (completed / applicable) if applicable else 0
			report_rows.append(row)

	file_bytes = _build_tracking_workbook(sections, report_rows)
	filename = f"Tracking_Report_{now_datetime().strftime('%Y%m%d_%H%M')}.xlsx"
	send_excel_response(file_bytes, filename)

	log_portal_access("export_tracking_report", doctype="Forwarding Job")
