## Freightmas API Endpoints
###################################
import frappe
from erpnext.stock.utils import get_incoming_rate

@frappe.whitelist()
def get_fuel_rate(item_code, warehouse, posting_date=None):
    if not posting_date:
        posting_date = frappe.utils.today()

    args = {
        "item_code": item_code,
        "warehouse": warehouse,
        "posting_date": posting_date,
        "qty": 1,
        "allow_zero_valuation": 1
    }

    rate = get_incoming_rate(args)
    return rate or 0


################################################################################
import frappe
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from frappe.utils import now_datetime
import re

@frappe.whitelist()
def export_report_to_excel(report_name, filters=None):
    import json
    from io import BytesIO
    import importlib

    if isinstance(filters, str):
        filters = json.loads(filters)

    # Get columns and data using the report's execute method
    report = frappe.get_doc("Report", report_name)
    if report.report_type != "Script Report":
        frappe.throw("Only Script Reports are supported.")

    module = importlib.import_module(
        f"freightmas.{frappe.scrub(report.module)}.report.{frappe.scrub(report.name)}.{frappe.scrub(report.name)}"
    )
    columns, data = module.execute(filters)

    wb = openpyxl.Workbook()
    ws = wb.active

    sheet_title = re.sub(r'[\\/*?:\[\]]', '', report_name)[:31]
    ws.title = sheet_title

    # Styles
    bold_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=13)
    filter_label_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="305496")  # Modern blue
    zebra_fill = PatternFill("solid", fgColor="F2F2F2")  # Light gray for zebra
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    ncols = len(columns)
    row_idx = 1

    # Company Name (merged)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
    ws.cell(row=row_idx, column=1, value=frappe.defaults.get_user_default("Company")).font = title_font
    row_idx += 1

    # Report Title (merged)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ncols)
    ws.cell(row=row_idx, column=1, value=report_name).font = subtitle_font
    row_idx += 1

    # Filters (merged label and value)
    if filters:
        for label, val in filters.items():
            if val:
                # Format date filters
                if "date" in label and val:
                    try:
                        val = frappe.utils.formatdate(val, "dd-MMM-yy")
                    except Exception:
                        pass
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=1)
                ws.cell(row=row_idx, column=1, value=f"{label.replace('_', ' ').title()}:").font = filter_label_font
                ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=ncols)
                ws.cell(row=row_idx, column=2, value=val)
                row_idx += 1

    # Timestamp (label in first column, value in next column, merged across remaining columns)
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=ncols)
    ws.cell(row=row_idx, column=1, value="Exported:").font = filter_label_font
    ws.cell(row=row_idx, column=1).alignment = left_align
    export_time = now_datetime().strftime("%d-%b-%Y %H:%M")
    ws.cell(row=row_idx, column=2, value=export_time)
    ws.cell(row_idx, column=2).alignment = left_align
    row_idx += 1

    # Table Header
    header_row = row_idx
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col["label"])
        cell.font = bold_font
        cell.alignment = left_align
        cell.fill = header_fill
        cell.border = border
    row_idx += 1

    # Freeze panes below the header row
    ws.freeze_panes = ws[f"A{header_row+1}"]

    # Table Data (zebra striping)
    for i, row in enumerate(data, start=1):
        fill = zebra_fill if i % 2 == 0 else None
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col["fieldname"], "")
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Format numbers and currency
            if col.get("fieldtype") in ["Int", "Float", "Currency"]:
                if isinstance(value, (int, float)):
                    cell.value = value
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = 0
                    cell.number_format = '#,##0.00'
            # Format dates
            elif "date" in col["fieldname"] and value:
                try:
                    cell.value = frappe.utils.formatdate(value, "dd-MMM-yy")
                except Exception:
                    cell.value = value
            else:
                cell.value = value
                
            cell.border = border
            if fill:
                cell.fill = fill
            # Alignment
            if col.get("fieldtype") in ["Int", "Float", "Currency"]:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
        row_idx += 1

    # Auto-fit columns using only header and data rows
    for col_idx, col in enumerate(columns, start=1):
        max_length = 0
        # Only check header and data rows
        for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_length + 2, 40))

    # Hide gridlines
    ws.sheet_view.showGridLines = False

    # Save to BytesIO and return as file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.local.response.filename = f"{report_name.replace(' ', '_')}.xlsx"
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "binary"



###################################################
# Freightmas PDF Export for Script Reports
# This module exports Script Reports to PDF format.
# It uses Jinja templates for rendering the report
import frappe
import importlib
from frappe.utils.pdf import get_pdf
from frappe.utils.jinja import render_template
from frappe.utils import now_datetime
from frappe import _

@frappe.whitelist()
def export_report_to_pdf(report_name, filters):
    import json
    filters = json.loads(filters)

    # Dynamically import the report module and call its execute function
    report = frappe.get_doc("Report", report_name)
    if report.report_type != "Script Report":
        frappe.throw("Only Script Reports are supported.")

    module = importlib.import_module(
        f"freightmas.{frappe.scrub(report.module)}.report.{frappe.scrub(report.name)}.{frappe.scrub(report.name)}"
    )
    columns, data = module.execute(filters)

    # Format filter dates
    def format_date(val):
        if not val:
            return ""
        try:
            return frappe.utils.formatdate(val, "dd-MMM-yy")
        except Exception:
            return val

    formatted_filters = {}
    for k, v in filters.items():
        if "date" in k and v:
            formatted_filters[k] = format_date(v)
        else:
            formatted_filters[k] = v

    context = {
        "company": frappe.defaults.get_user_default("Company"),
        "title": _(report_name),
        "filters": formatted_filters,
        "columns": columns,
        "data": data,
        "frappe": frappe,
        "exported_at": frappe.utils.now_datetime().strftime("%d-%b-%Y %H:%M"),
    }

    html = frappe.render_template(
        "freightmas/templates/report_pdf_template.html", context
    )

    pdf = frappe.utils.pdf.get_pdf(
        html,
        options={
            "orientation": "Landscape",
            "footer-right": "Page [page] of [topage]",
            "footer-font-size": "10",
        }
    )

    frappe.local.response.filename = f"{report_name.replace(' ', '_')}.pdf"
    frappe.local.response.filecontent = pdf
    frappe.local.response.type = "download"


########################################################
## Freightmas Truck Trip Summary PDF Export
# This module exports the Truck Trip Summary report to PDF format.
@frappe.whitelist()
def export_truck_trip_summary_to_pdf(report_name, filters):
    import json
    filters = json.loads(filters)

    # Get report document first
    report = frappe.get_doc("Report", report_name)
    if report.report_type != "Script Report":
        frappe.throw("Only Script Reports are supported.")

    # Get report data
    module = importlib.import_module(
        f"freightmas.{frappe.scrub(report.module)}.report.{frappe.scrub(report.name)}.{frappe.scrub(report.name)}"
    )
    columns, data = module.execute(filters)

    # Helper function to parse currency values
    def parse_currency(value):
        if not value:
            return 0
        if isinstance(value, (int, float)):
            return float(value)
        clean_value = str(value).replace('$', '').replace(',', '').replace(' ', '')
        try:
            return float(clean_value)
        except ValueError:
            return 0

    # Get list of trucks that appear in the report
    used_trucks = list(set([row.get('truck') for row in data if row.get('truck')]))
    
    # Get available trucks that are not in the report
    available_trucks = frappe.get_all(
        "Truck",
        filters={
            "truck_status": "Available",
            "name": ["not in", used_trucks] if used_trucks else ["!=", ""]
        },
        fields=["name"],
        order_by="name"
    )

    # Format filters for display
    filters_list = []
    for key, value in filters.items():
        if value:
            label = key.replace('_', ' ').title()
            if 'date' in key.lower() and not isinstance(value, str):
                value = frappe.utils.formatdate(value)
            filters_list.append(f"{label}: {value}")
    
    filters_html = " | ".join(filters_list)

    # Group data by truck
    grouped_data = {}
    for row in data:
        truck = row.get('truck')
        if truck not in grouped_data:
            grouped_data[truck] = {
                'trips': [],
                'total': 0
            }
        grouped_data[truck]['trips'].append(row)
        grouped_data[truck]['total'] += parse_currency(row.get('estimated_revenue', 0))

    context = {
        "report_title": report_name,
        "filters_html": filters_html,  # Now filters_html is defined
        "trucks": grouped_data,
        "available_trucks": available_trucks,
        "frappe": frappe,
        "company": frappe.defaults.get_user_default("Company"),
        "exported_at": frappe.utils.now_datetime().strftime("%d-%b-%Y %H:%M")
    }

    html = frappe.render_template(
        "freightmas/templates/truck_trip_summary.html", context
    )

    pdf = frappe.utils.pdf.get_pdf(
        html,
        options={
            "orientation": "Landscape",
            "footer-right": "Page [page] of [topage]",
            "footer-font-size": "10",
        }
    )

    frappe.local.response.filename = f"{report_name.replace(' ', '_')}.pdf"
    frappe.local.response.filecontent = pdf
    frappe.local.response.type = "download"


########################################################
## Freightmas Truck Trip Summary Excel Export
# This module exports the Truck Trip Summary report to Excel format.
@frappe.whitelist()
def export_truck_trip_summary_to_excel(report_name, filters):
    import json
    from io import BytesIO
    
    filters = json.loads(filters)

    # Get report document first
    report = frappe.get_doc("Report", report_name)
    if report.report_type != "Script Report":
        frappe.throw("Only Script Reports are supported.")

    # Get report data
    module = importlib.import_module(
        f"freightmas.{frappe.scrub(report.module)}.report.{frappe.scrub(report.name)}.{frappe.scrub(report.name)}"
    )
    columns, data = module.execute(filters)

    # Helper function to parse currency values
    def parse_currency(value):
        if not value:
            return 0
        if isinstance(value, (int, float)):
            return float(value)
        clean_value = str(value).replace('$', '').replace(',', '').replace(' ', '')
        try:
            return float(clean_value)
        except ValueError:
            return 0

    # Pre-process and group data
    grouped_data = {}
    for row in data:
        row['estimated_revenue'] = parse_currency(row.get('estimated_revenue', 0))
        truck = row.get('truck')
        if truck not in grouped_data:
            grouped_data[truck] = {
                'trips': [],
                'total': 0
            }
        grouped_data[truck]['trips'].append(row)
        grouped_data[truck]['total'] += row['estimated_revenue']

    # Create workbook and styles
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trip Summary"

    # Styles
    header_font = Font(bold=True)
    truck_header_font = Font(bold=True, size=12)
    total_font = Font(bold=True)
    currency_format = '#,##0.00'
    date_format = 'DD-MMM-YY'
    
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column widths
    column_widths = {
        'Driver': 25,
        'Trip ID': 15,
        'Route': 25,
        'Customer': 20,
        'Revenue': 15,
        'Load': 12,
        'Offload': 12,
        'Days': 8,
        'Status': 12
    }

    # Start writing data
    current_row = 1

    # Report title
    ws.merge_cells(f'A{current_row}:I{current_row}')
    ws['A1'] = report_name
    ws['A1'].font = Font(bold=True, size=14)
    current_row += 1

    # Filters
    filters_text = []
    for key, value in filters.items():
        if value:
            label = key.replace('_', ' ').title()
            if 'date' in key.lower() and not isinstance(value, str):
                value = frappe.utils.formatdate(value)
            filters_text.append(f"{label}: {value}")
    
    ws.merge_cells(f'A{current_row}:I{current_row}')
    ws['A' + str(current_row)] = ' | '.join(filters_text)
    current_row += 2  # Add extra space after filters

    # Write data for each truck
    grand_total = 0
    total_trips = 0
    total_trucks = len(grouped_data)

    for truck, truck_data in grouped_data.items():
        # Truck header
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = f"Truck: {truck}"
        ws[f'A{current_row}'].font = truck_header_font
        ws[f'A{current_row}'].fill = header_fill
        current_row += 1

        # Column headers
        headers = ['Driver', 'Trip ID', 'Route', 'Customer', 'Revenue', 'Load', 'Offload', 'Days', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col)].width = column_widths[header]

        current_row += 1

        # Write trips
        for trip in truck_data['trips']:
            ws.cell(row=current_row, column=1, value=trip.get('driver'))
            ws.cell(row=current_row, column=2, value=trip.get('trip_id'))
            ws.cell(row=current_row, column=3, value=trip.get('route'))
            ws.cell(row=current_row, column=4, value=trip.get('customer'))
            
            revenue_cell = ws.cell(row=current_row, column=5, value=trip.get('estimated_revenue'))
            revenue_cell.number_format = currency_format
            revenue_cell.alignment = right_align

            load_cell = ws.cell(row=current_row, column=6, value=trip.get('date_loaded'))
            load_cell.alignment = center_align
            
            offload_cell = ws.cell(row=current_row, column=7, value=trip.get('date_offloaded'))
            offload_cell.alignment = center_align
            
            days_cell = ws.cell(row=current_row, column=8, value=trip.get('transit_days'))
            days_cell.alignment = center_align
            
            ws.cell(row=current_row, column=9, value=trip.get('workflow_state'))

            # Apply borders to all cells in the row
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).border = thin_border

            current_row += 1

        # Truck total
        total_text = f"Total for {truck} ({len(truck_data['trips'])} {'trip' if len(truck_data['trips']) == 1 else 'trips'})"
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = total_text
        ws[f'A{current_row}'].font = total_font
        
        total_cell = ws.cell(row=current_row, column=5, value=truck_data['total'])
        total_cell.number_format = currency_format
        total_cell.font = total_font
        total_cell.alignment = right_align

        # Apply borders to total row
        for col in range(1, 10):
            ws.cell(row=current_row, column=col).border = thin_border

        grand_total += truck_data['total']
        total_trips += len(truck_data['trips'])
        current_row += 2  # Add space between trucks

    # Grand total
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = f"Total Revenue ({total_trips} {'trip' if total_trips == 1 else 'trips'} | {total_trucks} {'truck' if total_trucks == 1 else 'trucks'})"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    
    grand_total_cell = ws.cell(row=current_row, column=5, value=grand_total)
    grand_total_cell.number_format = currency_format
    grand_total_cell.font = Font(bold=True, size=12)
    grand_total_cell.alignment = right_align

    # Apply borders to grand total row
    for col in range(1, 10):
        ws.cell(row=current_row, column=col).border = thin_border

    current_row += 2  # Add space after grand total

    # Add available trucks section
    used_trucks = list(set([row.get('truck') for row in data if row.get('truck')]))
    available_trucks = frappe.get_all(
        "Truck",
        filters={
            "truck_status": "Available",
            "name": ["not in", used_trucks] if used_trucks else ["!=", ""]
        },
        fields=["name"],
        order_by="name"
    )

    if available_trucks:
        ws.merge_cells(f'A{current_row}:I{current_row}')
        truck_list = ", ".join([t.name for t in available_trucks])
        ws[f'A{current_row}'] = f"Available trucks not contributing: {truck_list}"
        ws[f'A{current_row}'].font = Font(italic=True)
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = f"Total available but not contributing = {len(available_trucks)} {'truck' if len(available_trucks) == 1 else 'trucks'}"
        ws[f'A{current_row}'].font = Font(italic=True)

    # Save to BytesIO and return as file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.local.response.filename = f"{report_name.replace(' ', '_')}.xlsx"
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "binary"

